from __future__ import annotations

import importlib.util
import csv
import json
import os
import re
import shutil
import subprocess
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook


def app_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


APP_DIR = app_dir()
REFERENCE_DIR = APP_DIR / "reference_data"
SCRIPTS_DIR = APP_DIR / "scripts"
PROJECT_ROOT = Path(os.environ.get("DATONG_WORKSPACE", Path.home() / "Documents" / "大統工作助手"))
CODEX_HOME = Path(os.environ.get("CODEX_HOME", Path.home() / ".codex"))
OCR_SCRIPT = CODEX_HOME / "skills" / "extract-vendor-invoice-image" / "scripts" / "local_paddleocr_invoice_to_xlsx.py"
SYSTEM_PYTHON = Path.home() / "AppData" / "Local" / "Programs" / "Python" / "Python312" / "python.exe"
VENV_PYTHON = PROJECT_ROOT / ".venv-paddleocr" / "Scripts" / "python.exe"


CODE_RE = re.compile(r"^\d{6}$")
SUMMARY_ROW_NAMES = {"總價格", "總價", "總計", "合計", "小計", "稅金", "稅額", "折扣", "總數量", "合計數量"}


@dataclass
class OcrConfirmRow:
    excel_row: int
    is_existing: bool
    raw_name: str
    quantity: str
    unit_cost: str
    amount: str
    matched_code: str
    matched_name: str
    candidates: str
    status: str


@dataclass
class AdjustmentRow:
    row_id: str
    excel_row: int
    source_row: int
    product_code: str
    name: str
    category: str
    category_display: str
    quantity: str
    unit_cost: str
    amount: str
    status: str


@dataclass
class WorkflowState:
    image_path: Path | None = None
    output_dir: Path | None = None
    raw_xlsx: Path | None = None
    match_xlsx: Path | None = None
    suggested_names_txt: Path | None = None
    adjusted_xlsx: Path | None = None
    new_product_file: Path | None = None
    purchase_file: Path | None = None
    vendor: str = ""
    row_count: int = 0
    needs_ocr_review: bool = False
    ocr_issues: list[str] = field(default_factory=list)
    excluded_items: list[str] = field(default_factory=list)
    build_summary: dict[str, Any] = field(default_factory=dict)
    tmp_dir: Path | None = None


def python_exe(project_root: Path = PROJECT_ROOT) -> Path:
    venv_python = project_root / ".venv-paddleocr" / "Scripts" / "python.exe"
    if venv_python.exists():
        return venv_python
    return SYSTEM_PYTHON


def product_csv() -> Path | None:
    candidates = sorted(
        REFERENCE_DIR.glob("產品資料輸出*.CSV"),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    return candidates[0] if candidates else None


def assert_product_csv_current(path: Path) -> None:
    timezone = ZoneInfo("Asia/Taipei")
    stat = path.stat()
    created_at = datetime.fromtimestamp(stat.st_ctime, timezone)
    modified_at = datetime.fromtimestamp(stat.st_mtime, timezone)
    today = datetime.now(timezone).date()
    if created_at.date() != today and modified_at.date() != today:
        raise RuntimeError(
            "產品資料輸出.CSV 不是今天建立或修改的版本。\n"
            f"目前檔案：{path}\n"
            f"建立時間：{created_at:%Y-%m-%d %H:%M:%S}\n"
            f"修改時間：{modified_at:%Y-%m-%d %H:%M:%S}\n"
            "請先把今天的產品資料輸出.CSV 放進 reference_data。"
        )


def read_csv_dicts(path: Path) -> list[dict[str, str]]:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "cp950", "big5", "utf-8"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                return [{k: (v or "").strip() for k, v in row.items()} for row in csv.DictReader(handle)]
        except UnicodeDecodeError as error:
            last_error = error
    raise RuntimeError(f"無法讀取 CSV：{path} ({last_error})")


def category_name_map() -> dict[str, str]:
    path = REFERENCE_DIR / "大類清單.csv"
    if not path.exists():
        return {}
    result: dict[str, str] = {}
    for row in read_csv_dicts(path):
        code = str(row.get("大類代號", "")).strip()
        name = str(row.get("大類名稱", "")).strip()
        if code and name:
            result[code] = name
    return result


def category_code(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    first = text.split(maxsplit=1)[0].strip()
    return first if first else text


def category_display(value: str, names: dict[str, str] | None = None) -> str:
    code = category_code(value)
    if not code:
        return ""
    lookup = names if names is not None else category_name_map()
    name = lookup.get(code)
    return f"{code} {name}" if name else f"{code} 未知大類"


def validate_runtime(project_root: Path = PROJECT_ROOT) -> list[str]:
    issues: list[str] = []
    required = [
        python_exe(project_root),
        OCR_SCRIPT,
        SCRIPTS_DIR / "match-existing-products.py",
        SCRIPTS_DIR / "review-invoice-product-check.py",
        SCRIPTS_DIR / "fill-import-templates.ps1",
        REFERENCE_DIR / "產品比對身份關鍵詞.csv",
        REFERENCE_DIR / "品牌括號命名規則.csv",
        REFERENCE_DIR / "大類清單.csv",
        REFERENCE_DIR / "廠商代號.xlsx",
        REFERENCE_DIR / "建檔用.xls",
        REFERENCE_DIR / "採購單匯入範例.xls",
    ]
    for path in required:
        if not path.exists():
            issues.append(f"找不到必要檔案：{path}")
    if python_exe(project_root).exists():
        completed = subprocess.run(
            [str(python_exe(project_root)), "-X", "utf8", "-c", "import paddleocr, paddle"],
            text=True,
            encoding="utf-8",
            errors="replace",
            capture_output=True,
            check=False,
        )
        if completed.returncode != 0:
            detail = "\n".join(part for part in [completed.stdout, completed.stderr] if part).strip()
            issues.append(
                "系統 OCR 尚未安裝完成，請確認全系統 Python 已安裝 PaddleOCR。\n"
                f"Python：{python_exe(project_root)}\n{detail}"
            )
    csv_path = product_csv()
    if csv_path is None:
        issues.append(f"找不到產品資料輸出.CSV：{REFERENCE_DIR}")
    return issues


def parse_json_summary(output: str) -> dict[str, Any]:
    matches = list(re.finditer(r"\{[\s\S]*?\}", output))
    for match in reversed(matches):
        try:
            data = json.loads(match.group(0))
        except json.JSONDecodeError:
            continue
        if isinstance(data, dict):
            return data
    raise RuntimeError("程式已結束，但無法讀取結果摘要。")


def run_command(command: list[str], cwd: Path, env: dict[str, str] | None = None) -> subprocess.CompletedProcess[str]:
    merged_env = os.environ.copy()
    merged_env["PYTHONUTF8"] = "1"
    if env:
        merged_env.update(env)
    return subprocess.run(
        command,
        cwd=str(cwd),
        env=merged_env,
        text=True,
        encoding="utf-8",
        errors="replace",
        capture_output=True,
        check=False,
    )


def run_ocr(state: WorkflowState, project_root: Path = PROJECT_ROOT) -> dict[str, Any]:
    if state.image_path is None or state.output_dir is None:
        raise RuntimeError("尚未選擇圖片或輸出資料夾。")
    state.tmp_dir = project_root / ".codex-tmp" / "invoice-ocr-gui" / state.image_path.stem
    command = [
        str(python_exe(project_root)),
        "-X",
        "utf8",
        str(OCR_SCRIPT),
        str(state.image_path),
        "--output-dir",
        str(state.output_dir),
        "--tmp-dir",
        str(state.tmp_dir),
        "--settings",
        str(project_root / "參考資料" / "OCR設定.json"),
    ]
    completed = run_command(
        command,
        cwd=project_root,
        env={"PADDLE_PDX_ENABLE_MKLDNN_BYDEFAULT": "0"},
    )
    output = "\n".join(part for part in [completed.stdout, completed.stderr] if part)
    if completed.returncode != 0:
        raise RuntimeError(output or f"OCR 程式結束代碼：{completed.returncode}")
    summary = parse_json_summary(output)
    state.raw_xlsx = Path(str(summary.get("output", "")))
    state.vendor = str(summary.get("vendor", ""))
    state.row_count = int(summary.get("row_count", 0) or 0)
    state.needs_ocr_review = bool(summary.get("needs_review"))
    state.ocr_issues = list(summary.get("issues", []) or [])
    return summary


def run_match(state: WorkflowState, project_root: Path = PROJECT_ROOT) -> dict[str, Any]:
    if state.raw_xlsx is None or not state.raw_xlsx.exists():
        raise RuntimeError("找不到 OCR 原始 Excel，無法產品比對。")
    csv_path = product_csv()
    if csv_path is None:
        raise RuntimeError(f"找不到產品資料輸出.CSV：{REFERENCE_DIR}")
    assert_product_csv_current(csv_path)
    command = [
        str(python_exe(project_root)),
        "-X",
        "utf8",
        str(SCRIPTS_DIR / "match-existing-products.py"),
        "--input-xlsx",
        str(state.raw_xlsx),
        "--csv",
        str(csv_path),
    ]
    completed = run_command(command, cwd=project_root)
    output = "\n".join(part for part in [completed.stdout, completed.stderr] if part)
    if completed.returncode != 0:
        raise RuntimeError(output or f"產品比對結束代碼：{completed.returncode}")
    state.match_xlsx = state.raw_xlsx.with_name(f"{state.raw_xlsx.stem}_產品比對檢查.xlsx")
    state.suggested_names_txt = state.match_xlsx.with_name(f"{state.match_xlsx.stem}_建議名稱.txt")
    if not state.match_xlsx.exists():
        raise RuntimeError(f"產品比對完成但找不到檢查檔：{state.match_xlsx}")
    return {"message": output.strip(), "output": str(state.match_xlsx)}


def load_review_module() -> Any:
    path = SCRIPTS_DIR / "review-invoice-product-check.py"
    spec = importlib.util.spec_from_file_location("review_invoice_product_check_local", path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"無法載入覆核腳本：{path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def find_header(ws: Any, required: list[str]) -> tuple[int, dict[str, int]]:
    for row in range(1, min(ws.max_row, 20) + 1):
        headers: dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row, col).value
            if value is not None:
                headers[str(value).strip()] = col
        if all(name in headers for name in required):
            return row, headers
    raise RuntimeError("找不到必要欄位：" + "、".join(required))


def ensure_column(ws: Any, header_row: int, headers: dict[str, int], name: str) -> int:
    if name in headers:
        return headers[name]
    column = ws.max_column + 1
    ws.cell(header_row, column).value = name
    headers[name] = column
    return column


def is_summary_row_name(value: str) -> bool:
    text = str(value or "").strip()
    return not text or text in SUMMARY_ROW_NAMES or any(text.startswith(name) for name in SUMMARY_ROW_NAMES)


def first_candidate(candidates: str) -> tuple[str, str]:
    first_line = next((line.strip() for line in str(candidates or "").splitlines() if line.strip()), "")
    match = re.match(r"^(\d{6})\s+(.+?)(?:\s+\([0-9.]+\))?$", first_line)
    if not match:
        return "", ""
    return match.group(1), match.group(2).strip()


def load_ocr_confirm_rows(state: WorkflowState) -> list[OcrConfirmRow]:
    if state.match_xlsx is None or not state.match_xlsx.exists():
        raise RuntimeError("找不到產品比對檢查檔。")
    wb = load_workbook(state.match_xlsx)
    ws = wb.active
    header_row, headers = find_header(ws, ["產品代號", "品名", "數量", "進價", "金額"])
    rows: list[OcrConfirmRow] = []
    for row in range(header_row + 1, ws.max_row + 1):
        raw_name = str(ws.cell(row, headers["品名"]).value or "").strip()
        product_code = str(ws.cell(row, headers["產品代號"]).value or "").strip()
        if is_summary_row_name(raw_name) or product_code in SUMMARY_ROW_NAMES:
            continue
        status = str(ws.cell(row, headers.get("比對狀態", 0)).value or "").strip() if "比對狀態" in headers else ""
        matched_code = str(ws.cell(row, headers.get("已建檔代號", 0)).value or "").strip() if "已建檔代號" in headers else ""
        matched_name = str(ws.cell(row, headers.get("已建檔品名", 0)).value or "").strip() if "已建檔品名" in headers else ""
        candidates = str(ws.cell(row, headers.get("相似候選", 0)).value or "").strip() if "相似候選" in headers else ""
        rows.append(
            OcrConfirmRow(
                excel_row=row,
                is_existing=status == "已建檔" and bool(CODE_RE.match(matched_code or product_code)),
                raw_name=raw_name,
                quantity=str(ws.cell(row, headers["數量"]).value or "").strip(),
                unit_cost=str(ws.cell(row, headers["進價"]).value or "").strip(),
                amount=str(ws.cell(row, headers["金額"]).value or "").strip(),
                matched_code=matched_code or product_code,
                matched_name=matched_name,
                candidates=candidates,
                status=status,
            )
        )
    return rows


def save_ocr_confirm_rows(state: WorkflowState, rows: list[OcrConfirmRow]) -> None:
    if state.match_xlsx is None or not state.match_xlsx.exists():
        raise RuntimeError("找不到產品比對檢查檔。")
    wb = load_workbook(state.match_xlsx)
    ws = wb.active
    header_row, headers = find_header(ws, ["產品代號", "品名", "數量", "進價", "金額"])
    status_col = ensure_column(ws, header_row, headers, "比對狀態")
    matched_code_col = ensure_column(ws, header_row, headers, "已建檔代號")
    matched_name_col = ensure_column(ws, header_row, headers, "已建檔品名")
    ensure_column(ws, header_row, headers, "相似候選")

    for item in rows:
        row = item.excel_row
        if row < header_row + 1 or row > ws.max_row:
            continue
        matched_code = str(item.matched_code).strip()
        matched_name = str(item.matched_name).strip()
        if item.is_existing and (not CODE_RE.match(matched_code) or not matched_name):
            raise RuntimeError(f"第 {row} 列已勾選已建檔，但缺少六位產品代號或已建檔品名。")

        ws.cell(row, headers["品名"]).value = str(item.raw_name).strip()
        ws.cell(row, headers["數量"]).value = item.quantity
        ws.cell(row, headers["進價"]).value = item.unit_cost
        ws.cell(row, headers["金額"]).value = item.amount
        if item.is_existing:
            ws.cell(row, headers["產品代號"]).value = matched_code
            ws.cell(row, headers["產品代號"]).number_format = "@"
            ws.cell(row, status_col).value = "已建檔"
            ws.cell(row, matched_code_col).value = matched_code
            ws.cell(row, matched_name_col).value = matched_name
        else:
            ws.cell(row, headers["產品代號"]).value = ""
            ws.cell(row, headers["產品代號"]).number_format = "@"
            if str(ws.cell(row, status_col).value or "").strip() == "已建檔":
                ws.cell(row, status_col).value = "有類似產品" if str(item.candidates).strip() else "確認為新品"
            ws.cell(row, matched_code_col).value = ""
            ws.cell(row, matched_name_col).value = ""
    wb.save(state.match_xlsx)


def prepare_review_table(state: WorkflowState) -> list[AdjustmentRow]:
    if state.match_xlsx is None or not state.match_xlsx.exists():
        raise RuntimeError("找不到產品比對檢查檔。")
    csv_path = product_csv()
    if csv_path is None:
        raise RuntimeError(f"找不到產品資料輸出.CSV：{REFERENCE_DIR}")
    review = load_review_module()
    wb = load_workbook(state.match_xlsx)
    ws = wb.active
    header_row, headers = review.find_header_row(ws)
    rows = review.item_rows(ws, header_row, headers)
    excluded_rows = [row for row in rows if review.is_excluded_item_row(ws, row, headers)]
    rows = [row for row in rows if row not in excluded_rows]
    review.fill_existing_product_codes(ws, rows, headers)

    catalog_rows = review.read_catalog(csv_path)
    catalog_names = review.catalog_product_names(catalog_rows)
    brand_rules_path = REFERENCE_DIR / "品牌括號命名規則.csv"
    brand_rule_rows = review.load_brand_rule_rows(brand_rules_path)
    brand_rules = review.load_brand_rules(brand_rules_path)
    category_rules_path = REFERENCE_DIR / "大類清單.csv"
    category_rules = review.load_category_rules(category_rules_path)
    name_col = headers["品名"]
    category_col = headers["大類"]
    code_col = headers["產品代號"]

    for row in rows:
        code_cell = ws.cell(row, code_col)
        code_text = "" if code_cell.value is None else str(code_cell.value).strip()
        code_cell.value = code_text
        code_cell.number_format = "@"

        raw_name = str(ws.cell(row, name_col).value or "").strip()
        matched_name_col = headers.get("已建檔品名")
        if review.is_existing_row(ws, row, headers) and matched_name_col:
            matched_name = str(ws.cell(row, matched_name_col).value or "").strip()
            adjusted_name = matched_name or raw_name
        elif review.is_existing_row(ws, row, headers):
            adjusted_name = raw_name
        elif review.CODE_RE.match(code_text) and review.PREFIX_RE.match(raw_name):
            adjusted_name = raw_name
        else:
            adjusted_name = review.adjust_name(raw_name, catalog_names, brand_rules, brand_rule_rows)
        ws.cell(row, name_col).value = adjusted_name

        if not review.is_existing_row(ws, row, headers):
            category_cell = ws.cell(row, category_col)
            if category_cell.value not in (None, ""):
                category_value = review.infer_category_value(category_cell.value, adjusted_name, catalog_rows, category_rules)
            else:
                category_value = review.infer_category_from_brand_rules(adjusted_name, brand_rule_rows)
                if category_value in (None, ""):
                    category_value = review.infer_category_value(category_cell.value, adjusted_name, catalog_rows, category_rules)
            if category_value not in (None, ""):
                category_cell.value = category_value

    wb.save(state.match_xlsx)
    state.excluded_items = [str(ws.cell(row, name_col).value or "").strip() for row in excluded_rows]
    return load_adjustment_rows(state)


def load_adjustment_rows(state: WorkflowState) -> list[AdjustmentRow]:
    if state.match_xlsx is None:
        return []
    wb = load_workbook(state.match_xlsx)
    ws = wb.active
    header_row, headers = find_header(ws, ["產品代號", "品名", "大類", "數量", "進價", "金額"])
    names = category_name_map()
    rows: list[AdjustmentRow] = []
    for row in range(header_row + 1, ws.max_row + 1):
        name = str(ws.cell(row, headers["品名"]).value or "").strip()
        code = str(ws.cell(row, headers["產品代號"]).value or "").strip()
        if not name or name in {"總價格", "總價", "總計", "合計", "小計"} or code in {"總價格", "總價", "總計", "合計", "小計"}:
            continue
        text_for_exclusion = "\n".join(
            str(ws.cell(row, headers[h]).value or "") for h in ("品名", "已建檔品名", "相似候選") if h in headers
        )
        if any(keyword in text_for_exclusion for keyword in ("一番賞", "抽賞", "Ichiban Kuji", "ICHIBAN KUJI", "遮蔽", "已遮蔽", "人工確認重複", "重複品項")):
            continue
        status = str(ws.cell(row, headers.get("比對狀態", 0)).value or "").strip() if "比對狀態" in headers else ""
        if status == "已建檔":
            continue
        category = str(ws.cell(row, headers["大類"]).value or "").strip()
        rows.append(
            AdjustmentRow(
                row_id=str(row),
                excel_row=row,
                source_row=row,
                product_code=code,
                name=name,
                category=category_code(category),
                category_display=category_display(category, names),
                quantity=str(ws.cell(row, headers["數量"]).value or "").strip(),
                unit_cost=str(ws.cell(row, headers["進價"]).value or "").strip(),
                amount=str(ws.cell(row, headers["金額"]).value or "").strip(),
                status=status,
            )
        )
    return rows


def validate_split_totals(ws: Any, headers: dict[str, int], rows: list[AdjustmentRow]) -> None:
    grouped: dict[int, list[AdjustmentRow]] = {}
    for item in rows:
        if item.source_row and item.source_row != item.excel_row:
            grouped.setdefault(item.source_row, []).append(item)
    for source_row, split_rows in grouped.items():
        original_quantity = str(ws.cell(source_row, headers["數量"]).value or "").strip()
        original_amount = str(ws.cell(source_row, headers["金額"]).value or "").strip()
        source_current = next((item for item in rows if item.excel_row == source_row), None)
        group = ([source_current] if source_current else []) + split_rows
        try:
            original_quantity_value = float(original_quantity)
            split_quantity_value = sum(float(str(item.quantity or "0")) for item in group)
        except ValueError:
            original_quantity_value = split_quantity_value = 0
        try:
            original_amount_value = float(original_amount)
            split_amount_value = sum(float(str(item.amount or "0")) for item in group)
        except ValueError:
            original_amount_value = split_amount_value = 0
        if abs(original_quantity_value - split_quantity_value) > 0.0001 or abs(original_amount_value - split_amount_value) > 0.0001:
            name = str(ws.cell(source_row, headers["品名"]).value or "").strip()
            raise RuntimeError(
                f"第 {source_row} 列「{name}」已拆分，但拆分後數量或金額合計不一致。\n"
                f"原數量/金額：{original_quantity} / {original_amount}\n"
                f"拆分後數量/金額：{split_quantity_value:g} / {split_amount_value:g}"
            )


def save_adjustment_rows(state: WorkflowState, rows: list[AdjustmentRow]) -> None:
    if state.match_xlsx is None or not state.match_xlsx.exists():
        raise RuntimeError("找不到產品比對檢查檔。")
    wb = load_workbook(state.match_xlsx)
    ws = wb.active
    _header_row, headers = find_header(ws, ["產品代號", "品名", "大類", "數量", "進價", "金額"])
    validate_split_totals(ws, headers, rows)
    for item in rows:
        row = item.excel_row
        if row <= 0 or row > ws.max_row:
            row = ws.max_row + 1
            item.excel_row = row
        ws.cell(row, headers["產品代號"]).value = str(item.product_code).strip()
        ws.cell(row, headers["產品代號"]).number_format = "@"
        ws.cell(row, headers["品名"]).value = str(item.name).strip()
        ws.cell(row, headers["大類"]).value = category_code(item.category)
        ws.cell(row, headers["數量"]).value = item.quantity
        ws.cell(row, headers["進價"]).value = item.unit_cost
        ws.cell(row, headers["金額"]).value = item.amount
        if "比對狀態" in headers:
            ws.cell(row, headers["比對狀態"]).value = item.status or "確認為新品"
    wb.save(state.match_xlsx)


def generate_adjusted_xlsx(state: WorkflowState, project_root: Path = PROJECT_ROOT) -> dict[str, Any]:
    if state.match_xlsx is None or not state.match_xlsx.exists():
        raise RuntimeError("找不到產品比對檢查檔。")
    csv_path = product_csv()
    if csv_path is None:
        raise RuntimeError(f"找不到產品資料輸出.CSV：{REFERENCE_DIR}")
    command = [
        str(python_exe(project_root)),
        "-X",
        "utf8",
        str(SCRIPTS_DIR / "review-invoice-product-check.py"),
        "--input-xlsx",
        str(state.match_xlsx),
        "--csv",
        str(csv_path),
        "--brand-rules",
        str(REFERENCE_DIR / "品牌括號命名規則.csv"),
        "--category-rules",
        str(REFERENCE_DIR / "大類清單.csv"),
        "--confirmed",
    ]
    completed = run_command(command, cwd=project_root)
    output = "\n".join(part for part in [completed.stdout, completed.stderr] if part)
    if completed.returncode != 0:
        raise RuntimeError(output or f"覆核調整結束代碼：{completed.returncode}")
    state.adjusted_xlsx = state.match_xlsx.with_name(f"{state.match_xlsx.stem}[調整]{state.match_xlsx.suffix}")
    if not state.adjusted_xlsx.exists():
        raise RuntimeError(f"覆核調整完成但找不到調整檔：{state.adjusted_xlsx}")
    return {"message": output.strip(), "output": str(state.adjusted_xlsx)}


def build_import_files(state: WorkflowState, project_root: Path = PROJECT_ROOT) -> dict[str, Any]:
    if state.adjusted_xlsx is None or not state.adjusted_xlsx.exists():
        raise RuntimeError("找不到 [調整].xlsx。")
    output_dir = project_root / "建檔進貨用"
    command = [
        "powershell",
        "-NoProfile",
        "-ExecutionPolicy",
        "Bypass",
        "-File",
        str(SCRIPTS_DIR / "fill-import-templates.ps1"),
        "-WorkspaceRoot",
        str(APP_DIR),
        "-ProductsXlsx",
        str(state.adjusted_xlsx),
        "-OutputDir",
        str(output_dir),
        "-NewProductTemplate",
        str(REFERENCE_DIR / "建檔用.xls"),
        "-PurchaseTemplate",
        str(REFERENCE_DIR / "採購單匯入範例.xls"),
        "-ConfirmedReviewed",
    ]
    completed = run_command(command, cwd=project_root)
    output = "\n".join(part for part in [completed.stdout, completed.stderr] if part)
    if completed.returncode != 0:
        raise RuntimeError(output or f"正式輸出結束代碼：{completed.returncode}")
    summary = parse_json_summary(output)
    new_file = summary.get("newProductFile")
    purchase_file = summary.get("purchaseImportFile")
    state.new_product_file = Path(new_file) if new_file else None
    state.purchase_file = Path(str(purchase_file)) if purchase_file else None
    state.build_summary = summary
    if state.new_product_file and not state.new_product_file.exists():
        raise RuntimeError(f"找不到建檔用成品：{state.new_product_file}")
    if state.purchase_file is None or not state.purchase_file.exists():
        raise RuntimeError(f"找不到採購單成品：{state.purchase_file}")
    return summary


def cleanup_intermediate_files(state: WorkflowState) -> list[Path]:
    if not ((state.new_product_file and state.new_product_file.exists()) or (state.purchase_file and state.purchase_file.exists())):
        raise RuntimeError("正式成品不存在，不清理中間檔。")
    candidates: list[Path] = []
    for path in [state.raw_xlsx, state.match_xlsx, state.suggested_names_txt, state.adjusted_xlsx]:
        if path is not None:
            candidates.append(path)
    if state.match_xlsx is not None:
        candidates.extend(state.match_xlsx.parent.glob(f"{state.match_xlsx.stem}*.txt"))
    deleted: list[Path] = []
    for path in candidates:
        try:
            if path.exists() and path.is_file():
                path.unlink()
                deleted.append(path)
        except OSError:
            pass
    if state.tmp_dir and state.tmp_dir.exists():
        try:
            shutil.rmtree(state.tmp_dir)
            deleted.append(state.tmp_dir)
        except OSError:
            pass
    return deleted
