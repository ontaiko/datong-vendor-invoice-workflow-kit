#!/usr/bin/env python3
from __future__ import annotations

import importlib.util
from pathlib import Path

from openpyxl import Workbook


SCRIPT = Path(__file__).with_name("match-existing-products.py")
SPEC = importlib.util.spec_from_file_location("match_existing_products", SCRIPT)
if SPEC is None or SPEC.loader is None:
    raise RuntimeError(f"無法載入：{SCRIPT}")
MODULE = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(MODULE)


def main() -> None:
    wb = Workbook()
    ws = wb.active
    headers = ["產品代號", "品名", "數量", "進價", "金額"]
    ws.append(headers)
    ws.append([14796, "測試商品", 1, 100, 100])
    columns = {header: index for index, header in enumerate(headers, start=1)}
    MODULE.write_match_columns(
        ws,
        1,
        columns,
        [2],
        [
            {
                "matchStatus": "exact",
                "matchedProductCode": "014796",
                "matchedProductName": "(吊飾)內心慌張-海綿寶寶卡通毛絨系列路人魚掛件",
                "similarCandidates": [],
            }
        ],
    )

    assert ws.cell(2, columns["產品代號"]).value == "014796"
    assert ws.cell(2, columns["產品代號"]).number_format == "@"
    assert ws.cell(2, columns["已建檔代號"]).value == "014796"
    assert ws.cell(2, columns["已建檔代號"]).number_format == "@"
    print("match-product-catalog regression tests passed: 4")


if __name__ == "__main__":
    main()
