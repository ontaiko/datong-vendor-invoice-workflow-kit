#!/usr/bin/env python3
import argparse
import csv
import re
import unicodedata
from copy import copy
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

try:
    from rapidfuzz import fuzz
except ImportError:
    fuzz = None


NON_IDENTITY_PARENS = re.compile(r"\((?:全?\d+種|再販|單)\)", re.IGNORECASE)
LEADING_CATEGORY = re.compile(r"^\([^)]*\)")
IGNORED_CHARS = re.compile(r"[\s\-_/\\・:：,，.。+＋]+")
EXCEL_TEXT_CODE = re.compile(r'^="([^"]*)"$')
SUMMARY_ROW_NAMES = {"總價格", "總價", "總計", "合計", "小計", "稅金", "稅額", "折扣", "總數量", "合計數量", "頁碼", "頁次"}
SUMMARY_ROW_PREFIXES = tuple(SUMMARY_ROW_NAMES)


def is_summary_text(value):
    text = str(value or "").strip()
    if not text:
        return False
    normalized = text.replace("：", "").replace(":", "").strip()
    return normalized in SUMMARY_ROW_NAMES or any(normalized.startswith(name) for name in SUMMARY_ROW_PREFIXES)


def suggested_name_from_invoice(name):
    text = str(name or "").strip()
    text = re.sub(r"\*\s*\d+\s*$", "", text).strip()
    return f"(){text}"
STATUS_LABELS = {
    "exact": "已建檔",
    "similar": "有類似產品",
    "new": "確認為新品",
}
IDENTITY_TOKEN_CSV = "產品比對身份關鍵詞.csv"
WARNING_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
PRODUCT_CODE_NAME_WARNING_THRESHOLD = 0.35


def normalize_name(value):
    text = unicodedata.normalize("NFKC", str(value or "")).strip()
    text = LEADING_CATEGORY.sub("", text)
    text = NON_IDENTITY_PARENS.sub("", text)
    return IGNORED_CHARS.sub("", text).casefold()


def normalize_product_code(value):
    text = str(value or "").strip()
    match = EXCEL_TEXT_CODE.match(text)
    if match:
        text = match.group(1)
    return text.strip()


def bigrams(text):
    if len(text) < 2:
        return {text} if text else set()
    return {text[index : index + 2] for index in range(len(text) - 1)}


def similarity(left, right):
    if not left or not right:
        return 0.0
    left_bigrams = bigrams(left)
    right_bigrams = bigrams(right)
    dice = (
        (2 * len(left_bigrams & right_bigrams)) / (len(left_bigrams) + len(right_bigrams))
        if left_bigrams and right_bigrams
        else 0.0
    )
    sequence = SequenceMatcher(None, left, right).ratio()
    containment = min(len(left), len(right)) / max(len(left), len(right)) if left in right or right in left else 0.0
    return max(dice, sequence, containment)


def extract_identity_tokens(value, version_tokens):
    text = unicodedata.normalize("NFKC", str(value or ""))
    return {token for token in version_tokens if token in text}


def extract_character_tokens(value, character_tokens):
    text = unicodedata.normalize("NFKC", str(value or ""))
    return {token for token in character_tokens if token in text}


def product_similarity(source_name, product_name, identity_tokens):
    source_normalized = normalize_name(source_name)
    product_normalized = normalize_name(product_name)
    if fuzz is not None:
        score = fuzz.WRatio(source_normalized, product_normalized) / 100
    else:
        score = similarity(source_normalized, product_normalized)
    source_tokens = extract_identity_tokens(source_name, identity_tokens["version"])
    product_tokens = extract_identity_tokens(product_name, identity_tokens["version"])
    source_characters = extract_character_tokens(source_name, identity_tokens["character"])
    product_characters = extract_character_tokens(product_name, identity_tokens["character"])
    if source_tokens and product_tokens and source_tokens != product_tokens:
        score *= 0.6
    elif source_tokens and not product_tokens:
        score *= 0.75
    if source_characters and product_characters:
        if source_characters == product_characters:
            score *= 1.15
        else:
            score *= 0.5
    elif source_characters and not product_characters:
        score *= 0.85
    return score


def load_identity_tokens(reference_dir):
    path = reference_dir / IDENTITY_TOKEN_CSV
    if not path.exists():
        raise SystemExit(f"找不到產品比對身份關鍵詞清單：{path}。請在參考資料資料夾建立此 CSV。")

    tokens = {"version": set(), "character": set()}
    for row in read_csv_rows(path):
        token_type = str(row.get("類型", "")).strip()
        token = str(row.get("關鍵詞", "")).strip()
        enabled = str(row.get("啟用", "Y")).strip().upper()
        if not token or enabled in {"N", "NO", "0", "FALSE"}:
            continue
        if "版本" in token_type:
            tokens["version"].add(token)
        elif "角色" in token_type or "機體" in token_type:
            tokens["character"].add(token)
    return tokens


def read_csv_rows(path):
    last_error = None
    for encoding in ("utf-8-sig", "cp950", "utf-8"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                return list(csv.DictReader(handle))
        except UnicodeDecodeError as error:
            last_error = error
    raise last_error


def assert_current_reference_csv(path):
    timezone = ZoneInfo("Asia/Taipei")
    stat = path.stat()
    created_at = datetime.fromtimestamp(stat.st_ctime, timezone)
    modified_at = datetime.fromtimestamp(stat.st_mtime, timezone)
    today = datetime.now(timezone).date()
    if created_at.date() != today and modified_at.date() != today:
        raise SystemExit(
            f"產品資料參考檔不是今天建立或修改：{path.name}，"
            f"建立時間為 {created_at:%Y-%m-%d %H:%M:%S}，"
            f"修改時間為 {modified_at:%Y-%m-%d %H:%M:%S}（台北時區）。"
            "請更新今天建立或修改的產品資料 CSV 後再繼續。"
        )


def display_name(item):
    return item.get("suggestedName") or item.get("name") or item.get("rawName") or ""


def find_header_row(ws, required_headers):
    for row_index in range(1, min(ws.max_row, 20) + 1):
        values = [str(ws.cell(row_index, column).value or "").strip() for column in range(1, ws.max_column + 1)]
        if all(header in values for header in required_headers):
            return row_index, {header: values.index(header) + 1 for header in values if header}
    raise SystemExit(f"找不到必要欄位：{', '.join(required_headers)}。請確認進貨 xlsx 表頭。")


def read_invoice_xlsx(path):
    if "_產品比對檢查" in path.stem:
        raise SystemExit(f"輸入檔看起來已經是產品比對檢查檔：{path.name}。請改用原始進貨試算表。")
    wb = load_workbook(path)
    ws = wb.active
    header_row, columns = find_header_row(ws, ["產品代號", "品名", "數量", "進價", "金額"])
    items = []
    row_numbers = []
    for row_index in range(header_row + 1, ws.max_row + 1):
        leading_values = {
            str(ws.cell(row_index, columns[header]).value or "").strip()
            for header in ("產品代號", "品名")
            if header in columns
        }
        if any(is_summary_text(value) for value in leading_values):
            continue
        name = str(ws.cell(row_index, columns["品名"]).value or "").strip()
        if not name:
            continue
        if is_summary_text(name):
            continue
        item = {
            "productCode": normalize_product_code(ws.cell(row_index, columns["產品代號"]).value),
            "rawName": name,
            "name": name,
            "quantity": ws.cell(row_index, columns["數量"]).value,
            "unitCost": ws.cell(row_index, columns["進價"]).value,
            "amount": ws.cell(row_index, columns["金額"]).value,
        }
        if "零售價" in columns:
            item["retailPrice"] = ws.cell(row_index, columns["零售價"]).value
        items.append(item)
        row_numbers.append(row_index)
    return wb, ws, header_row, columns, items, row_numbers


def write_match_columns(ws, header_row, columns, row_numbers, matched):
    next_column = ws.max_column + 1
    output_headers = [
        "比對狀態",
        "已建檔代號",
        "已建檔品名",
        "相似候選",
    ]
    for offset, header in enumerate(output_headers):
        column = columns.get(header, next_column + offset)
        ws.cell(header_row, column).value = header
        if header not in columns:
            columns[header] = column
        source = ws.cell(header_row, 1)
        target = ws.cell(header_row, column)
        if source.has_style:
            target._style = copy(source._style)
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.alignment = copy(source.alignment)
        if header == "相似候選":
            target.alignment = Alignment(wrap_text=True, vertical="top")

    for row_index, result in zip(row_numbers, matched):
        if result.get("matchStatus") == "exact" and result.get("matchedProductCode"):
            product_code_cell = ws.cell(row_index, columns["產品代號"])
            if not normalize_product_code(product_code_cell.value):
                product_code_cell.value = str(result["matchedProductCode"]).zfill(6)
                product_code_cell.number_format = "@"

        candidates = "\n".join(
            f"{candidate['productCode']} {candidate['productName']} ({candidate['score']})"
            for candidate in result.get("similarCandidates", [])
        )
        values = {
            "比對狀態": format_status(result),
            "已建檔代號": result.get("matchedProductCode", ""),
            "已建檔品名": result.get("matchedProductName", ""),
            "相似候選": candidates,
        }
        for header, value in values.items():
            cell = ws.cell(row_index, columns[header])
            cell.value = value
            if header == "相似候選":
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        if result.get("productCodeNameWarning"):
            warning = result["productCodeNameWarning"]
            for header in ("比對狀態", "已建檔品名"):
                cell = ws.cell(row_index, columns[header])
                cell.fill = copy(WARNING_FILL)
                cell.comment = Comment(warning, "Codex")
    autofit_columns(ws)


def format_status(result):
    status = STATUS_LABELS.get(result.get("matchStatus", ""), result.get("matchStatus", ""))
    return status


def display_width(value):
    width = 0
    for char in str(value or ""):
        width += 2 if unicodedata.east_asian_width(char) in {"F", "W", "A"} else 1
    return width


def autofit_columns(ws):
    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_width = max(display_width(cell.value) for cell in column_cells)
        ws.column_dimensions[column_letter].width = min(max(max_width + 2, 8), 60)



def default_suggested_names_txt(output_xlsx):
    return output_xlsx.with_name(f"{output_xlsx.stem}_建議名稱.txt")


def write_suggested_names_txt(output_xlsx, items, matched):
    pending = [
        item
        for item, result in zip(items, matched)
        if result.get("matchStatus") in {"similar", "new"}
    ]
    suggested_path = default_suggested_names_txt(output_xlsx)
    if not pending:
        if suggested_path.exists():
            suggested_path.unlink()
        return None
    lines = [suggested_name_from_invoice(item.get("name", "")) for item in pending]
    suggested_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return suggested_path

def default_output_xlsx(input_xlsx):
    return input_xlsx.with_name(f"{input_xlsx.stem}_產品比對檢查.xlsx")


def compare_item(item, products, products_by_code, identity_tokens, threshold, max_candidates):
    source_name = display_name(item)
    normalized = normalize_name(source_name)
    exact_matches = [product for product in products if normalized and product["normalizedName"] == normalized]
    entered_product_code = normalize_product_code(item.get("productCode"))

    result = dict(item)
    result.update(
        {
            "matchStatus": "new",
            "matchedProductCode": "",
            "matchedProductName": "",
            "similarCandidates": [],
            "matchNote": "產品資料 CSV 未找到合理候選。",
            "existingProduct": False,
            "enteredProductCode": entered_product_code,
            "productCodeCheck": "blank" if not entered_product_code else "not_found",
        }
    )

    if entered_product_code and entered_product_code in products_by_code:
        product = products_by_code[entered_product_code]
        code_name_score = product_similarity(source_name, product["productName"], identity_tokens)
        product_code_name_warning = ""
        if code_name_score < PRODUCT_CODE_NAME_WARNING_THRESHOLD:
            product_code_name_warning = (
                "已填產品代號存在，但進貨品名與該代號正式品名差異過大；"
                f"品名相似度 {code_name_score:.4f}，請覆核是否填錯代號。"
            )
        result.update(
            {
                "matchStatus": "exact",
                "matchedProductCode": product["productCode"],
                "matchedProductName": product["productName"],
                "matchNote": product_code_name_warning or "試算表已填產品代號，且今天產品資料 CSV 找到該代號。",
                "existingProduct": True,
                "productCode": product["productCode"],
                "name": product["productName"],
                "productCodeCheck": "found",
                "productCodeNameScore": round(code_name_score, 4),
                "productCodeNameWarning": product_code_name_warning,
            }
        )
        return result

    if len(exact_matches) == 1:
        product = exact_matches[0]
        result.update(
            {
                "matchStatus": "exact",
                "matchedProductCode": product["productCode"],
                "matchedProductName": product["productName"],
                "matchNote": "正規化名稱完全一致，自動沿用既有產品代號。",
                "existingProduct": True,
                "productCodeCheck": result["productCodeCheck"],
            }
        )
        if not result.get("productCode"):
            result["productCode"] = product["productCode"]
        if not result.get("name"):
            result["name"] = product["productName"]
        return result

    scored = []
    for product in products:
        score = product_similarity(source_name, product["productName"], identity_tokens)
        if score >= threshold:
            scored.append(
                {
                    "productCode": product["productCode"],
                    "productName": product["productName"],
                    "score": round(score, 4),
                }
            )

    scored.sort(key=lambda candidate: (-candidate["score"], candidate["productCode"]))
    result["similarCandidates"] = scored[:max_candidates]
    if scored:
        result["matchStatus"] = "similar"
        if len(exact_matches) > 1:
            result["matchNote"] = "有多筆正規化名稱一致的既有商品，需人工確認產品代號。"
        else:
            result["matchNote"] = "找到相似候選，需人工確認是否為同一商品。"
    if entered_product_code and result["productCodeCheck"] == "not_found":
        result["matchNote"] = f"試算表產品代號 {entered_product_code} 不在今天產品資料 CSV；{result['matchNote']}"
    return result


def main():
    parser = argparse.ArgumentParser(description="Match invoice xlsx rows against an exported product CSV.")
    parser.add_argument("--input-xlsx", required=True, type=Path)
    parser.add_argument("--csv", required=True, type=Path)
    parser.add_argument("--output-xlsx", type=Path)
    parser.add_argument("--similar-threshold", type=float, default=0.6)
    parser.add_argument("--max-candidates", type=int, default=5)
    parser.add_argument(
        "--no-suggestion-txt",
        action="store_true",
        help="完整進貨流程使用：只輸出產品比對檢查 xlsx，不建立建議名稱 txt。",
    )
    args = parser.parse_args()

    assert_current_reference_csv(args.csv)
    wb, ws, header_row, columns, items, row_numbers = read_invoice_xlsx(args.input_xlsx)
    rows = read_csv_rows(args.csv)
    identity_tokens = load_identity_tokens(args.csv.parent)
    products = []
    products_by_code = {}
    for row in rows:
        code = normalize_product_code(row.get("1.產品代號", ""))
        name = str(row.get("2.產品名稱", "")).strip()
        normalized = normalize_name(name)
        if code and name and normalized:
            product = {"productCode": code, "productName": name, "normalizedName": normalized}
            products.append(product)
            products_by_code.setdefault(code, product)

    matched = [
        compare_item(item, products, products_by_code, identity_tokens, args.similar_threshold, args.max_candidates)
        for item in items
    ]
    write_match_columns(ws, header_row, columns, row_numbers, matched)
    output_xlsx = args.output_xlsx or default_output_xlsx(args.input_xlsx)
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)
    if args.no_suggestion_txt:
        stale_path = default_suggested_names_txt(output_xlsx)
        if stale_path.exists():
            stale_path.unlink()
        print("完整流程模式：未建立建議名稱檔。")
    else:
        suggested_path = write_suggested_names_txt(output_xlsx, items, matched)
        if suggested_path:
            print(f"已建立建議名稱檔：{suggested_path}")
        else:
            print("全部商品皆已建檔，未建立建議名稱檔。")


if __name__ == "__main__":
    main()
