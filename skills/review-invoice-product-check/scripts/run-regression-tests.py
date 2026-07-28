#!/usr/bin/env python3
from __future__ import annotations

import importlib.util
from pathlib import Path

from openpyxl import Workbook


SCRIPT = Path(__file__).with_name("review-invoice-product-check.py")
SPEC = importlib.util.spec_from_file_location("review_invoice_product_check", SCRIPT)
if SPEC is None or SPEC.loader is None:
    raise RuntimeError(f"無法載入：{SCRIPT}")
MODULE = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(MODULE)


def rule(keywords: str, label: str, category: str, sample_raw: str = "", sample_adjusted: str = ""):
    return {
        "關鍵字": keywords,
        "括號名稱": label,
        "大類": category,
        "位置規則": "括號分類在最前方",
        "命名規則": "測試",
        "範例原始名稱": sample_raw,
        "範例調整後名稱": sample_adjusted,
        "啟用": "1",
        "說明": "測試",
    }


def test_existing_code_text() -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["產品代號", "品名", "比對狀態", "已建檔代號"])
    ws.append([14796, "測試商品", "已建檔", 14796])
    headers = {"產品代號": 1, "品名": 2, "比對狀態": 3, "已建檔代號": 4}
    MODULE.fill_existing_product_codes(ws, [2], headers)
    assert ws["A2"].value == "014796"
    assert ws["A2"].number_format == "@"
    assert ws["D2"].value == "014796"
    assert ws["D2"].number_format == "@"


def test_pingu_specificity() -> None:
    raw = "Pingu搞怪時刻米粒公仔12入/半箱4"
    expected = "(盲盒)Pingu搞怪時刻米粒公仔@12"
    rows = [
        rule("Pingu", "廣泛", "36"),
        rule("Pingu搞怪時刻米粒公仔", "盲盒", "34", raw, expected),
    ]
    assert MODULE.adjust_name_by_brand_rule(raw, rows) == expected
    assert MODULE.infer_category_from_brand_rules(expected, rows) == "34"


def test_chikawa_specificity() -> None:
    raw = "Chikawa三麗鷗草莓車車毛絨掛件-四款可選"
    rows = [
        rule("Chikawa", "廣泛", "36"),
        rule("Chikawa三麗鷗草莓車車毛絨掛件", "吊飾", "31"),
    ]
    adjusted = MODULE.adjust_name_by_brand_rule(raw, rows)
    assert adjusted is not None and adjusted.startswith("(吊飾)")
    assert MODULE.infer_category_from_brand_rules(adjusted, rows) == "31"


def main() -> None:
    test_existing_code_text()
    test_pingu_specificity()
    test_chikawa_specificity()
    print("review-invoice-product-check regression tests passed: 10")


if __name__ == "__main__":
    main()
