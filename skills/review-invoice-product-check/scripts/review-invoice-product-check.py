#!/usr/bin/env python3
"""Review invoice product-check workbooks and write an adjusted copy."""

from __future__ import annotations

import argparse
import csv
import re
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


CODE_RE = re.compile(r"^\d{6}$")
FORMAL_NAME_RE = re.compile(r"^\([^)]*\).+-.+")
PREFIX_RE = re.compile(r"^\(([^)]*)\)")
WORK_HINTS = [
    "航海王",
    "海賊王",
    "火影忍者",
    "福音戰士",
    "新世紀福音戰士",
    "蝙蝠俠",
    "七龍珠",
    "寶可夢",
    "鋼彈",
    "咒術迴戰",
    "鬼滅之刃",
    "我的英雄學院",
    "排球少年",
]
CLASS_HINTS = [
    ("景品", ("景品", "KING OF ARTIST", "VIBRATION STARS", "Grandista", "MAXIMATIC")),
    ("盲盒", ("盲盒", "盒玩", "中盒", "全", "一中盒")),
    ("模型", ("模型", "組裝模型", "Figure-rise", "Figure rise", "RG", "HG", "MG")),
]
BRAND_TO_CLASS = ("布魯可", "布魯克")
DROP_OUTPUT_COLUMNS = {"建檔代號", "已建檔代號", "已建檔品名", "相似候選"}
BRAND_RULES_FILENAME = "品牌括號命名規則.csv"
CATEGORY_LIST_FILENAME = "大類清單.csv"
SUMMARY_ROW_NAMES = {"總價格", "總價", "總計", "合計", "小計"}
EXCLUDE_ITEM_KEYWORDS = (
    "一番賞",
    "抽賞",
    "Ichiban Kuji",
    "ICHIBAN KUJI",
    "遮蔽",
    "已遮蔽",
    "人工確認重複",
    "重複品項",
)


def die(message: str, code: int = 1) -> None:
    print(message, file=sys.stderr)
    raise SystemExit(code)


def read_catalog(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        die(f"找不到產品資料 CSV：{path}")
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "cp950", "big5"):
        try:
            with path.open("r", encoding=encoding, newline="") as f:
                reader = csv.DictReader(f)
                rows = [{k: (v or "").strip() for k, v in row.items()} for row in reader]
            return rows
        except UnicodeDecodeError as exc:
            last_error = exc
    die(f"無法讀取產品資料 CSV 編碼：{path} ({last_error})")


def read_csv_dicts(path: Path) -> list[dict[str, str]]:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "cp950", "big5"):
        try:
            with path.open("r", encoding=encoding, newline="") as f:
                return [{k: (v or "").strip() for k, v in row.items()} for row in csv.DictReader(f)]
        except UnicodeDecodeError as exc:
            last_error = exc
    die(f"無法讀取 CSV 編碼：{path} ({last_error})")


def split_rule_values(value: str) -> list[str]:
    return [part.strip() for part in re.split(r"[;；]", value or "") if part.strip()]


def find_default_catalog(input_path: Path) -> Path | None:
    candidates: list[Path] = []
    for base in [input_path.parent, *input_path.parents, Path.cwd()]:
        ref_dir = base / "參考資料"
        if ref_dir.exists():
            candidates.extend(ref_dir.glob("產品資料輸出*.CSV"))
            candidates.extend(ref_dir.glob("產品資料輸出*.csv"))
    unique = {candidate.resolve(): candidate for candidate in candidates if candidate.exists()}
    if not unique:
        return None
    return max(unique.values(), key=lambda path: path.stat().st_mtime)


def find_default_brand_rules(input_path: Path, catalog_path: Path) -> Path | None:
    candidates: list[Path] = []
    for base in [input_path.parent, *input_path.parents, catalog_path.parent, *catalog_path.parents]:
        candidates.append(base / "參考資料" / BRAND_RULES_FILENAME)
        candidates.append(base / BRAND_RULES_FILENAME)
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def find_default_category_list(input_path: Path, catalog_path: Path) -> Path | None:
    candidates: list[Path] = []
    skill_root = Path(__file__).resolve().parents[1]
    candidates.append(skill_root / "references" / CATEGORY_LIST_FILENAME)
    for base in [input_path.parent, *input_path.parents, catalog_path.parent, *catalog_path.parents]:
        candidates.append(base / "參考資料" / CATEGORY_LIST_FILENAME)
        candidates.append(base / CATEGORY_LIST_FILENAME)
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def load_brand_rules(path: Path | None) -> tuple[str, ...]:
    brands: list[str] = []
    for row in load_brand_rule_rows(path):
        keywords = split_rule_values(row.get("關鍵字", ""))
        label = row.get("括號名稱", "").strip()
        for keyword in keywords:
            if keyword and keyword not in brands:
                brands.append(keyword)
        if label and label not in brands:
            brands.append(label)
    return tuple(brands) if brands else BRAND_TO_CLASS


def load_brand_rule_rows(path: Path | None) -> list[dict[str, str]]:
    if path is None or not path.exists():
        return []

    rules: list[dict[str, str]] = []
    for row in read_csv_dicts(path):
        enabled = row.get("啟用", "1").strip()
        if enabled in {"0", "否", "false", "False", "FALSE"}:
            continue
        keywords = split_rule_values(row.get("關鍵字", ""))
        label = row.get("括號名稱", "").strip()
        if not keywords and not label:
            continue
        rules.append(row)
    return rules


def load_category_rules(path: Path | None) -> list[dict[str, str]]:
    if path is None or not path.exists():
        return []

    rules: list[dict[str, str]] = []
    for row in read_csv_dicts(path):
        enabled = row.get("啟用", "1").strip()
        if enabled in {"0", "否", "false", "False", "FALSE"}:
            continue
        code = row.get("大類代號", "").strip()
        name = row.get("大類名稱", "").strip()
        if not code or not name:
            continue
        rules.append(
            {
                "code": code,
                "name": name,
                "label": row.get("括號分類", "").strip(),
                "keywords": row.get("關鍵字", "").strip(),
                "note": row.get("備註", "").strip(),
            }
        )
    return rules


def category_label(adjusted_name: str) -> str:
    match = PREFIX_RE.match(adjusted_name.strip())
    return match.group(1).strip() if match else ""


def infer_category_from_rules(adjusted_name: str, category_rules: list[dict[str, str]]) -> str | None:
    label = category_label(adjusted_name)
    for rule in category_rules:
        labels = {rule["name"], rule["label"]}
        keywords = {part.strip() for part in rule["keywords"].split(";") if part.strip()}
        if label and label in labels:
            return rule["code"]
        if label and label in keywords:
            return rule["code"]

    for rule in category_rules:
        keywords = {part.strip() for part in rule["keywords"].split(";") if part.strip()}
        if any(keyword and keyword in adjusted_name for keyword in keywords):
            return rule["code"]
    return None


def print_category_reference(category_rules: list[dict[str, str]]) -> None:
    if not category_rules:
        return
    print("")
    print("可參考大類清單：")
    for rule in category_rules:
        print(f"- {rule['code']}｜{rule['name']}")


def category_name_map(category_rules: list[dict[str, str]]) -> dict[str, str]:
    return {str(rule["code"]).strip(): rule["name"] for rule in category_rules if rule.get("code")}


def display_category_value(category, category_names: dict[str, str]) -> str:
    value = str(category or "").strip()
    if not value:
        return ""
    name = category_names.get(value)
    return f"{value} {name}" if name else value


def find_header_row(ws) -> tuple[int, dict[str, int]]:
    required = {"產品代號", "品名"}
    for row in range(1, min(ws.max_row, 20) + 1):
        headers: dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row, col).value
            if value is not None:
                headers[str(value).strip()] = col
        if required.intersection(headers):
            missing = required - set(headers)
            if missing:
                die("試算表缺少必要欄位：" + "、".join(sorted(missing)))
            if "大類" not in headers:
                category_col = ws.max_column + 1
                ws.cell(row, category_col).value = "大類"
                headers["大類"] = category_col
            return row, headers
    die("找不到含有 產品代號、品名、大類 的表頭列。")


def item_rows(ws, header_row: int, headers: dict[str, int]) -> list[int]:
    code_col = headers["產品代號"]
    name_col = headers["品名"]
    rows: list[int] = []
    for row in range(header_row + 1, ws.max_row + 1):
        code_text = str(ws.cell(row, code_col).value or "").strip()
        if code_text in SUMMARY_ROW_NAMES:
            continue
        name = ws.cell(row, name_col).value
        if name is None:
            continue
        name_text = str(name).strip()
        if not name_text or name_text in SUMMARY_ROW_NAMES:
            continue
        rows.append(row)
    return rows


def row_text_for_exclusion(ws, row: int, headers: dict[str, int]) -> str:
    texts: list[str] = []
    for header in ("品名", "已建檔品名", "相似候選"):
        col = headers.get(header)
        if col:
            texts.append(str(ws.cell(row, col).value or ""))
    return "\n".join(texts)


def is_excluded_item_row(ws, row: int, headers: dict[str, int]) -> bool:
    text = row_text_for_exclusion(ws, row, headers)
    return any(keyword in text for keyword in EXCLUDE_ITEM_KEYWORDS)


def print_excluded_items(ws, rows: list[int], headers: dict[str, int]) -> None:
    if not rows:
        return
    name_col = headers["品名"]
    print("已排除項目：")
    for row in rows:
        name = str(ws.cell(row, name_col).value or "").strip()
        print(f"- {name}")
    print("")


def validate_codes(ws, rows: list[int], headers: dict[str, int]) -> list[tuple[int, str, str]]:
    code_col = headers["產品代號"]
    name_col = headers["品名"]
    problems: list[tuple[int, str, str]] = []
    for row in rows:
        raw = ws.cell(row, code_col).value
        code = "" if raw is None else str(raw).strip()
        if not CODE_RE.match(code):
            name = str(ws.cell(row, name_col).value or "").strip()
            problems.append((row, name, code))
    return problems


def row_status(ws, row: int, headers: dict[str, int]) -> str:
    status_col = headers.get("比對狀態")
    if not status_col:
        return ""
    return str(ws.cell(row, status_col).value or "").strip()


def is_existing_row(ws, row: int, headers: dict[str, int]) -> bool:
    return row_status(ws, row, headers) == "已建檔"


def fill_existing_product_codes(ws, rows: list[int], headers: dict[str, int]) -> None:
    code_col = headers["產品代號"]
    matched_code_col = headers.get("已建檔代號")
    if not matched_code_col:
        return
    for row in rows:
        if not is_existing_row(ws, row, headers):
            continue
        code_cell = ws.cell(row, code_col)
        matched_code = str(ws.cell(row, matched_code_col).value or "").strip()
        if (code_cell.value is None or str(code_cell.value).strip() == "") and CODE_RE.match(matched_code):
            code_cell.value = matched_code
            code_cell.number_format = "@"


def split_words(text: str) -> list[str]:
    cleaned = re.sub(r"[()（）\[\]【】,，/／｜|+＋:：]", " ", text)
    return [part.strip() for part in cleaned.split() if part.strip()]


def catalog_product_names(rows: list[dict[str, str]]) -> list[str]:
    names: list[str] = []
    for row in rows:
        name = row.get("2.產品名稱", "").strip()
        if name:
            names.append(name)
    return names


def infer_class_label(raw_name: str, catalog_names: list[str], brand_rules: tuple[str, ...]) -> str:
    for brand in brand_rules:
        if brand in raw_name:
            return brand

    upper = raw_name.upper()
    for label, hints in CLASS_HINTS:
        for hint in hints:
            if hint.upper() in upper:
                return label
    return ""


def infer_work(words: list[str], catalog_names: list[str]) -> tuple[str, int | None]:
    joined = "".join(words)
    for hint in WORK_HINTS:
        if hint in joined:
            for index, word in enumerate(words):
                if hint in word or word in hint:
                    return hint, index
            return hint, None

    scores: Counter[str] = Counter()
    for word in words:
        if len(word) < 2:
            continue
        for catalog_name in catalog_names[:3000]:
            if word in catalog_name:
                scores[word] += 1
    if scores:
        work = scores.most_common(1)[0][0]
        return work, words.index(work) if work in words else None
    return "", None


def normalize_formal_brand_name(raw_name: str, brand_rules: tuple[str, ...]) -> str | None:
    match = re.match(r"^\(([^)]*)\)(.+)-(.+)$", raw_name)
    if not match:
        return None

    current_label, subject, tail = match.groups()
    found_brand = next((brand for brand in brand_rules if brand in raw_name), "")
    if not found_brand:
        return raw_name

    subject = subject.replace(found_brand, "").strip()
    tail = tail.replace(found_brand, "").strip()
    label = found_brand if current_label in {"", found_brand} else current_label
    if not subject or not tail:
        return f"({label}){subject}{tail}"
    return f"({label}){subject}-{tail}"


def box_count_suffix(raw_name: str) -> str:
    text = raw_name.replace("＋", "+")
    match = re.search(r"全(\d+)種\+隱(\d+)種", text)
    if match:
        return f"@{int(match.group(1)) + int(match.group(2))}"
    match = re.search(r"全(\d+)\+1隱(\d+)種", text)
    if match:
        return f"@{match.group(2)}"
    match = re.search(r"全(\d+)種", text)
    if match:
        return f"@{match.group(1)}"
    match = re.search(r"(\d+)入", text)
    if match:
        return f"@{match.group(1)}"
    return ""


def remove_count_notes(raw_name: str) -> str:
    text = re.sub(r"\(全[^)]*種[^)]*\)", "", raw_name)
    text = re.sub(r"\(再\)|\(再販\)", "", text)
    text = re.sub(r"[/／]\s*(?:半箱|箱|袋|優)\s*\d+", "", text)
    text = re.sub(r"(?:半箱|箱|袋|優)\s*\d+\s*$", "", text)
    text = re.sub(r"\*\s*\d+\s*$", "", text)
    text = re.sub(r"[/／]\s*$", "", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def adjust_name_by_known_pattern(raw_name: str) -> str | None:
    raw_name = raw_name.strip()
    count_suffix = box_count_suffix(raw_name)
    cleaned = remove_count_notes(raw_name)

    if cleaned.startswith("PalVerse 盒玩 "):
        subject = cleaned.replace("PalVerse 盒玩 ", "", 1).strip()
        return f"(盲盒){subject}-PalVerse{count_suffix}"

    match = re.match(r"^(葬送的芙莉蓮)\s+(.+?)\s+BOX版", cleaned, re.IGNORECASE)
    if match:
        return f"(盲盒){match.group(2).strip()}-{match.group(1)}BOX版{count_suffix}"

    match = re.match(r"^(勝利女神妮姬)\s+(.+?RICH BOX版\s*vol\.\d+)", cleaned, re.IGNORECASE)
    if match:
        return f"(盲盒){match.group(2).replace(' ', '')}-{match.group(1)}{count_suffix}"

    match = re.match(r"^(魔物獵人)\s+(.+?Plus)\s+(Vol\.\d+)", cleaned, re.IGNORECASE)
    if match:
        return f"(盲盒){match.group(3)}-{match.group(1)}{match.group(2).replace(' ', '')}{count_suffix}"

    match = re.match(r"^HG《ONE PIECE》(.+)$", cleaned, re.IGNORECASE)
    if match:
        return f"(公仔){match.group(1).strip()}-ONE PIECE HG"

    match = re.match(r"^(航海王)\s+KING OF ARTIST\s+(.+)$", cleaned, re.IGNORECASE)
    if match:
        return f"(景品){match.group(2).strip()}-{match.group(1)}KING OF ARTIST"

    match = re.match(r"^(航海王)\s+BATTLE RECORD COLLECTION\s+(.+)$", cleaned, re.IGNORECASE)
    if match:
        return f"(景品){match.group(2).strip()}-{match.group(1)}BATTLE RECORD COLLECTION"

    match = re.match(
        r"^(.+?)\s+(Relax time|MAXIMATIC|Grandista|Coreful Figure|Desktop Cute(?: Figure)?|KING OF ARTIST|VIBRATION STARS)\s+(.+)$",
        cleaned,
        re.IGNORECASE,
    )
    if match:
        work = match.group(1).strip()
        series = match.group(2).strip()
        subject = match.group(3).strip()
        return f"(景品){subject}-{work}{series}"

    match = re.match(r"^Pokepeace\s+絨毛玩偶\s+(.+)$", cleaned, re.IGNORECASE)
    if match:
        return f"(娃娃){match.group(1).strip()}-Pokepeace絨毛玩偶"

    match = re.match(r"^(戀上換裝娃娃)\s+Coreful Figure\s+(.+)$", cleaned, re.IGNORECASE)
    if match:
        return f"(景品){match.group(2).replace(' ', '')}-{match.group(1)}Coreful Figure"

    match = re.match(r"^(宇崎學妹想要玩!)\s+Desktop Cute Figure\s+(.+)$", cleaned, re.IGNORECASE)
    if match:
        return f"(景品){match.group(2).replace(' ', '')}-{match.group(1)}Desktop Cute Figure"

    match = re.match(r"^(魔女之旅)\s+Coreful Figure\s+(.+)$", cleaned, re.IGNORECASE)
    if match:
        return f"(景品){match.group(2).replace(' ', '')}-{match.group(1)}Coreful Figure"

    match = re.match(r"^(出包王女Darkness)\s+Desktop Cute\s+(.+)$", cleaned, re.IGNORECASE)
    if match:
        return f"(景品){match.group(2).replace(' ', '')}-{match.group(1)}Desktop Cute"

    match = re.match(r"^(櫻未來)\s+Newly Written Figure\s+(.+)$", cleaned, re.IGNORECASE)
    if match:
        return f"(景品){match.group(1)}{match.group(2).replace(' ', '')}-Newly Written Figure"

    return None


def clean_rule_subject(raw_name: str, keyword: str) -> str:
    subject = remove_count_notes(raw_name)
    subject = re.sub(r"\(\d+入\)|（\d+入）|\d+入", "", subject)
    subject = subject.replace(keyword, "")
    subject = re.sub(r"\b盒玩\b|盒玩", "", subject)
    subject = re.sub(r"[-－–—_]+", " ", subject)
    subject = re.sub(r"\s+", "", subject)
    return subject.strip()


def brand_rule_keywords(row: dict[str, str]) -> list[str]:
    return split_rule_values(row.get("關鍵字", ""))


def matched_brand_keyword(raw_name: str, row: dict[str, str]) -> str:
    for keyword in brand_rule_keywords(row):
        if keyword and keyword in raw_name:
            return keyword
    return ""


def sample_adjusted_name(raw_name: str, row: dict[str, str]) -> str | None:
    sample_raw = row.get("範例原始名稱", "").strip()
    sample_adjusted = row.get("範例調整後名稱", "").strip()
    if sample_raw and sample_adjusted and raw_name == sample_raw:
        return sample_adjusted
    return None


def adjust_name_by_brand_rule(raw_name: str, brand_rule_rows: list[dict[str, str]]) -> str | None:
    raw_name = raw_name.strip()
    for row in brand_rule_rows:
        example_name = sample_adjusted_name(raw_name, row)
        if example_name is not None:
            return example_name

        keyword = matched_brand_keyword(raw_name, row)
        label = row.get("括號名稱", "").strip()
        position_rule = row.get("位置規則", "").strip()
        if not keyword or not label:
            continue

        if "細項在前-品牌在後" in position_rule:
            subject = clean_rule_subject(raw_name, keyword)
            count_suffix = box_count_suffix(raw_name)
            if subject:
                return f"({label}){subject}-{keyword}{count_suffix}"

        if "括號分類在最前方" in position_rule:
            subject = raw_name.replace(keyword, "").strip() if keyword == label else raw_name.strip()
            subject = re.sub(r"\s+", " ", subject).strip()
            if subject:
                return f"({label}){subject}"
            return f"({label}){raw_name}"
    return None


def infer_category_from_brand_rules(adjusted_name: str, brand_rule_rows: list[dict[str, str]]) -> str | None:
    label_match = PREFIX_RE.match(adjusted_name)
    label = label_match.group(1) if label_match else ""
    for row in brand_rule_rows:
        category = row.get("大類", "").strip()
        if not category:
            continue
        row_label = row.get("括號名稱", "").strip()
        keywords = brand_rule_keywords(row)
        if (row_label and row_label == label) or any(keyword and keyword in adjusted_name for keyword in keywords):
            return category
    return None


def adjust_name(
    raw_name: str,
    catalog_names: list[str],
    brand_rules: tuple[str, ...],
    brand_rule_rows: list[dict[str, str]],
) -> str:
    raw_name = raw_name.strip()
    formal_name = normalize_formal_brand_name(raw_name, brand_rules)
    if formal_name is not None and FORMAL_NAME_RE.match(formal_name):
        return formal_name

    brand_rule_name = adjust_name_by_brand_rule(raw_name, brand_rule_rows)
    if brand_rule_name is not None:
        return brand_rule_name

    known_pattern_name = adjust_name_by_known_pattern(raw_name)
    if known_pattern_name is not None:
        return known_pattern_name

    normalized_name = remove_count_notes(raw_name)
    words = split_words(normalized_name)
    if not words:
        return normalized_name or raw_name

    class_label = infer_class_label(normalized_name, catalog_names, brand_rules)
    work, work_index = infer_work(words, catalog_names)

    brand = words[0] if len(words) >= 3 else ""
    if brand in brand_rules:
        brand = ""
    body_words = words[1:] if brand else words[:]
    body_words = [word for word in body_words if word not in brand_rules]

    if work_index is not None:
        work_in_body_index = max(work_index - (1 if brand else 0), 0)
        if words[0] in brand_rules and work_index > 0:
            work_in_body_index = work_index - 1
        before_work = body_words[:work_in_body_index]
        after_work = body_words[work_in_body_index + 1 :]
    else:
        before_work = []
        after_work = body_words[1:] if len(body_words) > 1 else []
        work = body_words[0] if body_words else ""

    if len(after_work) == 2 and after_work[0].endswith("版"):
        subject_words = after_work[-1:]
    else:
        subject_words = after_work[-2:] if len(after_work) >= 2 else after_work
    series_words = before_work + after_work[: max(len(after_work) - len(subject_words), 0)]

    subject = "".join(subject_words).strip()
    tail_parts = [work, brand, *series_words]
    tail = "".join(part for part in tail_parts if part).strip()

    if not subject or not tail:
        return f"({class_label}){normalized_name}"
    return f"({class_label}){subject}-{tail}"


def infer_category_value(
    original_value,
    adjusted_name: str,
    catalog_rows: list[dict[str, str]],
    category_rules: list[dict[str, str]],
):
    if original_value not in (None, ""):
        return original_value

    rule_value = infer_category_from_rules(adjusted_name, category_rules)
    if rule_value not in (None, ""):
        return rule_value

    category_headers = [h for h in (catalog_rows[0].keys() if catalog_rows else []) if h in {"大類", "大類代號"}]
    if not category_headers:
        return None

    product_name_col = "2.產品名稱"
    scores: Counter[str] = Counter()
    adjusted_words = split_words(adjusted_name)
    for row in catalog_rows:
        product_name = row.get(product_name_col, "")
        category = row.get(category_headers[0], "").strip()
        if not product_name or not category:
            continue
        score = sum(1 for word in adjusted_words if word and word in product_name)
        if score >= 2:
            scores[category] += score
    return scores.most_common(1)[0][0] if scores else None


def delete_output_columns(ws, header_row: int) -> None:
    drop_cols: list[int] = []
    for col in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col).value
        if value is not None and str(value).strip() in DROP_OUTPUT_COLUMNS:
            drop_cols.append(col)
    for col in sorted(drop_cols, reverse=True):
        ws.delete_cols(col)


def polish_worksheet(ws, header_row: int) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = ws.cell(header_row + 1, 1).coordinate
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"

    for cell in ws[header_row]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border

    for col in range(1, ws.max_column + 1):
        max_len = 8
        letter = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row, col).value
            if value is None:
                continue
            max_len = max(max_len, min(len(str(value)), 60))
        ws.column_dimensions[letter].width = max_len + 2


def print_confirmation_table(
    ws,
    rows: list[int],
    headers: dict[str, int],
    category_rules: list[dict[str, str]],
) -> None:
    code_col = headers["產品代號"]
    name_col = headers["品名"]
    category_col = headers["大類"]
    category_names = category_name_map(category_rules)

    print("請確認以下內容是否正確：")
    print("")
    print("| 產品代號 | 產品名稱 | 大類 |")
    print("| --- | --- | --- |")
    for row in rows:
        code = str(ws.cell(row, code_col).value or "").strip()
        name = str(ws.cell(row, name_col).value or "").strip()
        category = display_category_value(ws.cell(row, category_col).value, category_names)
        print(f"| {code} | {name} | {category} |")
    print("")
    print("若名稱與大類正確，請直接貼上產品代號；若要修改，也可一起貼上產品名稱或大類。")


def print_code_problems(problems: list[tuple[int, str, str]]) -> None:
    print("產品代號缺漏或格式錯誤，請先修正後再產生調整檔：")
    for row, name, code in problems:
        shown = "(空白)" if code == "" else code
        print(f"- 第 {row} 列｜{name}｜目前值：{shown}")


def output_path_for(input_path: Path) -> Path:
    return input_path.with_name(f"{input_path.stem}[調整]{input_path.suffix}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-xlsx", required=True)
    parser.add_argument("--csv")
    parser.add_argument("--output-xlsx")
    parser.add_argument("--brand-rules")
    parser.add_argument("--category-rules")
    parser.add_argument("--check-only", action="store_true")
    parser.add_argument("--confirmed", action="store_true")
    args = parser.parse_args()

    input_path = Path(args.input_xlsx)
    csv_path = Path(args.csv) if args.csv else find_default_catalog(input_path)
    if csv_path is None:
        die("找不到產品資料 CSV。請放在工作區 參考資料/產品資料輸出*.CSV，或使用 --csv 指定。")
    output_path = Path(args.output_xlsx) if args.output_xlsx else output_path_for(input_path)

    wb = load_workbook(input_path)
    ws = wb.active
    header_row, headers = find_header_row(ws)
    rows = item_rows(ws, header_row, headers)
    excluded_rows = [row for row in rows if is_excluded_item_row(ws, row, headers)]
    rows = [row for row in rows if row not in excluded_rows]

    fill_existing_product_codes(ws, rows, headers)

    if args.check_only:
        problems = validate_codes(ws, rows, headers)
        if problems:
            print_code_problems(problems)
            raise SystemExit(2)
        print(f"產品代號檢查通過：{len(rows)} 筆。")
        return

    catalog_rows = read_catalog(csv_path)
    catalog_names = catalog_product_names(catalog_rows)
    brand_rules_path = Path(args.brand_rules) if args.brand_rules else find_default_brand_rules(input_path, csv_path)
    brand_rule_rows = load_brand_rule_rows(brand_rules_path)
    brand_rules = load_brand_rules(brand_rules_path)
    category_rules_path = Path(args.category_rules) if args.category_rules else find_default_category_list(input_path, csv_path)
    category_rules = load_category_rules(category_rules_path)
    name_col = headers["品名"]
    category_col = headers["大類"]
    code_col = headers["產品代號"]
    category_problems: list[tuple[int, str]] = []

    for row in rows:
        code_cell = ws.cell(row, code_col)
        code_text = "" if code_cell.value is None else str(code_cell.value).strip()
        code_cell.value = code_text
        code_cell.number_format = "@"

        raw_name = str(ws.cell(row, name_col).value or "").strip()
        matched_name_col = headers.get("已建檔品名")
        if is_existing_row(ws, row, headers) and matched_name_col:
            matched_name = str(ws.cell(row, matched_name_col).value or "").strip()
            adjusted_name = matched_name or raw_name
        elif is_existing_row(ws, row, headers):
            adjusted_name = raw_name
        elif CODE_RE.match(code_text) and PREFIX_RE.match(raw_name):
            adjusted_name = raw_name
        else:
            adjusted_name = adjust_name(raw_name, catalog_names, brand_rules, brand_rule_rows)
        ws.cell(row, name_col).value = adjusted_name

        category_cell = ws.cell(row, category_col)
        if not is_existing_row(ws, row, headers):
            if category_cell.value not in (None, ""):
                category_value = infer_category_value(category_cell.value, adjusted_name, catalog_rows, category_rules)
            else:
                category_value = infer_category_from_brand_rules(adjusted_name, brand_rule_rows)
                if category_value in (None, ""):
                    category_value = infer_category_value(category_cell.value, adjusted_name, catalog_rows, category_rules)
            if category_value not in (None, ""):
                category_cell.value = category_value

    if not args.confirmed:
        print_excluded_items(ws, excluded_rows, headers)
        if not rows:
            print("排除後沒有需要覆核或輸出的商品。")
            return
        print_confirmation_table(ws, rows, headers, category_rules)
        has_blank_category = any(
            not is_existing_row(ws, row, headers)
            and not str(ws.cell(row, category_col).value or "").strip()
            for row in rows
        )
        if has_blank_category:
            print_category_reference(category_rules)
        return

    problems = validate_codes(ws, rows, headers)
    if problems:
        print_code_problems(problems)
        raise SystemExit(2)

    print(f"產品代號檢查通過：{len(rows)} 筆。")

    for row in rows:
        if is_existing_row(ws, row, headers):
            continue
        category = str(ws.cell(row, category_col).value or "").strip()
        if not category:
            name = str(ws.cell(row, name_col).value or "").strip()
            category_problems.append((row, name))

    if category_problems:
        print("以下商品大類空白，請先填入大類後再產生調整檔：")
        for row, name in category_problems:
            print(f"- 第 {row} 列｜{name}")
        print_category_reference(category_rules)
        raise SystemExit(3)

    if not rows:
        print_excluded_items(ws, excluded_rows, headers)
        die("排除後沒有需要輸出的商品。", code=4)

    for row in sorted(excluded_rows, reverse=True):
        ws.delete_rows(row)

    delete_output_columns(ws, header_row)
    polish_worksheet(ws, header_row)
    wb.save(output_path)
    print(f"已輸出：{output_path}")


if __name__ == "__main__":
    main()
