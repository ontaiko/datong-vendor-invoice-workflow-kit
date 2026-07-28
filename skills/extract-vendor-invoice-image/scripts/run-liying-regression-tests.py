from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook


def workbook_rows(path: Path) -> tuple[list[tuple], int | float | None]:
    wb = load_workbook(path, data_only=True)
    ws = wb["進貨明細"]
    headers = [ws.cell(2, column).value for column in range(1, 7)]
    expected_headers = ["產品代號", "品名", "零售價", "數量", "進價", "金額"]
    if headers != expected_headers:
        raise AssertionError(f"{path.name}: headers {headers!r} != {expected_headers!r}")

    rows: list[tuple] = []
    total = None
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] == "總價格":
            total = row[5]
            break
        rows.append(row)
    return rows, total


def assert_invoice(path: Path, expected_rows: int, expected_total: int, required_names: list[str]) -> None:
    rows, total = workbook_rows(path)
    amount_sum = sum(row[5] for row in rows if row[5] is not None)
    if len(rows) != expected_rows:
        raise AssertionError(f"{path.name}: rows {len(rows)} != {expected_rows}")
    if amount_sum != expected_total or total != expected_total:
        raise AssertionError(f"{path.name}: sum/total {amount_sum}/{total} != {expected_total}")
    names = [str(row[1]) for row in rows]
    missing = [name for name in required_names if name not in names]
    if missing:
        raise AssertionError(f"{path.name}: missing names {missing}")


def main() -> int:
    parser = argparse.ArgumentParser(description="Validate known Liying invoice OCR outputs.")
    parser.add_argument("--single", type=Path, required=True, help="1-row Liying workbook path")
    parser.add_argument("--multi", type=Path, required=True, help="26-row Liying workbook path")
    args = parser.parse_args()

    assert_invoice(args.single, 1, 995, ["DT 小熊學校(粉)"])
    assert_invoice(
        args.multi,
        26,
        18678,
        [
            "#007_879602 賓士AMG GT-R",
            "#066_102557 廣島電鐵650形",
            "#PRM32 本田NSX TYPE S",
            "AO-08 藍寶堅尼Reventon",
        ],
    )
    print("LIYING_REGRESSION_OK")
    return 0


if __name__ == "__main__":
    sys.exit(main())
