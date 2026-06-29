from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from statistics import median
from typing import Any

os.environ.setdefault("PADDLE_PDX_ENABLE_MKLDNN_BYDEFAULT", "0")

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageEnhance, ImageFilter, ImageOps
from paddleocr import PaddleOCR

try:
    import cv2
    import numpy as np
except ImportError:
    cv2 = None
    np = None


PROJECT_ROOT = Path(r"C:\Users\user\Documents\大統工作助手")
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "建檔進貨用" / "進貨圖片轉試算表"
DEFAULT_TMP_DIR = PROJECT_ROOT / ".codex-tmp" / "local-paddleocr"
DEFAULT_SETTINGS_PATH = PROJECT_ROOT / "參考資料" / "OCR設定.json"

KEY_TERMS = [
    "銷貨",
    "進貨",
    "貨號",
    "品號",
    "品名",
    "數量",
    "数量",
    "金額",
    "金额",
    "合計",
    "總計",
    "萬榮",
    "南波",
    "鉅霖",
    "BRICKROID",
    "BAP",
    "GSC",
]

PRODUCT_CODE_RE = re.compile(
    r"^\s*(?:\d+\s*)?((?:[A-Z]{2,8}-[A-Z0-9]+)|(?:[A-Z]{2,8}\d{4,})|(?:\d{6}))\s*(.*)$"
)
NUMBER_RE = re.compile(r"-?\d+(?:,\d{3})*(?:\.\d+)?")
MONEY_DECIMAL_RE = re.compile(r"\d,\d{3}\.\d+|\d{2,4}\.\d+")
MONEY_INTEGER_RE = re.compile(r"-?(?:\d{1,3}(?:,\d{3})+|\d{1,6})")
QUANTITY_RE = re.compile(r"[Vv]?\s*(\d{1,3})\s*(?:個|个|PCS|Pcs|pcs)")
TRADITIONAL_CHAR_MAP = str.maketrans(
    {
        "种": "種",
        "号": "號",
        "机": "機",
        "宝": "寶",
        "绒": "絨",
        "挂": "掛",
        "车": "車",
        "岛": "島",
        "与": "與",
        "坏": "壞",
        "猫": "貓",
        "樱": "櫻",
        "欧": "歐",
        "丽": "麗",
    }
)
SUSPICIOUS_SIMPLIFIED_CHARS = set("种号机宝绒挂车岛与坏猫樱欧丽")
UNIT_TOKENS = {"抽", "個", "个", "PCS", "Pcs", "pcs"}
GOOD_ENOUGH_PRODUCT_CODE_HITS = 2
GOOD_ENOUGH_KEY_HITS = 2


@dataclass
class OcrEntry:
    text: str
    score: float
    box: list[int]
    x: float
    y: float


@dataclass
class ProductRow:
    vendor_code: str
    name: str
    quantity: int | None
    unit_cost: float | None
    amount: float | None
    check: str
    issue: str


def roc_today() -> str:
    now = datetime.now()
    return f"{now.year - 1911:03d}{now.month:02d}{now.day:02d}"


def clean_number(text: str) -> float | None:
    if "%" in text:
        return None
    match = NUMBER_RE.search(text.replace("O", "0").replace("o", "0"))
    if not match:
        return None
    try:
        return float(match.group(0).replace(",", ""))
    except ValueError:
        return None


def parse_number_text(value: str) -> float | None:
    try:
        return float(value.replace(",", ""))
    except ValueError:
        return None


def extract_quantity(text: str) -> int | None:
    if text.strip().lower() == "d":
        return 4
    match = QUANTITY_RE.search(text)
    if match:
        return int(match.group(1))
    if is_unit_token(text):
        return None
    clean_integer = re.fullmatch(r"\s*[Vv]?\s*(\d{1,3})\s*", text)
    return int(clean_integer.group(1)) if clean_integer else None


def extract_decimal_money(text: str) -> float | None:
    # Handwritten check marks often get glued before the printed cost, e.g. "4個1L5T6,328.35".
    # Searching for the last money-like decimal preserves the printed cost and ignores the mark.
    segment = re.split(r"(?:個|个|PCS|Pcs|pcs)", text)[-1]
    matches = list(MONEY_DECIMAL_RE.finditer(segment))
    if not matches:
        matches = list(MONEY_DECIMAL_RE.finditer(text))
    if not matches:
        return None
    return parse_number_text(matches[-1].group(0))


def extract_integer_money(text: str) -> float | None:
    if "." in text:
        return None
    matches = list(MONEY_INTEGER_RE.finditer(text))
    if not matches:
        return None
    return parse_number_text(matches[-1].group(0))


def compact_amount(value: float | None) -> int | float | None:
    if value is None:
        return None
    if abs(value - round(value)) < 0.001:
        return int(round(value))
    return round(value, 3)


def suspicious_name_issues(name: str) -> list[str]:
    issues: list[str] = []
    suspicious_chars = sorted({char for char in name if char in SUSPICIOUS_SIMPLIFIED_CHARS})
    if suspicious_chars:
        issues.append(f"品名含疑似簡體或 OCR 錯字：{''.join(suspicious_chars)}")
    if re.search(r"\b[a-z]\d{3,}\b", name):
        issues.append("品名含疑似星號或符號誤讀的英文字母數字片段")
    return issues


def normalize_ocr_name(name: str) -> str:
    normalized = name.translate(TRADITIONAL_CHAR_MAP)
    normalized = re.sub(r"[＊*]\s*\d+(?:\.\d+)?\s*$", "", normalized)
    normalized = re.sub(r"(?<=[\u4e00-\u9fff])\s+(?=[\u4e00-\u9fff])", "", normalized)
    return normalized.strip()


def is_unit_token(text: str) -> bool:
    return text.strip().upper() == "PCS" or text.strip() in UNIT_TOKENS


def vendor_short(name: str) -> str:
    if "南波" in name:
        return "南波"
    if "萬榮" in name:
        return "萬榮"
    if "麗嬰" in name:
        return "麗嬰"
    if "鉅霖" in name:
        return "鉅霖"
    cleaned = re.sub(r"[^\w\u4e00-\u9fff]+", "", name)
    return cleaned[:4] or "進貨"


def unique_output_path(output_dir: Path, vendor: str) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    prefix = f"{vendor_short(vendor)}進貨單-{roc_today()}"
    for i in range(1, 100):
        candidate = output_dir / f"{prefix}-{i:02d}.xlsx"
        if not candidate.exists():
            return candidate
    raise RuntimeError(f"無法建立不重複檔名：{prefix}-NN.xlsx")


def process_is_running(pid: int) -> bool:
    if pid <= 0:
        return False
    if os.name == "nt":
        import ctypes

        process_query_limited_information = 0x1000
        handle = ctypes.windll.kernel32.OpenProcess(process_query_limited_information, False, pid)
        if not handle:
            return False
        ctypes.windll.kernel32.CloseHandle(handle)
        return True
    try:
        os.kill(pid, 0)
        return True
    except OSError:
        return False


def acquire_image_lock(image_path: Path) -> Path:
    lock_dir = DEFAULT_TMP_DIR / ".locks"
    lock_dir.mkdir(parents=True, exist_ok=True)
    digest = hashlib.sha1(str(image_path).lower().encode("utf-8")).hexdigest()[:12]
    lock_path = lock_dir / f"{image_path.stem}-{digest}.lock"

    for _ in range(2):
        try:
            fd = os.open(lock_path, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
        except FileExistsError:
            try:
                current_pid = int(lock_path.read_text(encoding="utf-8").splitlines()[0])
            except (OSError, ValueError, IndexError):
                current_pid = 0
            if process_is_running(current_pid):
                raise SystemExit(f"同一張圖片正在 OCR：{image_path}（PID {current_pid}）。請等待原程序完成。")
            lock_path.unlink(missing_ok=True)
            continue
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            f.write(f"{os.getpid()}\n{image_path}\n")
        return lock_path
    raise SystemExit(f"無法取得 OCR 圖片鎖：{lock_path}")


def release_image_lock(lock_path: Path) -> None:
    try:
        lines = lock_path.read_text(encoding="utf-8").splitlines()
        if lines and int(lines[0]) == os.getpid():
            lock_path.unlink(missing_ok=True)
    except (OSError, ValueError):
        pass


def edge_orientation_score(image: Image.Image) -> float:
    gray = ImageOps.grayscale(image)
    gray.thumbnail((900, 900))
    edges = gray.filter(ImageFilter.FIND_EDGES)
    width, height = edges.size
    pixel_data = edges.get_flattened_data() if hasattr(edges, "get_flattened_data") else edges.getdata()
    pixels = list(pixel_data)
    threshold = 48
    row_sums = [0] * height
    col_sums = [0] * width
    for y in range(height):
        offset = y * width
        for x in range(width):
            if pixels[offset + x] > threshold:
                row_sums[y] += 1
                col_sums[x] += 1
    top_rows = sorted(row_sums, reverse=True)[: max(3, height // 80)]
    top_cols = sorted(col_sums, reverse=True)[: max(3, width // 80)]
    horizontal = sum(top_rows)
    vertical = sum(top_cols)
    return horizontal / max(vertical, 1)


def auto_rotate_document(image: Image.Image) -> tuple[str, Image.Image]:
    candidates = [
        ("none", image),
        ("rot90cw", image.rotate(-90, expand=True)),
        ("rot90ccw", image.rotate(90, expand=True)),
        ("rot180", image.rotate(180, expand=True)),
    ]
    scored = [(edge_orientation_score(candidate), name, candidate) for name, candidate in candidates]
    best_score, name, best = max(scored, key=lambda item: item[0])
    none_score, _, none_image = scored[0]
    # Table borders often make 0/90-degree scores nearly identical. Keep the
    # source direction unless another direction is materially better.
    if none_score >= best_score * 0.95:
        return "none", none_image
    return name, best


def split_stacked_pages(image: Image.Image) -> list[tuple[str, Image.Image]]:
    if cv2 is None or np is None:
        return [("page1", image)]
    width, height = image.size
    if height < width * 1.15:
        return [("page1", image)]

    gray = cv2.cvtColor(np.array(image.convert("RGB")), cv2.COLOR_RGB2GRAY)
    ink_ratio = (gray < 180).mean(axis=1)
    edge_ratio = cv2.Canny(gray, 50, 150).mean(axis=1) / 255.0
    signal = ink_ratio + edge_ratio
    smooth = np.convolve(signal, np.ones(9) / 9, mode="same")
    low = int(height * 0.35)
    high = int(height * 0.65)
    center = smooth[low:high]
    if center.size == 0:
        return [("page1", image)]

    local_index = int(center.argmin())
    split_y = low + local_index
    minimum = float(center[local_index])
    baseline = float(np.median(center))
    if baseline < 0.06 or minimum > min(0.065, baseline * 0.5):
        return [("page1", image)]
    if not (height * 0.30 <= split_y <= height * 0.70):
        return [("page1", image)]

    threshold = minimum + (baseline - minimum) * 0.20
    band_top = split_y
    band_bottom = split_y
    while band_top > low and smooth[band_top - 1] <= threshold:
        band_top -= 1
    while band_bottom < high - 1 and smooth[band_bottom + 1] <= threshold:
        band_bottom += 1
    split_y = (band_top + band_bottom) // 2
    return [
        ("page1", image.crop((0, 0, width, split_y))),
        ("page2", image.crop((0, split_y, width, height))),
    ]


def make_photo_variant(
    image_path: Path,
    tmp_dir: Path,
    rotation: str,
    contrast: float,
    sharpness: float,
    page_split: str,
) -> list[tuple[str, Path]]:
    tmp_dir.mkdir(parents=True, exist_ok=True)
    image = ImageOps.exif_transpose(Image.open(image_path)).convert("RGB")
    variants: list[tuple[str, Path]] = []

    def save(name: str, img: Image.Image) -> None:
        out = tmp_dir / f"{image_path.stem}_tmp_{name}.jpg"
        img.save(out, quality=95)
        variants.append((name, out))

    pages = split_stacked_pages(image) if page_split == "auto" else [("page1", image)]
    multi_page = len(pages) > 1
    for page_name, page_image in pages:
        if rotation == "auto":
            rotate_name, page_image = auto_rotate_document(page_image)
        elif rotation == "cw":
            rotate_name, page_image = "rot90cw", page_image.rotate(-90, expand=True)
        elif rotation == "ccw":
            rotate_name, page_image = "rot90ccw", page_image.rotate(90, expand=True)
        elif rotation == "180":
            rotate_name, page_image = "rot180", page_image.rotate(180, expand=True)
        else:
            rotate_name = "none"

        enhanced = enhance_for_ocr(page_image, contrast, sharpness)
        page_part = f"_{page_name}" if multi_page else ""
        engine_part = "opencv_ocr" if cv2 is not None else "gray_contrast"
        save(f"photo{page_part}_{rotate_name}_{engine_part}", enhanced)
    return variants


def enhance_for_ocr(image: Image.Image, contrast: float, sharpness: float) -> Image.Image:
    if cv2 is None or np is None:
        gray = ImageOps.grayscale(image)
        enhanced = ImageEnhance.Contrast(gray).enhance(contrast)
        return ImageEnhance.Sharpness(enhanced).enhance(sharpness).convert("RGB")

    rgb = np.array(image.convert("RGB"))
    gray = cv2.cvtColor(rgb, cv2.COLOR_RGB2GRAY)
    denoised = cv2.fastNlMeansDenoising(gray, None, 8, 7, 21)
    clahe = cv2.createCLAHE(clipLimit=max(1.0, contrast), tileGridSize=(8, 8))
    enhanced = clahe.apply(denoised)
    if sharpness > 1:
        blurred = cv2.GaussianBlur(enhanced, (0, 0), 1.0)
        enhanced = cv2.addWeighted(enhanced, sharpness, blurred, 1 - sharpness, 0)
    return Image.fromarray(enhanced).convert("RGB")


def make_multi_variants(image_path: Path, tmp_dir: Path, contrast: float, sharpness: float) -> list[tuple[str, Path]]:
    tmp_dir.mkdir(parents=True, exist_ok=True)
    image = ImageOps.exif_transpose(Image.open(image_path)).convert("RGB")
    variants: list[tuple[str, Path]] = []

    def save(name: str, img: Image.Image) -> None:
        out = tmp_dir / f"{image_path.stem}_tmp_{name}.jpg"
        img.save(out, quality=95)
        variants.append((name, out))

    save("raw", image)
    save("rot90cw", image.rotate(-90, expand=True))
    save("rot90ccw", image.rotate(90, expand=True))
    save("rot180", image.rotate(180, expand=True))

    enhanced = enhance_for_ocr(image, contrast, sharpness)
    save("gray_contrast", enhanced)
    save("rot90cw_gray_contrast", enhance_for_ocr(image.rotate(-90, expand=True), contrast, sharpness))
    save("rot90ccw_gray_contrast", enhance_for_ocr(image.rotate(90, expand=True), contrast, sharpness))
    return variants


def extract_entries(page: Any) -> list[OcrEntry]:
    texts = page["rec_texts"]
    scores = page["rec_scores"]
    boxes = page["rec_boxes"]
    entries: list[OcrEntry] = []
    for text, score, box in zip(texts, scores, boxes):
        cleaned = str(text).strip()
        if not cleaned:
            continue
        b = [int(x) for x in box]
        entries.append(
            OcrEntry(
                text=cleaned,
                score=float(score),
                box=b,
                x=(b[0] + b[2]) / 2,
                y=(b[1] + b[3]) / 2,
            )
        )
    return entries


def is_good_enough_result(result: dict[str, Any]) -> bool:
    return (
        result["product_code_hits"] >= GOOD_ENOUGH_PRODUCT_CODE_HITS
        and result["key_hits"] >= GOOD_ENOUGH_KEY_HITS
        and result["line_count"] >= 10
    )


def is_bad_auto_rotation_result(result: dict[str, Any]) -> bool:
    return result["key_hits"] == 0 or result["product_code_hits"] < GOOD_ENOUGH_PRODUCT_CODE_HITS


def score_ocr_result(result: dict[str, Any]) -> tuple[int, int, int, int, int, float]:
    try:
        rows, _issues = parse_result_rows(result)
    except Exception:
        rows = []
    valid_rows = sum(1 for row in rows if row.check == "通過")
    return (
        valid_rows,
        len(rows),
        result["key_hits"],
        result["product_code_hits"],
        result["line_count"],
        result["avg_score"],
    )


def parse_result_rows(result: dict[str, Any]) -> tuple[list[ProductRow], list[str]]:
    page_entries = result.get("page_entries") or [result["entries"]]
    rows: list[ProductRow] = []
    issues: list[str] = []
    for page_index, entries in enumerate(page_entries, start=1):
        page_rows, page_issues = parse_rows(entries)
        rows.extend(page_rows)
        issues.extend(f"第{page_index}頁：{issue}" for issue in page_issues)
    return rows, issues


def merge_page_entries(page_entries: list[list[OcrEntry]]) -> list[OcrEntry]:
    merged: list[OcrEntry] = []
    y_offset = 0
    for entries in page_entries:
        page_bottom = 0
        for entry in entries:
            shifted_box = [entry.box[0], entry.box[1] + y_offset, entry.box[2], entry.box[3] + y_offset]
            merged.append(
                OcrEntry(
                    text=entry.text,
                    score=entry.score,
                    box=shifted_box,
                    x=entry.x,
                    y=entry.y + y_offset,
                )
            )
            page_bottom = max(page_bottom, entry.box[3])
        y_offset += page_bottom + 80
    return merged


def opposite_rotation_from_variant(variant: str) -> str | None:
    if "rot90cw" in variant:
        return "ccw"
    if "rot90ccw" in variant:
        return "cw"
    if "rot180" in variant:
        return "none"
    if "none" in variant:
        return "180"
    return None


def create_ocr() -> PaddleOCR:
    return PaddleOCR(
        lang="ch",
        use_doc_orientation_classify=False,
        use_doc_unwarping=False,
        use_textline_orientation=False,
    )


def run_ocr(
    image_path: Path,
    tmp_dir: Path,
    rotation: str,
    multi_variant: bool,
    contrast: float,
    sharpness: float,
    page_split: str,
) -> tuple[dict[str, Any], list[dict[str, Any]], dict[str, float]]:
    timings: dict[str, float] = {}
    start = time.perf_counter()
    ocr = create_ocr()
    timings["model_load_seconds"] = round(time.perf_counter() - start, 3)

    results: list[dict[str, Any]] = []
    variant_batches: list[list[tuple[str, Path]]] = []
    if multi_variant:
        variant_batches.append(make_multi_variants(image_path, tmp_dir, contrast, sharpness))
    else:
        variant_batches.append(make_photo_variant(image_path, tmp_dir, rotation, contrast, sharpness, page_split))

    for batch_index, variants in enumerate(variant_batches):
        if not multi_variant and len(variants) > 1:
            page_entries: list[list[OcrEntry]] = []
            page_paths: list[str] = []
            page_names: list[str] = []
            predict_seconds = 0.0
            for name, variant_path in variants:
                predict_start = time.perf_counter()
                pages = ocr.predict(str(variant_path))
                predict_seconds += time.perf_counter() - predict_start
                entries = extract_entries(pages[0]) if pages else []
                page_entries.append(entries)
                page_paths.append(str(variant_path))
                page_names.append(name)
            entries = merge_page_entries(page_entries)
            key_hits = sum(1 for item in entries if any(term in item.text for term in KEY_TERMS))
            product_code_hits = sum(1 for item in entries if PRODUCT_CODE_RE.match(item.text))
            avg_score = round(sum(item.score for item in entries) / len(entries), 4) if entries else 0
            results.append(
                {
                    "variant": f"split_{len(variants)}pages:" + ",".join(page_names),
                    "path": "\n".join(page_paths),
                    "line_count": len(entries),
                    "avg_score": avg_score,
                    "key_hits": key_hits,
                    "product_code_hits": product_code_hits,
                    "predict_seconds": round(predict_seconds, 3),
                    "entries": entries,
                    "page_entries": page_entries,
                }
            )
        else:
            for name, variant_path in variants:
                predict_start = time.perf_counter()
                pages = ocr.predict(str(variant_path))
                predict_seconds = round(time.perf_counter() - predict_start, 3)
                entries = extract_entries(pages[0]) if pages else []
                key_hits = sum(1 for item in entries if any(term in item.text for term in KEY_TERMS))
                product_code_hits = sum(1 for item in entries if PRODUCT_CODE_RE.match(item.text))
                avg_score = round(sum(item.score for item in entries) / len(entries), 4) if entries else 0
                result = {
                    "variant": name,
                    "path": str(variant_path),
                    "line_count": len(entries),
                    "avg_score": avg_score,
                    "key_hits": key_hits,
                    "product_code_hits": product_code_hits,
                    "predict_seconds": predict_seconds,
                    "entries": entries,
                }
                results.append(result)
                if multi_variant and is_good_enough_result(result):
                    break
        if multi_variant:
            break

        if rotation == "auto" and batch_index == 0 and results and is_bad_auto_rotation_result(results[-1]):
            retry_rotation = opposite_rotation_from_variant(results[-1]["variant"])
            if retry_rotation:
                variant_batches.append(make_photo_variant(image_path, tmp_dir, retry_rotation, contrast, sharpness, page_split))
                continue
        break
    best = sorted(
        results,
        key=score_ocr_result,
        reverse=True,
    )[0]
    timings["total_ocr_seconds"] = round(time.perf_counter() - start, 3)
    return best, results, timings


def load_settings(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    with path.open("r", encoding="utf-8-sig") as f:
        data = json.load(f)
    return data if isinstance(data, dict) else {}


def infer_vendor(entries: list[OcrEntry]) -> str:
    for entry in entries[:20]:
        if "南波" in entry.text or "萬榮" in entry.text or "麗嬰" in entry.text or "鉅霖" in entry.text:
            return entry.text
    for entry in entries[:20]:
        if "公司" in entry.text or "商行" in entry.text:
            return entry.text
    return "未辨識廠商"


def header_positions(entries: list[OcrEntry]) -> tuple[float, dict[str, float]]:
    positions: dict[str, list[float]] = {
        "code": [],
        "name": [],
        "quantity": [],
        "unit": [],
        "list_price": [],
        "discount": [],
        "unit_cost": [],
        "amount": [],
    }
    header_y_values = []
    width = max((entry.box[2] for entry in entries), default=1200)
    for entry in entries:
        text = entry.text
        matched = False
        if "貨號" in text or "品號" in text:
            positions["code"].append(entry.x)
            matched = True
        if "品名" in text:
            positions["name"].append(entry.x)
            matched = True
        if "數量" in text or "数量" in text:
            positions["quantity"].append(entry.x)
            matched = True
        if "單位" in text:
            positions["unit"].append(entry.x)
            matched = True
        if "售價" in text:
            positions["list_price"].append(entry.x)
            matched = True
        if "折" in text:
            positions["discount"].append(entry.x)
            matched = True
        if "進價" in text or "單價" in text or ("價" in text and entry.x > width * 0.55):
            positions["unit_cost"].append(entry.x)
            matched = True
        if "金額" in text or "金额" in text or ("金" in text and entry.x > width * 0.65):
            positions["amount"].append(entry.x)
            matched = True
        if matched:
            header_y_values.append(entry.y)
    defaults = {
        "code": width * 0.14,
        "name": width * 0.36,
        "quantity": width * 0.55,
        "unit": width * 0.62,
        "list_price": width * 0.70,
        "discount": width * 0.76,
        "unit_cost": width * 0.82,
        "amount": width * 0.93,
    }
    resolved = {key: (median(value) if value else defaults[key]) for key, value in positions.items()}
    header_y = max(header_y_values) if header_y_values else min((entry.y for entry in entries), default=0)
    return header_y, resolved


def nearest_column(entry: OcrEntry, positions: dict[str, float]) -> str:
    return min(positions, key=lambda key: abs(entry.x - positions[key]))


def parse_rows(entries: list[OcrEntry]) -> tuple[list[ProductRow], list[str]]:
    header_y, positions = header_positions(entries)
    code_like_entries = [entry for entry in entries if PRODUCT_CODE_RE.match(entry.text)]
    if code_like_entries:
        header_y = min(header_y, min(entry.y for entry in code_like_entries) - 40)
    total_y = min(
        (entry.y for entry in entries if any(term in entry.text for term in ["合計", "總計", "总计"])),
        default=max((entry.y for entry in entries), default=10_000),
    )

    code_entries: list[tuple[OcrEntry, str, str]] = []
    for entry in entries:
        if entry.y <= header_y or entry.y >= total_y:
            continue
        match = PRODUCT_CODE_RE.match(entry.text)
        if not match:
            continue
        if abs(entry.x - positions["code"]) > max(240, positions["name"] - positions["code"] + 80):
            continue
        code_entries.append((entry, match.group(1), match.group(2).strip()))

    code_entries.sort(key=lambda item: item[0].y)
    row_spacings = [
        code_entries[i + 1][0].y - code_entries[i][0].y
        for i in range(len(code_entries) - 1)
        if code_entries[i + 1][0].y > code_entries[i][0].y
    ]
    typical_spacing = median(row_spacings) if row_spacings else 50
    rows: list[ProductRow] = []
    global_issues: list[str] = []
    for idx, (code_entry, code, trailing_name) in enumerate(code_entries):
        previous_code = code_entries[idx - 1][0] if idx > 0 else None
        next_code = code_entries[idx + 1][0] if idx + 1 < len(code_entries) else None
        row_start_y = (
            previous_code.y + (code_entry.y - previous_code.y) * 0.90
            if previous_code is not None
            else code_entry.box[1] - 3
        )
        next_y = (
            code_entry.y + (next_code.y - code_entry.y) * 0.90
            if next_code is not None
            else min(total_y, code_entry.y + typical_spacing * 1.8)
        )
        name_start_y = row_start_y
        name_end_y = next_y
        numeric_start_y = row_start_y
        numeric_end_y = next_y
        row_entries = [entry for entry in entries if name_start_y <= entry.y < numeric_end_y]

        name_parts: list[str] = []
        if trailing_name:
            name_parts.append(trailing_name)
        numeric_by_col: dict[str, list[float]] = {key: [] for key in positions}
        quantity_candidates: list[int] = []
        unit_cost_candidates: list[float] = []
        amount_candidates: list[float] = []
        name_left = positions["code"] + 20
        name_right_candidates = [positions["quantity"] - 20]
        if positions["code"] < positions["unit"] < positions["quantity"]:
            name_right_candidates.append(positions["unit"] - 12)
        name_right = min(name_right_candidates)
        quantity_left = positions["quantity"] - 120
        unit_cost_left = min(positions["unit_cost"], positions["quantity"]) - 70
        amount_left = positions["amount"] - max(20, (positions["amount"] - positions["unit_cost"]) * 0.25)
        for entry in sorted(row_entries, key=lambda item: (item.y, item.x)):
            if entry is code_entry:
                continue
            col = nearest_column(entry, positions)
            if is_unit_token(entry.text):
                continue
            if numeric_start_y <= entry.y < numeric_end_y:
                number = clean_number(entry.text)
                if number is not None:
                    numeric_by_col[col].append(number)
                if entry.x >= quantity_left:
                    quantity = extract_quantity(entry.text)
                    if quantity is not None:
                        quantity_candidates.append(quantity)
                if unit_cost_left <= entry.x < positions["amount"]:
                    unit_cost_value = extract_decimal_money(entry.text)
                    if unit_cost_value is not None:
                        unit_cost_candidates.append(unit_cost_value)
                if entry.x >= amount_left:
                    amount_value = extract_integer_money(entry.text)
                    if amount_value is not None:
                        amount_candidates.append(amount_value)
            if name_start_y <= entry.y < name_end_y and name_left < entry.x < name_right:
                if not PRODUCT_CODE_RE.match(entry.text) and "以下" not in entry.text:
                    name_parts.append(entry.text)

        quantity = None
        if quantity_candidates:
            quantity = quantity_candidates[0]
        elif numeric_by_col["quantity"]:
            integer_values = [int(round(n)) for n in numeric_by_col["quantity"] if abs(n - round(n)) < 0.001]
            quantity = integer_values[0] if integer_values else int(round(numeric_by_col["quantity"][0]))

        unit_cost = None
        if unit_cost_candidates:
            unit_cost = unit_cost_candidates[-1]
        elif numeric_by_col["unit_cost"]:
            unit_cost = numeric_by_col["unit_cost"][0]

        amount = None
        if amount_candidates:
            amount = amount_candidates[-1]

        if unit_cost is None and quantity and amount is not None:
            unit_cost = amount / quantity
        if amount is None and quantity is not None and unit_cost is not None:
            amount = quantity * unit_cost

        name = " ".join(part.strip() for part in name_parts if part.strip())
        name = re.sub(r"\s+", " ", name).strip()
        name = normalize_ocr_name(name)
        if not name:
            name = code

        issue_parts: list[str] = []
        issue_parts.extend(suspicious_name_issues(name))
        if quantity is None:
            issue_parts.append("數量未穩定辨識")
        if unit_cost is None:
            issue_parts.append("進價未穩定辨識")
        if amount is None:
            issue_parts.append("金額未穩定辨識")

        check = "通過"
        if quantity is not None and unit_cost is not None and amount is not None:
            diff = abs(quantity * unit_cost - amount)
            tolerance = max(1, abs(amount) * 0.01)
            if diff > tolerance:
                check = "不符"
                issue_parts.append(f"數量×進價={quantity * unit_cost:.3f}，與金額 {amount:.3f} 不符")
        else:
            check = "需確認"

        rows.append(
            ProductRow(
                vendor_code=code,
                name=name,
                quantity=quantity,
                unit_cost=compact_amount(unit_cost),
                amount=compact_amount(amount),
                check=check,
                issue="；".join(issue_parts) if issue_parts else "無",
            )
        )

    if not rows:
        global_issues.append("未能從 OCR 座標穩定切出商品列，請查看 OCR測試紀錄工作表。")
    return rows, global_issues


def infer_total(entries: list[OcrEntry], rows: list[ProductRow]) -> float | None:
    for i, entry in enumerate(entries):
        if "總計" in entry.text or "总计" in entry.text:
            nearby = entries[i : i + 5]
            for candidate in nearby:
                number = clean_number(candidate.text)
                if number is not None and number > 0:
                    return number
    amounts = [float(row.amount) for row in rows if row.amount is not None]
    return sum(amounts) if amounts else None


def write_workbook(
    image_path: Path,
    output_path: Path,
    vendor: str,
    best: dict[str, Any],
    all_results: list[dict[str, Any]],
    rows: list[ProductRow],
    global_issues: list[str],
    timings: dict[str, float],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "進貨明細"
    ws.append([f"廠商：{vendor}", None, None, None, None])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws.append(["產品代號", "品名", "數量", "進價", "金額"])
    for row in rows:
        ws.append(["", row.name, row.quantity, row.unit_cost, row.amount])
    total = infer_total(best["entries"], rows)
    ws.append(["總價格", None, None, None, compact_amount(total)])
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=4)

    header_fill = PatternFill("solid", fgColor="D9EAF0")
    title_fill = PatternFill("solid", fgColor="1F4E5F")
    thin = Side(style="thin", color="D9E2E6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws["A1"].fill = title_fill
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=(cell.column == 2))
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=3, max_col=5):
        for cell in row:
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right")
    for cell in ws["A"]:
        cell.number_format = "@"
    widths = {"A": 12, "B": 58, "C": 12, "D": 12, "E": 12}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A3"

    note = wb.create_sheet("OCR測試紀錄")
    note.append(["來源圖片", str(image_path)])
    note.append(["最佳版本", best["variant"]])
    note.append(["平均信心", best["avg_score"]])
    note.append(["關鍵字命中", best["key_hits"]])
    note.append(["商品列數", len(rows)])
    note.append(["偵測頁數", len(best.get("page_entries") or [best["entries"]])])
    note.append(["模型載入秒數", timings.get("model_load_seconds", "")])
    note.append(["OCR總秒數", timings.get("total_ocr_seconds", "")])
    note.append(["整體疑點", "；".join(global_issues) if global_issues else "無"])
    note.append([])
    note.append(["版本", "行數", "平均信心", "關鍵字命中", "貨號命中", "辨識秒數", "暫存圖"])
    for result in sorted(
        all_results,
        key=lambda r: (r["product_code_hits"], r["key_hits"], r["line_count"], r["avg_score"]),
        reverse=True,
    ):
        note.append(
            [
                result["variant"],
                result["line_count"],
                result["avg_score"],
                result["key_hits"],
                result["product_code_hits"],
                result.get("predict_seconds", ""),
                result["path"],
            ]
        )
    note.append([])
    note.append(["列號", "廠商貨號", "品名", "數量", "進價", "金額", "金額核對", "OCR疑點"])
    for idx, row in enumerate(rows, start=1):
        note.append([idx, row.vendor_code, row.name, row.quantity, row.unit_cost, row.amount, row.check, row.issue])
    note.append([])
    note.append(["原始OCR文字", "信心", "x1", "y1", "x2", "y2"])
    for entry in best["entries"]:
        note.append([entry.text, round(entry.score, 4), *entry.box])

    for sheet in [note]:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col in range(1, sheet.max_column + 1):
            letter = get_column_letter(col)
            sheet.column_dimensions[letter].width = 18
        sheet.column_dimensions["A"].width = 36
        sheet.column_dimensions["C"].width = 50
        sheet.freeze_panes = "A2"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Use local PaddleOCR to convert a vendor invoice image into review xlsx.")
    parser.add_argument("image", type=Path, help="進貨單圖片路徑")
    parser.add_argument("--output", type=Path, default=None, help="指定輸出 xlsx 路徑")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR, help="輸出資料夾")
    parser.add_argument("--tmp-dir", type=Path, default=DEFAULT_TMP_DIR, help="OCR 暫存資料夾")
    parser.add_argument("--settings", type=Path, default=DEFAULT_SETTINGS_PATH, help="OCR 設定 JSON 路徑")
    parser.add_argument(
        "--rotation",
        choices=["auto", "none", "cw", "ccw", "180"],
        default=None,
        help="照片前處理旋轉方式；未指定時讀取 OCR設定.json，預設 auto。",
    )
    parser.add_argument("--multi-variant", action="store_true", help="手動開啟舊式多版本 OCR，命中足夠結果後早停。")
    parser.add_argument("--contrast", type=float, default=None, help="灰階對比增強倍率；未指定時讀取 OCR設定.json。")
    parser.add_argument("--sharpness", type=float, default=None, help="銳化倍率；未指定時讀取 OCR設定.json。")
    parser.add_argument(
        "--page-split",
        choices=["auto", "off"],
        default=None,
        help="上下多頁照片自動分割；未指定時讀取 OCR設定.json，預設 auto。",
    )
    args = parser.parse_args()

    image_path = args.image.expanduser().resolve()
    if not image_path.exists():
        raise SystemExit(f"找不到圖片：{image_path}")

    lock_path = acquire_image_lock(image_path)
    try:
        settings = load_settings(args.settings.expanduser().resolve()) if args.settings else {}
        rotation = args.rotation or str(settings.get("rotation", "auto"))
        multi_variant = bool(args.multi_variant or settings.get("multi_variant", False))
        contrast = float(args.contrast if args.contrast is not None else settings.get("contrast", 1.8))
        sharpness = float(args.sharpness if args.sharpness is not None else settings.get("sharpness", 1.15))
        page_split = args.page_split or str(settings.get("page_split", "auto"))

        best, all_results, timings = run_ocr(
            image_path,
            args.tmp_dir / image_path.stem,
            rotation,
            multi_variant,
            contrast,
            sharpness,
            page_split,
        )
        vendor = infer_vendor(best["entries"])
        rows, global_issues = parse_result_rows(best)
        output_path = args.output.expanduser().resolve() if args.output else unique_output_path(args.output_dir, vendor)
        write_workbook(image_path, output_path, vendor, best, all_results, rows, global_issues, timings)

        summary = {
            "ok": True,
            "output": str(output_path),
            "vendor": vendor,
            "best_variant": best["variant"],
            "row_count": len(rows),
            "page_count": len(best.get("page_entries") or [best["entries"]]),
            "model_load_seconds": timings.get("model_load_seconds"),
            "total_ocr_seconds": timings.get("total_ocr_seconds"),
            "needs_review": bool(global_issues or any(row.issue != "無" or row.check != "通過" for row in rows)),
            "issues": global_issues + [
                f"{row.vendor_code}: {row.issue}" for row in rows if row.issue != "無" or row.check != "通過"
            ],
        }
        print(json.dumps(summary, ensure_ascii=False, indent=2))
        return 0
    finally:
        release_image_lock(lock_path)


if __name__ == "__main__":
    sys.exit(main())
