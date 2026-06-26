from __future__ import annotations

import csv
import json
import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook


PACKAGE_ROOT = Path(__file__).resolve().parents[1]
MATCH_SCRIPT = PACKAGE_ROOT / "skills" / "match-product-catalog" / "scripts" / "match-existing-products.py"

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


def make_fixture(test_dir: Path) -> tuple[Path, Path]:
    input_xlsx = test_dir / "測試進貨單.xlsx"
    catalog_csv = test_dir / "產品資料輸出.CSV"
    identity_csv = test_dir / "產品比對身份關鍵詞.csv"

    wb = Workbook()
    ws = wb.active
    ws.title = "進貨明細"
    ws.append(["廠商：測試代理商"])
    ws.append(["產品代號", "品名", "數量", "進價", "金額"])
    ws.append(["", "NS2 瑪利歐賽車世界 中文版", 3, 1590, 4770])
    ws.append(["", "PS5 魔物獵人荒野", 2, 1990, 3980])
    ws.append(["", "NS2 咚奇剛 蕉力全開", 1, 1790, 1790])
    ws.append(["000123", "測試已填代號商品", 1, 1000, 1000])
    ws.append(["", "總價格", "", "", 11540])
    wb.save(input_xlsx)

    write_csv(
        catalog_csv,
        [
            {"1.產品代號": "000001", "2.產品名稱": "NS2 瑪利歐賽車 世界 中文版"},
            {"1.產品代號": "000002", "2.產品名稱": "PS5 魔物獵人 荒野 中文版"},
            {"1.產品代號": "000123", "2.產品名稱": "PS5 DualSense 無線控制器 白色"},
            {"1.產品代號": "000004", "2.產品名稱": "NS 薩爾達無雙 封印戰記 中文版"},
        ],
    )
    write_csv(
        identity_csv,
        [
            {"類型": "版本詞", "關鍵詞": "中文版", "啟用": "Y", "備註": ""},
            {"類型": "角色機體", "關鍵詞": "咚奇剛", "啟用": "Y", "備註": ""},
        ],
    )
    return input_xlsx, catalog_csv


def read_results(output_xlsx: Path) -> dict[str, object]:
    wb = load_workbook(output_xlsx)
    ws = wb.active
    headers = {str(cell.value): index + 1 for index, cell in enumerate(ws[2]) if cell.value}
    rows = []
    for row_index in range(3, 7):
        status_cell = ws.cell(row_index, headers["比對狀態"])
        product_name_cell = ws.cell(row_index, headers["已建檔品名"])
        rows.append(
            {
                "品名": str(ws.cell(row_index, headers["品名"]).value or ""),
                "比對狀態": str(status_cell.value or ""),
                "相似候選": str(ws.cell(row_index, headers["相似候選"]).value or ""),
                "比對狀態註解": status_cell.comment.text if status_cell.comment else "",
                "已建檔品名底色": str(product_name_cell.fill.fgColor.rgb or ""),
            }
        )
    return {"rows": rows}


def main() -> None:
    with tempfile.TemporaryDirectory(prefix="datong-local-engine-test-") as tmp:
        test_dir = Path(tmp)
        input_xlsx, catalog_csv = make_fixture(test_dir)
        output_xlsx = input_xlsx.with_name(f"{input_xlsx.stem}_產品比對檢查.xlsx")

        completed = subprocess.run(
            [
                sys.executable,
                str(MATCH_SCRIPT),
                "--input-xlsx",
                str(input_xlsx),
                "--csv",
                str(catalog_csv),
            ],
            check=True,
            cwd=test_dir,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
        )

        result = read_results(output_xlsx)
        rows = result["rows"]
        expected_statuses = ["已建檔", "有類似產品", "確認為新品", "已建檔"]
        actual_statuses = [row["比對狀態"] for row in rows]
        result["expected_statuses"] = expected_statuses
        result["actual_statuses"] = actual_statuses
        result["code_warning_ok"] = "差異過大" in rows[3]["比對狀態註解"]
        result["warning_fill_ok"] = rows[3]["已建檔品名底色"] == "00FFF2CC"
        result["ok"] = (
            actual_statuses == expected_statuses
            and result["code_warning_ok"]
            and result["warning_fill_ok"]
        )
        print(json.dumps(result, ensure_ascii=False, indent=2))
        if not result["ok"]:
            raise SystemExit(1)


if __name__ == "__main__":
    main()
