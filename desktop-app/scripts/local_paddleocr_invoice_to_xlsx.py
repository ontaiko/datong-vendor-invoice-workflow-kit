from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from statistics import median
from typing import Any

os.environ.setdefault("PADDLE_PDX_ENABLE_MKLDNN_BYDEFAULT", "0")

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageEnhance, ImageFilter, ImageOps
from paddleocr import PaddleOCR

try:
    import cv2
    import numpy as np
except ImportError:
    cv2 = None
    np = None


PROJECT_ROOT = Path(
    os.path.expandvars(
        os.environ.get("DATONG_WORKSPACE")
        or str(Path.home() / "Documents" / "大統工作助手")
    )
).expanduser()
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "建檔進貨用" / "進貨圖片轉試算表"
DEFAULT_TMP_DIR = PROJECT_ROOT / ".codex-tmp" / "local-paddleocr"
DEFAULT_SETTINGS_PATH = PROJECT_ROOT / "參考資料" / "OCR設定.json"

KEY_TERMS = [
    "銷貨",
    "進貨",
    "貨號",
    "品號",
    "品名",
    "數量",
    "数量",
    "金額",
    "金额",
    "合計",
    "總計",
    "萬榮",
    "南波",
    "鉅霖",
    "元大",
    "BRICKROID",
    "BAP",
    "GSC",
]

PRODUCT_CODE_RE = re.compile(
    r"^\s*(?:\d+\s*)?((?:[A-Z]{2,8}-[A-Z0-9]+)|(?:[A-Z]{2,8}\d{4,})|(?:\d{6}))\s*(.*)$"
)
YUANDA_ITEM_RE = re.compile(r"^\s*(?:\d+\s+)?(\d{9})(.*)$")
LIYING_ITEM_RE = re.compile(r"^\s*(TM[A-Z0-9]{3,6}|[A-Z]{2}\d{5})\s*$")
NUMBER_RE = re.compile(r"-?\d+(?:,\d{3})*(?:\.\d+)?")
MONEY_DECIMAL_RE = re.compile(r"\d,\d{3}\.\d+|\d{2,4}\.\d+")
MONEY_INTEGER_RE = re.compile(r"-?(?:\d{1,3}(?:,\d{3})+|\d{1,6})")
QUANTITY_RE = re.compile(r"[Vv]?\s*(\d{1,3})\s*(?:個|个|PCS|Pcs|pcs)")
TRADITIONAL_CHAR_MAP = str.maketrans(
    {
        "种": "種",
        "号": "號",
        "机": "機",
        "宝": "寶",
        "绒": "絨",
        "挂": "掛",
        "车": "車",
        "岛": "島",
        "与": "與",
        "坏": "壞",
        "猫": "貓",
        "樱": "櫻",
        "欧": "歐",
        "丽": "麗",
    }
)
SUSPICIOUS_SIMPLIFIED_CHARS = set("种号机宝绒挂车岛与坏猫樱欧丽")
UNIT_TOKENS = {"抽", "個", "个", "PCS", "Pcs", "pcs"}
GOOD_ENOUGH_PRODUCT_CODE_HITS = 2
GOOD_ENOUGH_KEY_HITS = 2
PAGE_NUMBER_RE = re.compile(r"(?<!\d)(\d{1,2})\s*[/／]\s*(\d{1,2})(?!\d)")
NAME_HEADER_RE = re.compile(
    r"^\s*(?:品名規格|品名规格|品名/?批號|品名/?批号|"
    r"商品名稱|商品名称|產品名稱|产品名称)\s*[:：]?\s*"
)
SEPARATOR_NOISE_RE = re.compile(
    r"(?:[＊*※]{5,}|(?:以下|人下|以卜).{0,4}(?:空白|空自|空日)|"
    r"(?:冰木|木冰|冰木木|卡韋|卡韦|字常辛){2,})"
)


@dataclass
class OcrEntry:
    text: str
    score: float
    box: list[int]
    x: float
    y: float


@dataclass
class ProductRow:
    vendor_code: str
    name: str
    quantity: int | None
    unit_cost: float | None
    amount: float | None
    check: str
    issue: str
    retail_price: int | None = None
    audit_note: str = ""


def roc_today() -> str:
    now = datetime.now()
    return f"{now.year - 1911:03d}{now.month:02d}{now.day:02d}"


def clean_number(text: str) -> float | None:
    if "%" in text:
        return None
    match = NUMBER_RE.search(text.replace("O", "0").replace("o", "0"))
    if not match:
        return None
    try:
        return float(match.group(0).replace(",", ""))
    except ValueError:
        return None


def parse_number_text(value: str) -> float | None:
    try:
        return float(value.replace(",", ""))
    except ValueError:
        return None


def extract_quantity(text: str) -> int | None:
    if text.strip().lower() == "d":
        return 4
    match = QUANTITY_RE.search(text)
    if match:
        return int(match.group(1))
    if is_unit_token(text):
        return None
    clean_integer = re.fullmatch(r"\s*[Vv]?\s*(\d{1,3})\s*", text)
    return int(clean_integer.group(1)) if clean_integer else None


def extract_decimal_money(text: str) -> float | None:
    # Handwritten check marks often get glued before the printed cost, e.g. "4個1L5T6,328.35".
    # Searching for the last money-like decimal preserves the printed cost and ignores the mark.
    segment = re.split(r"(?:個|个|PCS|Pcs|pcs)", text)[-1]
    matches = list(MONEY_DECIMAL_RE.finditer(segment))
    if not matches:
        matches = list(MONEY_DECIMAL_RE.finditer(text))
    if not matches:
        return None
    return parse_number_text(matches[-1].group(0))


def extract_integer_money(text: str) -> float | None:
    if "." in text:
        return None
    matches = list(MONEY_INTEGER_RE.finditer(text))
    if not matches:
        return None
    return parse_number_text(matches[-1].group(0))


def extract_amount_integer(text: str) -> float | None:
    stripped = text.strip()
    if re.fullmatch(r"\d{1,3}\.\d{3}", stripped):
        return parse_number_text(stripped.replace(".", ""))
    return extract_integer_money(text)


def compact_amount(value: float | None) -> int | float | None:
    if value is None:
        return None
    if abs(value - round(value)) < 0.001:
        return int(round(value))
    return round(value, 3)


def rounded_invoice_amount(value: float | None) -> int | None:
    if value is None:
        return None
    return int(value + 0.5)


def inferred_unit_cost(quantity: int | None, amount: float | None) -> float | None:
    if quantity is None or quantity <= 0 or amount is None:
        return None
    inferred = amount / quantity
    if inferred <= 0:
        return None
    return inferred


def swapped_quantity_unit_cost(
    quantity: int | None,
    unit_cost: float | None,
    amount: float | None,
    tolerance: float,
) -> tuple[int, float] | None:
    if quantity is None or unit_cost is None or amount is None:
        return None
    if quantity <= 0 or unit_cost <= 0:
        return None
    if abs(unit_cost - round(unit_cost)) >= 0.001:
        return None
    swapped_quantity = int(round(unit_cost))
    swapped_unit_cost = float(quantity)
    if swapped_quantity <= 0 or swapped_unit_cost <= 0:
        return None
    if abs(swapped_quantity * swapped_unit_cost - amount) <= tolerance:
        return swapped_quantity, swapped_unit_cost
    return None


def invoice_amount_matches(quantity: int | None, unit_cost: float | None, amount: float | None) -> bool:
    if quantity is None or unit_cost is None or amount is None:
        return False
    if quantity <= 0 or unit_cost <= 0 or amount <= 0:
        return False
    raw = quantity * unit_cost
    tolerance = max(1, abs(amount) * 0.01)
    return abs(raw - amount) <= tolerance or rounded_invoice_amount(raw) == rounded_invoice_amount(amount)


def infer_total_quantity(entries: list[OcrEntry]) -> int | None:
    labels = [entry for entry in entries if "合計數量" in entry.text or "數量總計" in entry.text]
    for label in labels:
        nearby = [
            entry
            for entry in entries
            if entry is not label
            and entry.x > label.x
            and entry.x < label.x + 360
            and abs(entry.y - label.y) <= 70
        ]
        for entry in sorted(nearby, key=lambda item: (abs(item.y - label.y), item.x)):
            quantity = extract_quantity(entry.text)
            if quantity is not None and quantity > 0:
                return quantity
    return None


def suspicious_name_issues(name: str) -> list[str]:
    issues: list[str] = []
    suspicious_chars = sorted({char for char in name if char in SUSPICIOUS_SIMPLIFIED_CHARS})
    if suspicious_chars:
        issues.append(f"品名含疑似簡體或 OCR 錯字：{''.join(suspicious_chars)}")
    if re.search(r"\b[a-z]\d{3,}\b", name):
        issues.append("品名含疑似星號或符號誤讀的英文字母數字片段")
    if re.search(r"[A-Z]{5,}[\u4e00-\u9fff]", name):
        issues.append("品名英文與中文疑似黏字，請核對原圖")
    if re.search(r"([\u4e00-\u9fff])\1{3,}", name):
        issues.append("品名含異常重複字元，請核對原圖")
    return issues


def normalize_ocr_name(name: str) -> str:
    normalized = name.translate(TRADITIONAL_CHAR_MAP)
    normalized = NAME_HEADER_RE.sub("", normalized)
    separator_match = SEPARATOR_NOISE_RE.search(normalized)
    if separator_match:
        normalized = normalized[: separator_match.start()]
    normalized = re.sub(r"[＊*]\s*\d+(?:\.\d+)?\s*$", "", normalized)
    normalized = re.sub(r"^[＊*\s]+|[＊*\s]+$", "", normalized)
    normalized = re.sub(r"(?<=[\u4e00-\u9fff])\s+(?=[\u4e00-\u9fff])", "", normalized)
    normalized = re.sub(r"\bwith\s+", "with", normalized, flags=re.IGNORECASE)
    normalized = normalized.replace("腊大黨", "膽大黨")
    return normalized.strip()


def is_unit_token(text: str) -> bool:
    return text.strip().upper() == "PCS" or text.strip() in UNIT_TOKENS


def vendor_short(name: str) -> str:
    if "南波" in name:
        return "南波"
    if "萬榮" in name:
        return "萬榮"
    if "麗嬰" in name:
        return "麗嬰"
    if "鉅霖" in name:
        return "鉅霖"
    if "元大玩具" in name or "元大" in name:
        return "元大玩具"
    if "金亞特" in name:
        return "金亞特"
    cleaned = re.sub(r"[^\w\u4e00-\u9fff]+", "", name)
    return cleaned[:4] or "進貨"


def unique_output_path(output_dir: Path, vendor: str) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    prefix = f"{vendor_short(vendor)}進貨單-{roc_today()}"
    for i in range(1, 100):
        candidate = output_dir / f"{prefix}-{i:02d}.xlsx"
        if not candidate.exists():
            return candidate
    raise RuntimeError(f"無法建立不重複檔名：{prefix}-NN.xlsx")


def process_is_running(pid: int) -> bool:
    if pid <= 0:
        return False
    if os.name == "nt":
        import ctypes

        process_query_limited_information = 0x1000
        handle = ctypes.windll.kernel32.OpenProcess(process_query_limited_information, False, pid)
        if not handle:
            return False
        ctypes.windll.kernel32.CloseHandle(handle)
        return True
    try:
        os.kill(pid, 0)
        return True
    except OSError:
        return False


def acquire_image_lock(image_path: Path) -> Path:
    lock_dir = DEFAULT_TMP_DIR / ".locks"
    lock_dir.mkdir(parents=True, exist_ok=True)
    digest = hashlib.sha1(str(image_path).lower().encode("utf-8")).hexdigest()[:12]
    lock_path = lock_dir / f"{image_path.stem}-{digest}.lock"

    for _ in range(2):
        try:
            fd = os.open(lock_path, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
        except FileExistsError:
            try:
                current_pid = int(lock_path.read_text(encoding="utf-8").splitlines()[0])
            except (OSError, ValueError, IndexError):
                current_pid = 0
            if process_is_running(current_pid):
                raise SystemExit(f"同一張圖片正在 OCR：{image_path}（PID {current_pid}）。請等待原程序完成。")
            lock_path.unlink(missing_ok=True)
            continue
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            f.write(f"{os.getpid()}\n{image_path}\n")
        return lock_path
    raise SystemExit(f"無法取得 OCR 圖片鎖：{lock_path}")


def release_image_lock(lock_path: Path) -> None:
    try:
        lines = lock_path.read_text(encoding="utf-8").splitlines()
        if lines and int(lines[0]) == os.getpid():
            lock_path.unlink(missing_ok=True)
    except (OSError, ValueError):
        pass


def edge_orientation_score(image: Image.Image) -> float:
    gray = ImageOps.grayscale(image)
    gray.thumbnail((900, 900))
    edges = gray.filter(ImageFilter.FIND_EDGES)
    width, height = edges.size
    pixel_data = edges.get_flattened_data() if hasattr(edges, "get_flattened_data") else edges.getdata()
    pixels = list(pixel_data)
    threshold = 48
    row_sums = [0] * height
    col_sums = [0] * width
    for y in range(height):
        offset = y * width
        for x in range(width):
            if pixels[offset + x] > threshold:
                row_sums[y] += 1
                col_sums[x] += 1
    top_rows = sorted(row_sums, reverse=True)[: max(3, height // 80)]
    top_cols = sorted(col_sums, reverse=True)[: max(3, width // 80)]
    horizontal = sum(top_rows)
    vertical = sum(top_cols)
    return horizontal / max(vertical, 1)


def auto_rotate_document(image: Image.Image) -> tuple[str, Image.Image]:
    # EXIF transpose already normalizes ordinary phone photos.  Table borders
    # make edge-only 0/90-degree scores unreliable and previously rotated
    # readable invoices sideways.  Keep the source direction first; run_ocr
    # only tries a 90-degree alternative when the source result is structurally
    # unusable.
    return "none", image


def _split_plan_for_count(
    smooth: Any,
    edge_ratio: Any,
    height: int,
    page_count: int,
) -> tuple[float, list[int]] | None:
    if page_count not in (2, 3):
        return None
    cuts: list[int] = []
    strengths: list[float] = []
    search_radius = int(height * (0.16 if page_count == 2 else 0.105))
    minimum_page_height = int(height * 0.20)
    for index in range(1, page_count):
        target = int(height * index / page_count)
        low = max(minimum_page_height, target - search_radius)
        high = min(height - minimum_page_height, target + search_radius)
        if high - low < 10:
            return None
        window = smooth[low:high]
        local_index = int(window.argmin())
        cut = low + local_index
        minimum = float(smooth[cut])
        baseline_low = max(0, cut - int(height * 0.09))
        baseline_high = min(height, cut + int(height * 0.09))
        baseline = float(np.percentile(smooth[baseline_low:baseline_high], 75))
        edge_minimum = float(edge_ratio[cut])
        edge_baseline = float(np.percentile(edge_ratio[baseline_low:baseline_high], 75))
        valley_strength = 1.0 - minimum / max(baseline, 0.0001)
        edge_strength = 1.0 - edge_minimum / max(edge_baseline, 0.0001)
        if baseline < 0.035 or valley_strength < 0.30 or edge_strength < 0.24:
            return None

        threshold = minimum + (baseline - minimum) * 0.22
        band_top = cut
        band_bottom = cut
        while band_top > low and smooth[band_top - 1] <= threshold:
            band_top -= 1
        while band_bottom < high - 1 and smooth[band_bottom + 1] <= threshold:
            band_bottom += 1
        cuts.append((band_top + band_bottom) // 2)
        strengths.append(valley_strength + edge_strength * 0.45)

    boundaries = [0, *cuts, height]
    page_heights = [boundaries[i + 1] - boundaries[i] for i in range(page_count)]
    if min(page_heights) < minimum_page_height:
        return None
    average_height = height / page_count
    uniformity = 1.0 - max(abs(page_height - average_height) for page_height in page_heights) / average_height
    if uniformity < 0.45:
        return None
    return sum(strengths) + uniformity * 0.35 + page_count * 0.03, cuts


def split_stacked_pages(image: Image.Image) -> list[tuple[str, Image.Image]]:
    if cv2 is None or np is None:
        return [("page1", image)]
    width, height = image.size
    if height < width * 1.15:
        return [("page1", image)]

    gray = cv2.cvtColor(np.array(image.convert("RGB")), cv2.COLOR_RGB2GRAY)
    ink_ratio = (gray < 180).mean(axis=1)
    edge_ratio = cv2.Canny(gray, 50, 150).mean(axis=1) / 255.0
    signal = ink_ratio + edge_ratio
    smooth = np.convolve(signal, np.ones(9) / 9, mode="same")
    plans = [
        plan
        for page_count in (3, 2)
        if (plan := _split_plan_for_count(smooth, edge_ratio, height, page_count)) is not None
    ]
    if not plans:
        return [("page1", image)]
    _score, cuts = max(plans, key=lambda item: item[0])
    boundaries = [0, *cuts, height]
    return [
        (f"page{index + 1}", image.crop((0, boundaries[index], width, boundaries[index + 1])))
        for index in range(len(boundaries) - 1)
    ]


def make_photo_variant(
    image_path: Path,
    tmp_dir: Path,
    rotation: str,
    contrast: float,
    sharpness: float,
    page_split: str,
) -> list[tuple[str, Path]]:
    tmp_dir.mkdir(parents=True, exist_ok=True)
    image = ImageOps.exif_transpose(Image.open(image_path)).convert("RGB")
    variants: list[tuple[str, Path]] = []

    def save(name: str, img: Image.Image) -> None:
        out = tmp_dir / f"{image_path.stem}_tmp_{name}.jpg"
        img.save(out, quality=95)
        variants.append((name, out))

    pages = split_stacked_pages(image) if page_split == "auto" else [("page1", image)]
    multi_page = len(pages) > 1
    for page_name, page_image in pages:
        if rotation == "auto":
            rotate_name, page_image = auto_rotate_document(page_image)
        elif rotation == "cw":
            rotate_name, page_image = "rot90cw", page_image.rotate(-90, expand=True)
        elif rotation == "ccw":
            rotate_name, page_image = "rot90ccw", page_image.rotate(90, expand=True)
        elif rotation == "180":
            rotate_name, page_image = "rot180", page_image.rotate(180, expand=True)
        else:
            rotate_name = "none"

        enhanced = enhance_for_ocr(page_image, contrast, sharpness)
        page_part = f"_{page_name}" if multi_page else ""
        engine_part = "opencv_ocr" if cv2 is not None else "gray_contrast"
        save(f"photo{page_part}_{rotate_name}_{engine_part}", enhanced)
    return variants


def enhance_for_ocr(image: Image.Image, contrast: float, sharpness: float) -> Image.Image:
    if cv2 is None or np is None:
        gray = ImageOps.grayscale(image)
        enhanced = ImageEnhance.Contrast(gray).enhance(contrast)
        return ImageEnhance.Sharpness(enhanced).enhance(sharpness).convert("RGB")

    rgb = np.array(image.convert("RGB"))
    gray = cv2.cvtColor(rgb, cv2.COLOR_RGB2GRAY)
    denoised = cv2.fastNlMeansDenoising(gray, None, 8, 7, 21)
    clahe = cv2.createCLAHE(clipLimit=max(1.0, contrast), tileGridSize=(8, 8))
    enhanced = clahe.apply(denoised)
    if sharpness > 1:
        blurred = cv2.GaussianBlur(enhanced, (0, 0), 1.0)
        enhanced = cv2.addWeighted(enhanced, sharpness, blurred, 1 - sharpness, 0)
    return Image.fromarray(enhanced).convert("RGB")


def make_multi_variants(image_path: Path, tmp_dir: Path, contrast: float, sharpness: float) -> list[tuple[str, Path]]:
    tmp_dir.mkdir(parents=True, exist_ok=True)
    image = ImageOps.exif_transpose(Image.open(image_path)).convert("RGB")
    variants: list[tuple[str, Path]] = []

    def save(name: str, img: Image.Image) -> None:
        out = tmp_dir / f"{image_path.stem}_tmp_{name}.jpg"
        img.save(out, quality=95)
        variants.append((name, out))

    save("raw", image)
    save("rot90cw", image.rotate(-90, expand=True))
    save("rot90ccw", image.rotate(90, expand=True))
    save("rot180", image.rotate(180, expand=True))

    enhanced = enhance_for_ocr(image, contrast, sharpness)
    save("gray_contrast", enhanced)
    save("rot90cw_gray_contrast", enhance_for_ocr(image.rotate(-90, expand=True), contrast, sharpness))
    save("rot90ccw_gray_contrast", enhance_for_ocr(image.rotate(90, expand=True), contrast, sharpness))
    return variants


def extract_entries(page: Any) -> list[OcrEntry]:
    texts = page["rec_texts"]
    scores = page["rec_scores"]
    boxes = page["rec_boxes"]
    entries: list[OcrEntry] = []
    for text, score, box in zip(texts, scores, boxes):
        cleaned = str(text).strip()
        if not cleaned:
            continue
        b = [int(x) for x in box]
        entries.append(
            OcrEntry(
                text=cleaned,
                score=float(score),
                box=b,
                x=(b[0] + b[2]) / 2,
                y=(b[1] + b[3]) / 2,
            )
        )
    return entries


def is_good_enough_result(result: dict[str, Any]) -> bool:
    return (
        result["product_code_hits"] >= GOOD_ENOUGH_PRODUCT_CODE_HITS
        and result["key_hits"] >= GOOD_ENOUGH_KEY_HITS
        and result["line_count"] >= 10
    )


def is_bad_auto_rotation_result(result: dict[str, Any]) -> bool:
    try:
        rows, _issues = parse_result_rows(result)
    except Exception:
        return True
    anchors = estimate_product_anchor_count(result)
    complete_positive = sum(
        1
        for row in rows
        if row.quantity is not None
        and row.unit_cost is not None
        and row.amount is not None
        and row.quantity > 0
        and row.unit_cost > 0
        and row.amount > 0
        and normalize_ocr_name(row.name).upper() != row.vendor_code.upper()
    )
    return (
        result["key_hits"] == 0
        or not rows
        or complete_positive == 0
        or (anchors > 0 and len(rows) < max(1, anchors // 2))
    )


def page_number_pairs(entries: list[OcrEntry]) -> list[tuple[int, int]]:
    pairs: list[tuple[int, int]] = []
    for entry in entries:
        for current, total in PAGE_NUMBER_RE.findall(entry.text):
            current_number = int(current)
            total_number = int(total)
            if 1 <= current_number <= total_number <= 20:
                pairs.append((current_number, total_number))
    return pairs


def page_number_audit_issues(result: dict[str, Any]) -> list[str]:
    page_entries = result.get("page_entries") or [result["entries"]]
    pairs = [pair for entries in page_entries for pair in page_number_pairs(entries)]
    if not pairs:
        return []
    declared_total = max(total for _current, total in pairs)
    captured = sorted({current for current, total in pairs if total == declared_total})
    expected = list(range(1, declared_total + 1))
    issues: list[str] = []
    if captured != expected:
        missing = sorted(set(expected) - set(captured))
        issues.append(
            f"單據頁碼顯示共 {declared_total} 頁，目前辨識到 {captured or '無'}，"
            f"缺少頁碼 {missing or '無'}；請補齊同單號頁面後再核對總額。"
        )
    if len(page_entries) < len(captured):
        issues.append(
            f"圖片包含至少 {len(captured)} 個頁碼，但只切成 {len(page_entries)} 頁，可能有頁面未分割。"
        )
    return issues


def result_quality_metrics(result: dict[str, Any]) -> dict[str, int | float]:
    try:
        rows, _issues = parse_result_rows(result)
    except Exception:
        rows = []
    anchors = estimate_product_anchor_count(result)
    valid_rows = sum(1 for row in rows if row.check == "通過")
    complete_rows = sum(
        1
        for row in rows
        if row.quantity is not None
        and row.unit_cost is not None
        and row.amount is not None
        and row.quantity > 0
        and row.unit_cost > 0
        and row.amount > 0
    )
    named_rows = sum(
        1
        for row in rows
        if row.name
        and normalize_ocr_name(row.name).upper() != row.vendor_code.upper()
        and not PRODUCT_CODE_RE.fullmatch(normalize_ocr_name(row.name))
    )
    total = infer_result_total(result, rows)
    total_matches = 0
    if rows and total is not None and all(row.amount is not None and row.amount >= 0 for row in rows):
        row_total = sum(rounded_invoice_amount(float(row.amount)) or 0 for row in rows)
        total_matches = int(rounded_invoice_amount(row_total) == rounded_invoice_amount(total))
    header_company_hits = sum(
        1
        for entry in result["entries"]
        if any(term in entry.text for term in ("有限公司", "股份有限公司", "商行", "銷貨憑單", "出貨單", "訂單"))
    )
    page_penalty = len(page_number_audit_issues(result))
    negative_values = sum(
        1
        for row in rows
        for value in (row.quantity, row.unit_cost, row.amount)
        if value is not None and value < 0
    )
    return {
        "valid_rows": valid_rows,
        "complete_rows": complete_rows,
        "named_rows": named_rows,
        "row_count": len(rows),
        "anchor_count": anchors,
        "total_matches": total_matches,
        "header_company_hits": header_company_hits,
        "page_penalty": page_penalty,
        "negative_values": negative_values,
    }


def score_ocr_result(result: dict[str, Any]) -> tuple[int, int, int, int, int, int, int, int, int, float]:
    metrics = result_quality_metrics(result)
    return (
        int(metrics["complete_rows"]),
        int(metrics["valid_rows"]),
        int(metrics["named_rows"]),
        int(metrics["total_matches"]),
        min(int(metrics["row_count"]), int(metrics["anchor_count"]) or int(metrics["row_count"])),
        -abs(int(metrics["anchor_count"]) - int(metrics["row_count"])),
        -int(metrics["negative_values"]),
        -int(metrics["page_penalty"]),
        int(metrics["header_company_hits"]),
        result["key_hits"],
        result["avg_score"],
    )


def parse_result_rows(result: dict[str, Any]) -> tuple[list[ProductRow], list[str]]:
    page_entries = result.get("page_entries") or [result["entries"]]
    rows: list[ProductRow] = []
    issues: list[str] = []
    for page_index, entries in enumerate(page_entries, start=1):
        if is_yuanda_invoice(entries):
            page_rows, page_issues = parse_yuanda_rows(entries)
        elif is_liying_invoice(entries):
            page_rows, page_issues = parse_liying_rows(entries)
        else:
            page_rows, page_issues = parse_rows(entries)
        rows.extend(page_rows)
        issues.extend(f"第{page_index}頁：{issue}" for issue in page_issues)
    return rows, issues


def estimate_product_anchor_count(result: dict[str, Any]) -> int:
    page_entries = result.get("page_entries") or [result["entries"]]
    return sum(len(product_anchor_codes(entries)) for entries in page_entries)


def product_anchor_codes(entries: list[OcrEntry]) -> list[str]:
    if not entries:
        return []
    if is_yuanda_invoice(entries):
        width = max((entry.box[2] for entry in entries), default=1700)
        anchors = [
            match.group(1)
            for entry in entries
            if (match := YUANDA_ITEM_RE.match(entry.text)) and entry.x < width * 0.35
        ]
        return dedupe_nearby_anchor_codes(anchors)
    if is_liying_invoice(entries):
        anchors = [
            match.group(1)
            for entry in entries
            if (match := LIYING_ITEM_RE.match(entry.text.strip()))
        ]
        return dedupe_nearby_anchor_codes(anchors)

    _header_y, positions = header_positions(entries)
    code_left = positions["code"] - 140
    code_right = positions["name"] + 30
    anchors: list[str] = []
    for entry in sorted(entries, key=lambda item: (item.y, item.x)):
        match = PRODUCT_CODE_RE.match(entry.text)
        if not match:
            continue
        if not (code_left <= entry.box[0] <= code_right):
            continue
        code = match.group(1)
        if code.upper().startswith(("PH", "PK")) and "-" not in code:
            continue
        anchors.append(code)
    return dedupe_nearby_anchor_codes(anchors)


def is_yuanda_invoice(entries: list[OcrEntry]) -> bool:
    nearby_text = " ".join(entry.text for entry in entries[:60])
    if "元大玩具" in nearby_text:
        return True
    if "2913-8859" in nearby_text or "2914-6377" in nearby_text or "寶興路45巷3號2樓" in nearby_text:
        return True
    has_yuanda_amount_header = any("原幣金額" in entry.text for entry in entries)
    has_yuanda_total = any("原幣合計" in entry.text for entry in entries)
    has_9_digit_anchor = any(YUANDA_ITEM_RE.match(entry.text) for entry in entries)
    return has_yuanda_amount_header and has_yuanda_total and has_9_digit_anchor


def is_wanrong_invoice(entries: list[OcrEntry]) -> bool:
    text = " ".join(entry.text for entry in entries)
    if "萬榮" in text:
        return True
    has_wanrong_contact = "2836" in text or "士東路45" in text
    has_wanrong_layout = all(term in text for term in ("銷貨憑單", "售價", "折", "進價", "金額"))
    gsc_count = sum(1 for entry in entries if entry.text.strip().upper().startswith("GSC-"))
    has_pcs = any(is_unit_token(entry.text) for entry in entries)
    return bool(has_wanrong_contact or (has_wanrong_layout and has_pcs and gsc_count >= 2))


def is_liying_invoice(entries: list[OcrEntry]) -> bool:
    nearby_text = " ".join(entry.text for entry in entries)
    if "麗嬰" in nearby_text or "丽婴" in nearby_text:
        return True
    has_liying_headers = any("零售價" in entry.text for entry in entries) and any(
        "日幣" in entry.text or "廠價" in entry.text for entry in entries
    )
    return has_liying_headers and len(liying_item_entries(entries)) >= 1


def dedupe_nearby_anchor_codes(codes: list[str]) -> list[str]:
    deduped: list[str] = []
    for code in codes:
        normalized = code.strip().upper()
        if not normalized:
            continue
        if deduped and deduped[-1] == normalized:
            continue
        deduped.append(normalized)
    return deduped


def row_count_audit_issues(result: dict[str, Any], rows: list[ProductRow]) -> list[str]:
    issues: list[str] = page_number_audit_issues(result)
    expected = estimate_product_anchor_count(result)
    if expected > len(rows):
        issues.append(
            f"OCR 原始文字疑似有 {expected} 個商品列錨點，但輸出明細只有 {len(rows)} 筆，可能漏列；請查看 OCR測試紀錄原始文字。"
        )
    elif len(rows) > expected > 0:
        issues.append(
            f"輸出明細有 {len(rows)} 筆，但 OCR 商品列錨點只有 {expected} 個，可能誤把表頭或續行當商品。"
        )

    for row in rows:
        if row.name.upper() == row.vendor_code.upper() or PRODUCT_CODE_RE.fullmatch(row.name):
            issues.append(f"{row.vendor_code}: 品名等於廠商貨號，禁止自動繼續。")
        for field_name, value in (
            ("數量", row.quantity),
            ("進價", row.unit_cost),
            ("金額", row.amount),
        ):
            if value is not None and value < 0:
                issues.append(f"{row.vendor_code}: {field_name}為負數 {value}，禁止自動繼續。")

    invoice_total = infer_result_total(result, rows)
    missing_amount_count = sum(row.amount is None for row in rows)
    if rows and missing_amount_count:
        issues.append(f"有 {missing_amount_count} 筆商品金額缺漏，無法核對商品合計與單據總額。")
    elif rows and invoice_total is not None:
        rounded_rows_total = sum(rounded_invoice_amount(float(row.amount)) or 0 for row in rows)
        if rounded_invoice_amount(rounded_rows_total) != rounded_invoice_amount(invoice_total):
            issues.append(
                "商品列金額合計與單據總額不符："
                f"商品列合計 {rounded_rows_total}，單據總額 {compact_amount(invoice_total)}；請查看 OCR測試紀錄原始文字。"
            )
    elif rows:
        issues.append("最末頁未穩定辨識到總計／合計標籤與右側金額，無法核對單據總額。")
    return list(dict.fromkeys(issues))


def auto_align_numeric_columns(
    rows: list[ProductRow],
    entries: list[OcrEntry],
    code_entries: list[tuple[OcrEntry, str, str]],
    positions: dict[str, float],
    header_y: float,
    total_y: float,
    typical_spacing: float,
) -> list[ProductRow]:
    row_count = len(code_entries)
    if row_count < 2 or len(rows) != row_count:
        return rows

    quantity_left = positions["quantity"] - 120
    quantity_right = positions.get("unit", positions["quantity"] + 90) + 20
    discount_x = positions.get("discount")
    if discount_x is not None and positions["quantity"] < discount_x < positions["unit_cost"]:
        unit_cost_left = (discount_x + positions["unit_cost"]) / 2
    else:
        unit_cost_left = min(positions["unit_cost"], positions["quantity"]) - 70
    unit_cost_right = positions["amount"] - max(20, (positions["amount"] - positions["unit_cost"]) * 0.25)
    amount_left = positions["amount"] - max(20, (positions["amount"] - positions["unit_cost"]) * 0.25)
    detail_top = max(header_y, code_entries[0][0].y - max(50, typical_spacing * 1.25))
    detail_bottom = min(total_y, code_entries[-1][0].y + max(60, typical_spacing * 2.20))

    quantity_candidates: list[tuple[OcrEntry, int]] = []
    unit_cost_candidates: list[tuple[OcrEntry, float]] = []
    amount_candidates: list[tuple[OcrEntry, float]] = []
    for entry in entries:
        if entry.y <= detail_top or entry.y >= detail_bottom or is_unit_token(entry.text):
            continue
        if quantity_left <= entry.x < quantity_right:
            quantity = extract_quantity(entry.text)
            if quantity is not None:
                quantity_candidates.append((entry, quantity))
        if unit_cost_left <= entry.x < unit_cost_right:
            unit_cost = extract_decimal_money(entry.text)
            if unit_cost is not None:
                unit_cost_candidates.append((entry, unit_cost))
        if entry.x >= amount_left:
            amount = extract_amount_integer(entry.text)
            if amount is not None:
                amount_candidates.append((entry, amount))

    if len(unit_cost_candidates) != row_count or len(amount_candidates) != row_count:
        return rows

    ordered_unit_costs = [value for _entry, value in sorted(unit_cost_candidates, key=lambda item: item[0].y)]
    ordered_amounts = [value for _entry, value in sorted(amount_candidates, key=lambda item: item[0].y)]

    quantity_by_row: dict[int, tuple[int, float]] = {}
    max_distance = max(45, typical_spacing * 0.90)
    for entry, quantity in quantity_candidates:
        nearest_index = min(range(row_count), key=lambda index: abs(entry.y - code_entries[index][0].y))
        distance = abs(entry.y - code_entries[nearest_index][0].y)
        if distance > max_distance:
            continue
        existing = quantity_by_row.get(nearest_index)
        if existing is None or distance < existing[1]:
            quantity_by_row[nearest_index] = (quantity, distance)

    for index, (unit_cost, amount) in enumerate(zip(ordered_unit_costs, ordered_amounts)):
        if index in quantity_by_row:
            continue
        inferred = inferred_unit_cost(unit_cost, amount)
        if inferred is None or abs(inferred - round(inferred)) > 0.001:
            continue
        inferred_quantity = int(round(inferred))
        if invoice_amount_matches(inferred_quantity, unit_cost, amount):
            quantity_by_row[index] = (inferred_quantity, 0.0)

    total_quantity = infer_total_quantity(entries)
    if len(quantity_by_row) == row_count - 1 and total_quantity is not None:
        missing_indexes = [index for index in range(row_count) if index not in quantity_by_row]
        inferred_quantity = total_quantity - sum(quantity for quantity, _distance in quantity_by_row.values())
        if len(missing_indexes) == 1 and inferred_quantity > 0:
            quantity_by_row[missing_indexes[0]] = (inferred_quantity, 0.0)

    if len(quantity_by_row) != row_count:
        return rows

    ordered_quantities = [quantity_by_row[index][0] for index in range(row_count)]
    if total_quantity is not None and sum(ordered_quantities) != total_quantity:
        return rows
    if not all(
        invoice_amount_matches(quantity, unit_cost, amount)
        for quantity, unit_cost, amount in zip(ordered_quantities, ordered_unit_costs, ordered_amounts)
    ):
        return rows

    total_amount = infer_total(entries, [])
    if total_amount is None:
        return rows
    amount_sum = sum(ordered_amounts)
    if rounded_invoice_amount(amount_sum) != rounded_invoice_amount(total_amount):
        return rows

    current_clean_rows = sum(1 for row in rows if row.check == "通過" and row.issue == "無")
    differs = any(
        row.quantity != quantity
        or row.unit_cost is None
        or abs(float(row.unit_cost) - unit_cost) > 0.001
        or row.amount is None
        or abs(float(row.amount) - amount) > 0.001
        for row, quantity, unit_cost, amount in zip(rows, ordered_quantities, ordered_unit_costs, ordered_amounts)
    )
    if not differs and current_clean_rows == row_count:
        return rows

    audit_note = "欄位順序自動修正：依數量、進價、金額欄由上到下對齊，且逐列金額與總額核對通過"
    corrected: list[ProductRow] = []
    for index, (row, quantity, unit_cost, amount) in enumerate(
        zip(rows, ordered_quantities, ordered_unit_costs, ordered_amounts)
    ):
        issue_parts = suspicious_name_issues(row.name)
        corrected.append(
            ProductRow(
                vendor_code=row.vendor_code,
                name=row.name,
                quantity=quantity,
                unit_cost=compact_amount(unit_cost),
                amount=compact_amount(amount),
                check="通過",
                issue="；".join(issue_parts) if issue_parts else "無",
                audit_note=audit_note if index == 0 else "",
            )
        )
    return corrected


def merge_page_entries(page_entries: list[list[OcrEntry]]) -> list[OcrEntry]:
    merged: list[OcrEntry] = []
    y_offset = 0
    for entries in page_entries:
        page_bottom = 0
        for entry in entries:
            shifted_box = [entry.box[0], entry.box[1] + y_offset, entry.box[2], entry.box[3] + y_offset]
            merged.append(
                OcrEntry(
                    text=entry.text,
                    score=entry.score,
                    box=shifted_box,
                    x=entry.x,
                    y=entry.y + y_offset,
                )
            )
            page_bottom = max(page_bottom, entry.box[3])
        y_offset += page_bottom + 80
    return merged


def opposite_rotation_from_variant(variant: str) -> str | None:
    if "rot90cw" in variant:
        return "ccw"
    if "rot90ccw" in variant:
        return "cw"
    if "rot180" in variant:
        return "none"
    if "none" in variant:
        return "cw"
    return None


def create_ocr() -> PaddleOCR:
    return PaddleOCR(
        lang="ch",
        use_doc_orientation_classify=False,
        use_doc_unwarping=False,
        use_textline_orientation=False,
    )


def run_ocr(
    image_path: Path,
    tmp_dir: Path,
    rotation: str,
    multi_variant: bool,
    contrast: float,
    sharpness: float,
    page_split: str,
) -> tuple[dict[str, Any], list[dict[str, Any]], dict[str, float]]:
    timings: dict[str, float] = {}
    start = time.perf_counter()
    ocr = create_ocr()
    timings["model_load_seconds"] = round(time.perf_counter() - start, 3)

    results: list[dict[str, Any]] = []
    variant_batches: list[list[tuple[str, Path]]] = []
    if multi_variant:
        variant_batches.append(make_multi_variants(image_path, tmp_dir, contrast, sharpness))
    else:
        variant_batches.append(make_photo_variant(image_path, tmp_dir, rotation, contrast, sharpness, page_split))

    for batch_index, variants in enumerate(variant_batches):
        if not multi_variant and len(variants) > 1:
            page_entries: list[list[OcrEntry]] = []
            page_paths: list[str] = []
            page_names: list[str] = []
            predict_seconds = 0.0
            for name, variant_path in variants:
                predict_start = time.perf_counter()
                pages = ocr.predict(str(variant_path))
                predict_seconds += time.perf_counter() - predict_start
                entries = extract_entries(pages[0]) if pages else []
                page_entries.append(entries)
                page_paths.append(str(variant_path))
                page_names.append(name)
            entries = merge_page_entries(page_entries)
            key_hits = sum(1 for item in entries if any(term in item.text for term in KEY_TERMS))
            product_code_hits = sum(1 for item in entries if PRODUCT_CODE_RE.match(item.text))
            avg_score = round(sum(item.score for item in entries) / len(entries), 4) if entries else 0
            results.append(
                {
                    "variant": f"split_{len(variants)}pages:" + ",".join(page_names),
                    "path": "\n".join(page_paths),
                    "line_count": len(entries),
                    "avg_score": avg_score,
                    "key_hits": key_hits,
                    "product_code_hits": product_code_hits,
                    "predict_seconds": round(predict_seconds, 3),
                    "entries": entries,
                    "page_entries": page_entries,
                }
            )
        else:
            for name, variant_path in variants:
                predict_start = time.perf_counter()
                pages = ocr.predict(str(variant_path))
                predict_seconds = round(time.perf_counter() - predict_start, 3)
                entries = extract_entries(pages[0]) if pages else []
                key_hits = sum(1 for item in entries if any(term in item.text for term in KEY_TERMS))
                product_code_hits = sum(1 for item in entries if PRODUCT_CODE_RE.match(item.text))
                avg_score = round(sum(item.score for item in entries) / len(entries), 4) if entries else 0
                result = {
                    "variant": name,
                    "path": str(variant_path),
                    "line_count": len(entries),
                    "avg_score": avg_score,
                    "key_hits": key_hits,
                    "product_code_hits": product_code_hits,
                    "predict_seconds": predict_seconds,
                    "entries": entries,
                }
                results.append(result)
                if multi_variant and is_good_enough_result(result):
                    break
        if multi_variant:
            break

        if rotation == "auto" and batch_index == 0 and results and is_bad_auto_rotation_result(results[-1]):
            retry_rotation = opposite_rotation_from_variant(results[-1]["variant"])
            if retry_rotation:
                variant_batches.append(make_photo_variant(image_path, tmp_dir, retry_rotation, contrast, sharpness, page_split))
                continue
        break
    best = sorted(
        results,
        key=score_ocr_result,
        reverse=True,
    )[0]
    timings["total_ocr_seconds"] = round(time.perf_counter() - start, 3)
    return best, results, timings


def load_settings(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    with path.open("r", encoding="utf-8-sig") as f:
        data = json.load(f)
    return data if isinstance(data, dict) else {}


def infer_vendor(entries: list[OcrEntry]) -> str:
    known_vendors = ["南波", "萬榮", "麗嬰", "鉅霖", "元大玩具", "金亞特"]
    for vendor in known_vendors:
        if any(vendor in entry.text for entry in entries):
            if vendor == "麗嬰":
                return "麗嬰國際股份有限公司"
            if vendor == "元大玩具":
                return "元大玩具股份有限公司"
            if vendor == "金亞特":
                return "金亞特創藝玩具股份有限公司"
            for entry in entries:
                if vendor in entry.text:
                    return entry.text
    if is_liying_invoice(entries):
        return "麗嬰國際股份有限公司"
    if is_yuanda_invoice(entries):
        return "元大玩具股份有限公司"
    if is_wanrong_invoice(entries):
        return "萬榮國際企業股份有限公司"
    for entry in entries[:20]:
        if "公司" in entry.text or "商行" in entry.text:
            return entry.text
    return "未辨識廠商"


def header_positions(entries: list[OcrEntry]) -> tuple[float, dict[str, float]]:
    positions: dict[str, list[float]] = {
        "code": [],
        "name": [],
        "quantity": [],
        "unit": [],
        "list_price": [],
        "discount": [],
        "unit_cost": [],
        "amount": [],
    }
    header_y_values = []
    width = max((entry.box[2] for entry in entries), default=1200)
    for entry in entries:
        text = entry.text
        matched = False
        if "貨號" in text or "品號" in text:
            positions["code"].append(entry.x)
            matched = True
        if "品名" in text:
            positions["name"].append(entry.x)
            matched = True
        if "數量" in text or "数量" in text:
            positions["quantity"].append(entry.x)
            matched = True
        if "單位" in text:
            positions["unit"].append(entry.x)
            matched = True
        if "售價" in text:
            positions["list_price"].append(entry.x)
            matched = True
        if "折" in text:
            positions["discount"].append(entry.x)
            matched = True
        is_factory_price_label = "日幣" in text or "廠價" in text or "厂价" in text
        if "進價" in text or "單價" in text or (
            "價" in text and "售價" not in text and not is_factory_price_label and entry.x > width * 0.55
        ):
            positions["unit_cost"].append(entry.x)
            matched = True
        is_total_label = any(term in text for term in ["總金額", "总金额", "總額", "总额", "合計"])
        if not is_total_label and ("金額" in text or "金额" in text or ("金" in text and entry.x > width * 0.65)):
            positions["amount"].append(entry.x)
            matched = True
        if matched:
            header_y_values.append(entry.y)
    defaults = {
        "code": width * 0.14,
        "name": width * 0.36,
        "quantity": width * 0.55,
        "unit": width * 0.62,
        "list_price": width * 0.70,
        "discount": width * 0.76,
        "unit_cost": width * 0.82,
        "amount": width * 0.93,
    }
    resolved = {key: (median(value) if value else defaults[key]) for key, value in positions.items()}
    header_y = max(header_y_values) if header_y_values else min((entry.y for entry in entries), default=0)
    return header_y, resolved


def nearest_column(entry: OcrEntry, positions: dict[str, float]) -> str:
    return min(positions, key=lambda key: abs(entry.x - positions[key]))


def liying_item_entries(entries: list[OcrEntry]) -> list[OcrEntry]:
    return sorted(
        [entry for entry in entries if LIYING_ITEM_RE.match(entry.text.strip())],
        key=lambda entry: (entry.y, entry.x),
    )


def is_noise_or_header_text(text: str) -> bool:
    stripped = text.strip()
    if not stripped:
        return True
    if stripped in UNIT_TOKENS:
        return True
    header_terms = [
        "貨號",
        "客戶品號",
        "條碼",
        "品名規格",
        "品名规格",
        "商品名稱",
        "商品名称",
        "產品名稱",
        "产品名称",
        "零售價",
        "數量",
        "数量",
        "單價",
        "金額",
        "日幣",
        "廠價",
        "備註",
    ]
    return any(term in stripped for term in header_terms)


def is_barcode_text(text: str) -> bool:
    return bool(re.fullmatch(r"\d{12,14}", text.strip()))


def is_plain_number_text(text: str) -> bool:
    return bool(re.fullmatch(r"\s*-?\d+(?:,\d{3})*(?:\.\d+)?\s*", text))


def nearest_value_by_x(
    entries: list[OcrEntry],
    target_x: float,
    parser: Any,
    max_dx: float,
    validator: Any | None = None,
) -> Any | None:
    candidates = []
    for entry in entries:
        value = parser(entry.text)
        if value is None:
            continue
        if validator is not None and not validator(value):
            continue
        dx = abs(entry.x - target_x)
        if dx <= max_dx:
            candidates.append((dx, entry.y, value))
    return min(candidates, key=lambda item: (item[0], item[1]))[2] if candidates else None


def nearest_value_by_y(
    entries: list[OcrEntry],
    target_y: float,
    parser: Any,
    max_dy: float,
    validator: Any | None = None,
) -> Any | None:
    candidates = []
    for entry in entries:
        value = parser(entry.text)
        if value is None:
            continue
        if validator is not None and not validator(value):
            continue
        dy = abs(entry.y - target_y)
        if dy <= max_dy:
            candidates.append((dy, entry.x, value))
    return min(candidates, key=lambda item: (item[0], item[1]))[2] if candidates else None


def ordered_values_by_x(
    entries: list[OcrEntry],
    target_x: float,
    parser: Any,
    max_dx: float,
    validator: Any | None = None,
) -> list[Any]:
    candidates = []
    for entry in entries:
        value = parser(entry.text)
        if value is None:
            continue
        if validator is not None and not validator(value):
            continue
        dx = abs(entry.x - target_x)
        if dx <= max_dx:
            candidates.append((entry.y, entry.x, value))
    return [value for _y, _x, value in sorted(candidates, key=lambda item: (item[0], item[1]))]


def liying_horizontal_column_sequences(
    entries: list[OcrEntry],
    positions: dict[str, float],
    row_count: int,
) -> dict[str, list[Any]]:
    sequence_specs = {
        "retail": (positions["list_price"], extract_quantity, 80, lambda value: 50 <= value <= 999),
        "quantity": (positions["quantity"], extract_quantity, 40, lambda value: 1 <= value <= 999),
        "unit_cost": (positions["unit_cost"], extract_decimal_money, 75, lambda value: value > 0),
        "amount": (positions["amount"], extract_integer_money, 80, lambda value: value > 0),
    }
    sequences: dict[str, list[Any]] = {}
    for key, (target_x, parser, max_dx, validator) in sequence_specs.items():
        values = ordered_values_by_x(entries, target_x, parser, max_dx, validator)
        if len(values) == row_count:
            sequences[key] = values
    return sequences


def infer_liying_retail(unit_cost: float | None, retail: int | None) -> int | None:
    valid_prices = {150, 195, 250, 280, 295}
    if retail in valid_prices:
        return retail
    if unit_cost is None:
        return retail
    by_cost = {
        99.45: 150,
        129.29: 195,
        165.75: 250,
        185.64: 280,
        195.59: 295,
    }
    nearest_cost, inferred = min(by_cost.items(), key=lambda item: abs(item[0] - unit_cost))
    if abs(nearest_cost - unit_cost) <= 0.08:
        return inferred
    return retail


def normalize_liying_name(name: str) -> str:
    normalized = normalize_ocr_name(name).replace("_日產", " 日產")
    normalized = normalized.replace("噴火龍Y公 存", "噴火龍Y公仔")
    normalized = normalized.replace("噴火龍Y公存", "噴火龍Y公仔")
    marker = re.search(r"(?:亞版#\d+_\d+|#\d+_\d+|#PRM\d+|A[O0]-\d+|UTR\d+)", normalized)
    if marker and marker.start() > 0:
        normalized = f"{normalized[marker.start():]} {normalized[:marker.start()]}".strip()
    normalized = normalized.replace("A0-", "AO-")
    normalized = re.sub(r"#(\d{3})\s+(\d{6})", r"#\1_\2", normalized)
    normalized = re.sub(r"#066\s+650形\s+102557\s+廣島電鐵", "#066_102557 廣島電鐵650形", normalized)
    normalized = normalized.replace("#PRM28 航空自衛隊F-", "#PRM28 航空自衛隊F-35")
    normalized = normalized.replace("#139_798323 動物運輸", "#139_798323 動物運輸車")
    normalized = normalized.replace("#PRM32 本田NSX TYPE", "#PRM32 本田NSX TYPE S")
    normalized = re.sub(r"(?<=[\u4e00-\u9fff])\s+(?=[\u4e00-\u9fff])", "", normalized)
    normalized = re.sub(r"\s+", " ", normalized)
    return normalized.strip()


def validate_liying_numbers(
    quantity: int | None,
    unit_cost: float | None,
    amount: float | None,
    issue_parts: list[str],
) -> tuple[float | None, str]:
    if quantity is None or unit_cost is None:
        return amount, "需確認"
    expected = rounded_invoice_amount(quantity * unit_cost)
    if amount is None:
        issue_parts.append(f"金額由數量×進價四捨五入為 {expected}，請確認")
        return float(expected), "通過"
    if expected is not None and abs(amount - expected) <= 1:
        return amount, "通過"
    issue_parts.append(f"金額 OCR {amount:g} 與數量×進價四捨五入 {expected} 不符，改用 {expected}")
    return float(expected) if expected is not None else amount, "通過"


def liying_rotated_positions(entries: list[OcrEntry]) -> dict[str, float]:
    values: dict[str, list[float]] = {
        "name": [],
        "retail": [],
        "quantity": [],
        "unit_cost": [],
        "amount": [],
        "factory_price": [],
    }
    for entry in entries:
        text = entry.text
        if "產品名稱" in text:
            values["name"].append(entry.y)
        if "零售價" in text:
            values["retail"].append(entry.y)
        if "數量" in text or "数量" in text:
            values["quantity"].append(entry.y)
        if "單價" in text:
            values["unit_cost"].append(entry.y)
        if "金額" in text:
            values["amount"].append(entry.y)
        if "日幣" in text or "廠價" in text:
            values["factory_price"].append(entry.y)
    defaults = {
        "name": 340.0,
        "retail": 465.0,
        "quantity": 530.0,
        "unit_cost": 600.0,
        "amount": 670.0,
        "factory_price": 735.0,
    }
    return {key: (median(value) if value else defaults[key]) for key, value in values.items()}


def parse_liying_rows(entries: list[OcrEntry]) -> tuple[list[ProductRow], list[str]]:
    anchors = liying_item_entries(entries)
    if not anchors:
        return [], ["麗嬰單據未辨識到 TM 貨號商品列。"]
    x_range = max(entry.x for entry in anchors) - min(entry.x for entry in anchors)
    y_range = max(entry.y for entry in anchors) - min(entry.y for entry in anchors)
    if len(anchors) > 1 and x_range > max(80, y_range * 1.4):
        return parse_liying_rotated_rows(entries, anchors)
    return parse_liying_horizontal_rows(entries, anchors)


def liying_name_row_entries(
    entries: list[OcrEntry],
    anchor: OcrEntry,
    row_start: float,
    next_anchor: OcrEntry | None,
    typical_spacing: float,
) -> list[OcrEntry]:
    # A wrapped product-name line can sit exactly at the midpoint between two
    # item-code centres.  Centre-to-centre row bands therefore move that line
    # into the next product and create a cascading name shift.  Product names
    # use the printed item-code top edges as their independent row boundary;
    # numeric columns keep the existing centre bands and sequence validation.
    name_start = max(row_start, float(anchor.box[1]))
    name_end = (
        next_anchor.box[1]
        if next_anchor is not None
        else anchor.y + typical_spacing * 1.8
    )
    return [
        entry
        for entry in entries
        if name_start < entry.y < name_end
    ]


def parse_liying_horizontal_rows(entries: list[OcrEntry], anchors: list[OcrEntry]) -> tuple[list[ProductRow], list[str]]:
    _header_y, positions = header_positions(entries)
    anchors = sorted(anchors, key=lambda entry: entry.y)
    spacings = [anchors[i + 1].y - anchors[i].y for i in range(len(anchors) - 1) if anchors[i + 1].y > anchors[i].y]
    typical_spacing = median(spacings) if spacings else 45
    table_top = min(anchor.box[1] for anchor in anchors) - typical_spacing
    table_bottom = max(anchor.box[3] for anchor in anchors) + typical_spacing
    table_entries = [entry for entry in entries if table_top <= entry.y <= table_bottom]
    column_sequences = liying_horizontal_column_sequences(table_entries, positions, len(anchors))
    rows: list[ProductRow] = []
    for idx, anchor in enumerate(anchors):
        previous_anchor = anchors[idx - 1] if idx > 0 else None
        next_anchor = anchors[idx + 1] if idx + 1 < len(anchors) else None
        row_start = previous_anchor.y + (anchor.y - previous_anchor.y) * 0.50 if previous_anchor else anchor.box[1] - 8
        row_end = anchor.y + (next_anchor.y - anchor.y) * 0.50 if next_anchor else anchor.y + typical_spacing * 1.8
        row_entries = [entry for entry in entries if row_start <= entry.y < row_end]
        name_row_entries = liying_name_row_entries(
            entries,
            anchor,
            row_start,
            next_anchor,
            typical_spacing,
        )
        name_left = positions["name"] - 90
        name_right = positions["list_price"] - 12
        name_parts = [
            entry.text
            for entry in sorted(name_row_entries, key=lambda item: (item.y, item.x))
            if name_left <= entry.x < name_right
            and entry is not anchor
            and not is_noise_or_header_text(entry.text)
            and not is_barcode_text(entry.text)
            and not is_plain_number_text(entry.text)
        ]
        name = normalize_liying_name(" ".join(name_parts)) or anchor.text
        issue_parts = suspicious_name_issues(name)
        retail = column_sequences.get("retail", [None] * len(anchors))[idx]
        if retail is None:
            retail = nearest_value_by_x(
                row_entries,
                positions["list_price"],
                extract_quantity,
                80,
                lambda value: 50 <= value <= 999,
            )
        quantity = column_sequences.get("quantity", [None] * len(anchors))[idx]
        if quantity is None:
            quantity = nearest_value_by_x(
                row_entries,
                positions["quantity"],
                extract_quantity,
                45,
                lambda value: 1 <= value <= 999,
            )
        unit_cost = column_sequences.get("unit_cost", [None] * len(anchors))[idx]
        if unit_cost is None:
            unit_cost = nearest_value_by_x(row_entries, positions["unit_cost"], extract_decimal_money, 95)
        amount = column_sequences.get("amount", [None] * len(anchors))[idx]
        if amount is None:
            amount = nearest_value_by_x(row_entries, positions["amount"], extract_integer_money, 95)
        retail = infer_liying_retail(unit_cost, retail)
        amount, check = validate_liying_numbers(quantity, unit_cost, amount, issue_parts)
        if retail is None:
            issue_parts.append("零售價未穩定辨識")
        if quantity is None:
            issue_parts.append("數量未穩定辨識")
        if unit_cost is None:
            issue_parts.append("進價未穩定辨識")
        rows.append(
            ProductRow(
                vendor_code=anchor.text.strip(),
                name=name,
                quantity=quantity,
                unit_cost=compact_amount(unit_cost),
                amount=compact_amount(amount),
                check=check,
                issue="；".join(issue_parts) if issue_parts else "無",
                retail_price=retail,
            )
        )
    return rows, []


def parse_liying_rotated_rows(entries: list[OcrEntry], anchors: list[OcrEntry]) -> tuple[list[ProductRow], list[str]]:
    positions = liying_rotated_positions(entries)
    anchors = sorted(anchors, key=lambda entry: entry.x)
    rows_by_x: list[ProductRow] = []
    for idx, anchor in enumerate(anchors):
        left = (anchors[idx - 1].x + anchor.x) / 2 if idx > 0 else anchor.x - 28
        right = (anchor.x + anchors[idx + 1].x) / 2 if idx + 1 < len(anchors) else anchor.x + 28
        name_column_entries = [entry for entry in entries if left <= entry.x < right]
        column_entries = [entry for entry in entries if left - 6 <= entry.x < right + 6]
        name_top = positions["name"] - 80
        name_bottom = positions["retail"] - 10
        name_parts = [
            entry
            for entry in name_column_entries
            if name_top <= entry.y < name_bottom
            and entry is not anchor
            and entry.score >= 0.55
            and not is_noise_or_header_text(entry.text)
            and not is_barcode_text(entry.text)
            and not is_plain_number_text(entry.text)
        ]
        ordered_name_parts = sorted(name_parts, key=lambda item: (round(item.y / 8), -item.x))
        name = normalize_liying_name(" ".join(entry.text for entry in ordered_name_parts)) or anchor.text
        issue_parts = suspicious_name_issues(name)
        retail_y = (positions["retail"] + positions["quantity"]) / 2
        quantity_y = positions["quantity"] + (positions["unit_cost"] - positions["quantity"]) * 0.55
        unit_cost_y = (positions["unit_cost"] + positions["amount"]) / 2
        amount_y = (positions["amount"] + positions["factory_price"]) / 2
        retail = nearest_value_by_y(
            column_entries,
            retail_y,
            extract_quantity,
            55,
            lambda value: 50 <= value <= 999,
        )
        quantity = nearest_value_by_y(
            column_entries,
            quantity_y,
            extract_quantity,
            45,
            lambda value: 1 <= value <= 999,
        )
        unit_cost = nearest_value_by_y(column_entries, unit_cost_y, extract_decimal_money, 60)
        amount = nearest_value_by_y(column_entries, amount_y, extract_integer_money, 75)
        retail = infer_liying_retail(unit_cost, retail)
        amount, check = validate_liying_numbers(quantity, unit_cost, amount, issue_parts)
        if retail is None:
            issue_parts.append("零售價未穩定辨識")
        if quantity is None:
            issue_parts.append("數量未穩定辨識")
        if unit_cost is None:
            issue_parts.append("進價未穩定辨識")
        rows_by_x.append(
            ProductRow(
                vendor_code=anchor.text.strip(),
                name=name,
                quantity=quantity,
                unit_cost=compact_amount(unit_cost),
                amount=compact_amount(amount),
                check=check,
                issue="；".join(issue_parts) if issue_parts else "無",
                retail_price=retail,
            )
        )
    return list(reversed(rows_by_x)), []


def fill_yuanda_missing_amounts_when_total_matches(
    rows: list[ProductRow],
    entries: list[OcrEntry],
) -> list[ProductRow]:
    if not rows or any("贈" in row.name for row in rows):
        return rows
    if any(
        row.quantity is None
        or row.quantity <= 0
        or row.unit_cost is None
        or row.unit_cost <= 0
        for row in rows
    ):
        return rows
    total = infer_total(entries, rows)
    if total is None:
        return rows
    candidate_amounts = [
        float(row.amount)
        if row.amount is not None
        else float(rounded_invoice_amount(float(row.quantity) * float(row.unit_cost)) or 0)
        for row in rows
    ]
    if rounded_invoice_amount(sum(candidate_amounts)) != rounded_invoice_amount(total):
        return rows

    corrected: list[ProductRow] = []
    for row, amount in zip(rows, candidate_amounts):
        if row.amount is not None:
            corrected.append(row)
            continue
        issue_parts = [
            issue
            for issue in row.issue.split("；")
            if issue and issue != "無" and issue != "金額未穩定辨識"
        ]
        corrected.append(
            ProductRow(
                vendor_code=row.vendor_code,
                name=row.name,
                quantity=row.quantity,
                unit_cost=row.unit_cost,
                amount=compact_amount(amount),
                check="通過",
                issue="；".join(issue_parts) if issue_parts else "無",
                retail_price=row.retail_price,
                audit_note="元大金額安全補值：數量×進價，且全列計算合計與原幣合計一致",
            )
        )
    return corrected


def parse_yuanda_rows(entries: list[OcrEntry]) -> tuple[list[ProductRow], list[str]]:
    width = max((entry.box[2] for entry in entries), default=1700)
    anchors: list[tuple[OcrEntry, str]] = []
    for entry in entries:
        match = YUANDA_ITEM_RE.match(entry.text)
        if match and entry.x < width * 0.35:
            anchors.append((entry, match.group(1)))
    anchors.sort(key=lambda item: item[0].y)
    if not anchors:
        return [], ["元大單據未辨識到 9 位品號商品列。"]

    first_y = anchors[0][0].y

    def header_x(terms: tuple[str, ...], default_ratio: float) -> float:
        matches = [
            entry.x
            for entry in entries
            if entry.y < first_y and any(term in entry.text for term in terms)
        ]
        return median(matches) if matches else width * default_ratio

    quantity_x = header_x(("數量", "数量"), 0.69)
    unit_cost_x = header_x(("單價",), 0.78)
    amount_x = header_x(("原幣金額", "金額", "金额"), 0.90)
    unit_amount_boundary = (unit_cost_x + amount_x) / 2
    stop_entries = [
        entry
        for entry in entries
        if any(term in entry.text for term in ("以下空白", "原幣未稅", "原幣合計", "本幣合計"))
    ]

    rows: list[ProductRow] = []
    global_issues: list[str] = []
    for index, (anchor, code) in enumerate(anchors):
        previous_anchor_y = anchors[index - 1][0].y if index > 0 else None
        next_anchor_y = anchors[index + 1][0].y if index + 1 < len(anchors) else None
        next_stop_y = min((entry.y for entry in stop_entries if entry.y > anchor.y), default=float("inf"))
        # YuanDa prints the first name line slightly above or below the 9-digit
        # item anchor, while the second name line can sit close to the next item.
        # Split rows at the midpoint between adjacent anchor centres.  Using the
        # next anchor's top edge makes the next product's first line leak into
        # the previous product and also lets bag-count values win over quantity.
        row_start_y = (
            (previous_anchor_y + anchor.y) / 2
            if previous_anchor_y is not None
            else anchor.y - 30
        )
        row_end_y = (
            (anchor.y + next_anchor_y) / 2
            if next_anchor_y is not None
            else anchor.y + 100
        )
        row_end_y = min(row_end_y, next_stop_y)
        row_entries = [entry for entry in entries if row_start_y <= entry.y < row_end_y]

        # Names use a slightly different boundary from numeric columns.  On a
        # skewed YuanDa printout, amount/quantity values align best to anchor
        # midpoints, but a wrapped product name belongs between consecutive
        # 9-digit anchor top edges when compared by OCR box centre.
        name_start_y = anchor.box[1]
        name_end_y = (
            anchors[index + 1][0].box[1]
            if index + 1 < len(anchors)
            else min(next_stop_y, anchor.y + 100)
        )
        name_parts: list[str] = []
        for entry in sorted(entries, key=lambda item: (item.y, item.x)):
            if not (name_start_y <= entry.y < name_end_y):
                continue
            if not (width * 0.28 <= entry.x <= width * 0.54):
                continue
            text = entry.text.strip()
            if entry.box[2] > width * 0.46:
                text = re.sub(r"(?<=[\u4e00-\u9fff])\s*\d{1,4}$", "", text)
            if not text or re.fullmatch(r"[\d\s,.\-]+", text):
                continue
            if YUANDA_ITEM_RE.match(text) or PRODUCT_CODE_RE.match(text):
                continue
            if any(term in text for term in ("品名", "規格", "以下空白", "台幣", "外幣")):
                continue
            name_parts.append(text)

        def nearest_value(
            target_x: float,
            max_distance: float,
            parser: Any,
            min_x: float | None = None,
            max_x: float | None = None,
        ) -> float | None:
            candidates: list[tuple[float, float, float]] = []
            for entry in row_entries:
                if min_x is not None and entry.x < min_x:
                    continue
                if max_x is not None and entry.x > max_x:
                    continue
                if abs(entry.x - target_x) > max_distance:
                    continue
                value = parser(entry.text)
                if value is not None:
                    candidates.append((abs(entry.x - target_x), abs(entry.y - anchor.y), float(value)))
            # Adjacent rows can share nearly identical x coordinates. Prefer the
            # value nearest this row vertically, then use x distance as a tie-break.
            return min(candidates, key=lambda item: (item[1], item[0]))[2] if candidates else None

        quantity_value = nearest_value(
            quantity_x,
            max(90, width * 0.06),
            extract_quantity,
            min_x=quantity_x - max(60, width * 0.04),
            max_x=(quantity_x + unit_cost_x) / 2,
        )
        unit_cost = nearest_value(unit_cost_x, max(100, width * 0.07), extract_decimal_money)
        if unit_cost is None:
            unit_cost = nearest_value(unit_cost_x, max(100, width * 0.07), clean_number)
        amount_right = min(width * 0.91, amount_x + max(70, width * 0.05))
        amount = nearest_value(
            amount_x,
            max(130, width * 0.09),
            extract_integer_money,
            min_x=unit_amount_boundary,
            max_x=amount_right,
        )

        inner_values: list[int] = []
        bag_values: list[int] = []
        for entry in row_entries:
            text = entry.text.strip()
            if width * 0.46 <= entry.x <= width * 0.56 and re.fullmatch(r"\d{1,3}", text):
                value = int(text)
                if 0 < value <= 100:
                    inner_values.append(value)
            if width * 0.55 <= entry.x <= width * 0.66 and re.fullmatch(r"0\d+", text):
                value = int(text[1:])
                if value > 0:
                    bag_values.append(value)
        inner_count = inner_values[-1] if inner_values else None
        bag_count = bag_values[-1] if bag_values else None
        calculated_quantity = inner_count * bag_count if inner_count and bag_count else None
        quantity = int(quantity_value) if quantity_value is not None else calculated_quantity

        issue_parts: list[str] = []
        name = normalize_ocr_name(" ".join(name_parts))
        name = name.replace("貓咪頭市貓公", "貓咪頭巾貓公")
        name = re.sub(r"公\s*存", "公仔", name)
        name = name.replace("倚局", "倚肩").replace("爱", "愛").replace("莲", "蓮")
        name = re.sub(r"角色造型吊(?!飾)", "角色造型吊飾", name)
        name = re.sub(r"\bvo1\.(?=\d)", "vol.", name, flags=re.IGNORECASE)
        name = re.sub(r"(?<=[\u4e00-\u9fff])\s+(?=P\d)", "", name)
        issue_parts.extend(suspicious_name_issues(name))
        if not name:
            name = code
            issue_parts.append("品名未穩定辨識")
        if quantity is None:
            issue_parts.append("數量未穩定辨識")
        elif calculated_quantity is not None and quantity != calculated_quantity:
            issue_parts.append(
                f"元大內裝數×袋數={calculated_quantity}，與數量 {quantity} 不符"
            )
        if unit_cost is None:
            issue_parts.append("進價未穩定辨識")
        if amount is None:
            issue_parts.append("金額未穩定辨識")

        check = "需確認"
        if quantity is not None and unit_cost is not None and amount is not None:
            diff = abs(quantity * unit_cost - amount)
            tolerance = max(1, abs(amount) * 0.01)
            if diff <= tolerance:
                check = "通過"
            else:
                swapped = swapped_quantity_unit_cost(quantity, unit_cost, amount, tolerance)
                if swapped is not None:
                    quantity, unit_cost = swapped
                    check = "通過"
                else:
                    inferred = inferred_unit_cost(quantity, amount)
                    if inferred is not None and abs(inferred - unit_cost) > tolerance:
                        check = "不符"
                        issue_parts.append(
                            f"金額÷數量為 {inferred:.3f}，與 OCR 進價 {unit_cost:.3f} 差異過大；未自動覆寫"
                        )
                    else:
                        check = "不符"
                        issue_parts.append(f"數量×進價={quantity * unit_cost:.3f}，與金額 {amount:.3f} 不符")

        rows.append(
            ProductRow(
                vendor_code=code,
                name=name,
                quantity=quantity,
                unit_cost=compact_amount(unit_cost),
                amount=compact_amount(amount),
                check=check,
                issue="；".join(issue_parts) if issue_parts else "無",
            )
        )

    return fill_yuanda_missing_amounts_when_total_matches(rows, entries), global_issues


def name_entry_quality_issues(name: str, name_entries: list[OcrEntry]) -> list[str]:
    issues = suspicious_name_issues(name)
    if any(entry.score < 0.86 and re.search(r"[A-Za-z]", entry.text) for entry in name_entries):
        issues.append("品名英文辨識信心偏低，請核對原圖")
    return list(dict.fromkeys(issues))


def parse_wanrong_rows(entries: list[OcrEntry]) -> tuple[list[ProductRow], list[str]]:
    header_y, positions = header_positions(entries)
    code_entries: list[tuple[OcrEntry, str, str]] = []
    code_left = positions["code"] - 160
    code_right = (positions["code"] + positions["name"]) / 2
    for entry in entries:
        match = PRODUCT_CODE_RE.match(entry.text)
        if not match or not (code_left <= entry.box[0] <= code_right):
            continue
        code = match.group(1)
        if code.upper().startswith(("PH", "PK")) and "-" not in code:
            continue
        code_entries.append((entry, code, match.group(2).strip()))
    code_entries.sort(key=lambda item: item[0].y)
    if not code_entries:
        return [], ["萬榮單據未辨識到商品貨號列。"]

    header_y = min(header_y, code_entries[0][0].box[1] - 10)
    footer_candidates = [
        entry.y
        for entry in entries
        if entry.y > code_entries[-1][0].y
        and any(term in entry.text for term in ("以下空白", "合計數量", "總計", "总计", "匯率"))
    ]
    total_y = min(footer_candidates, default=max((entry.box[3] for entry in entries), default=10_000))
    row_spacings = [
        code_entries[index + 1][0].box[1] - code_entries[index][0].box[1]
        for index in range(len(code_entries) - 1)
        if code_entries[index + 1][0].box[1] > code_entries[index][0].box[1]
    ]
    typical_spacing = median(row_spacings) if row_spacings else 55

    name_left = positions["code"] + max(20, (positions["name"] - positions["code"]) * 0.18)
    name_right = (positions["name"] + positions["quantity"]) / 2
    quantity_left = (positions["name"] + positions["quantity"]) / 2
    quantity_right = (positions["quantity"] + positions.get("unit", positions["quantity"] + 80)) / 2
    discount_x = positions.get("discount", positions["unit_cost"] - 80)
    unit_cost_left = (discount_x + positions["unit_cost"]) / 2
    unit_cost_right = (positions["unit_cost"] + positions["amount"]) / 2
    amount_left = unit_cost_right
    page_right = max((entry.box[2] for entry in entries), default=1200)

    rows: list[ProductRow] = []
    for index, (anchor, code, trailing_name) in enumerate(code_entries):
        row_start_y = max(header_y, anchor.box[1] - 1)
        row_end_y = (
            code_entries[index + 1][0].box[1] - 1
            if index + 1 < len(code_entries)
            else min(total_y, anchor.box[1] + max(70, typical_spacing * 2.0))
        )
        row_entries = [
            entry for entry in entries if row_start_y <= entry.y < row_end_y and entry is not anchor
        ]

        name_entries = [
            entry
            for entry in row_entries
            if name_left <= entry.x < name_right
            and not PRODUCT_CODE_RE.match(entry.text)
            and not is_unit_token(entry.text)
            and not is_plain_number_text(entry.text)
            and not is_noise_or_header_text(entry.text)
        ]
        raw_name_parts = ([trailing_name] if trailing_name else []) + [
            entry.text for entry in sorted(name_entries, key=lambda item: (item.y, item.x))
        ]
        raw_name = re.sub(r"\s+", " ", " ".join(part for part in raw_name_parts if part)).strip()
        name = normalize_ocr_name(raw_name)

        quantity_candidates: list[tuple[float, float, int]] = []
        unit_cost_candidates: list[tuple[float, float, float]] = []
        amount_candidates: list[tuple[float, float, float]] = []
        for entry in row_entries:
            if quantity_left <= entry.x < quantity_right:
                quantity = extract_quantity(entry.text)
                if quantity is not None and quantity > 0:
                    quantity_candidates.append((abs(entry.y - anchor.y), abs(entry.x - positions["quantity"]), quantity))
            if unit_cost_left <= entry.x < unit_cost_right:
                unit_cost = extract_decimal_money(entry.text)
                if unit_cost is None:
                    unit_cost = clean_number(entry.text)
                if unit_cost is not None and unit_cost > 0:
                    unit_cost_candidates.append(
                        (abs(entry.y - anchor.y), abs(entry.x - positions["unit_cost"]), float(unit_cost))
                    )
            if amount_left <= entry.x <= page_right:
                amount = extract_amount_integer(entry.text)
                if amount is not None and amount > 0:
                    amount_candidates.append((abs(entry.y - anchor.y), -entry.x, float(amount)))

        quantity = min(quantity_candidates)[2] if quantity_candidates else None
        unit_cost = min(unit_cost_candidates)[2] if unit_cost_candidates else None
        amount = min(amount_candidates)[2] if amount_candidates else None

        issue_parts = name_entry_quality_issues(name, name_entries)
        if NAME_HEADER_RE.search(raw_name) or SEPARATOR_NOISE_RE.search(raw_name):
            issue_parts.append("品名已移除表頭或分隔線雜訊，請確認")
        if not name or name.upper() == code.upper() or PRODUCT_CODE_RE.fullmatch(name):
            name = name or code
            issue_parts.append("品名未穩定辨識或等於廠商貨號")
        if quantity is None:
            issue_parts.append("數量未穩定辨識")
        if unit_cost is None:
            issue_parts.append("進價未穩定辨識")
        if amount is None:
            issue_parts.append("金額未穩定辨識")

        check = "需確認"
        if quantity is not None and unit_cost is not None and amount is not None:
            if invoice_amount_matches(quantity, unit_cost, amount):
                check = "通過"
            else:
                check = "不符"
                issue_parts.append(f"數量×進價={quantity * unit_cost:.3f}，與金額 {amount:.3f} 不符")
        rows.append(
            ProductRow(
                vendor_code=code,
                name=name,
                quantity=quantity,
                unit_cost=compact_amount(unit_cost),
                amount=compact_amount(amount),
                check=check,
                issue="；".join(dict.fromkeys(issue_parts)) if issue_parts else "無",
            )
        )

    rows = auto_align_numeric_columns(
        rows, entries, code_entries, positions, header_y, total_y, typical_spacing
    )
    return rows, []


def parse_rows(entries: list[OcrEntry]) -> tuple[list[ProductRow], list[str]]:
    if is_wanrong_invoice(entries):
        return parse_wanrong_rows(entries)
    return parse_generic_rows(entries)


def parse_generic_rows(entries: list[OcrEntry]) -> tuple[list[ProductRow], list[str]]:
    header_y, positions = header_positions(entries)
    code_like_entries = [entry for entry in entries if PRODUCT_CODE_RE.match(entry.text)]
    if code_like_entries:
        header_y = min(header_y, min(entry.y for entry in code_like_entries) - 40)
    total_y = min(
        (entry.y for entry in entries if any(term in entry.text for term in ["合計", "總計", "总计"])),
        default=max((entry.y for entry in entries), default=10_000),
    )

    code_entries: list[tuple[OcrEntry, str, str]] = []
    for entry in entries:
        if entry.y <= header_y or entry.y >= total_y:
            continue
        match = PRODUCT_CODE_RE.match(entry.text)
        if not match:
            continue
        if not (
            positions["code"] - 180
            <= entry.box[0]
            <= positions["name"] + 30
        ):
            continue
        code_entries.append((entry, match.group(1), match.group(2).strip()))

    code_entries.sort(key=lambda item: item[0].y)
    row_spacings = [
        code_entries[i + 1][0].y - code_entries[i][0].y
        for i in range(len(code_entries) - 1)
        if code_entries[i + 1][0].y > code_entries[i][0].y
    ]
    typical_spacing = median(row_spacings) if row_spacings else 50
    rows: list[ProductRow] = []
    global_issues: list[str] = []
    for idx, (code_entry, code, trailing_name) in enumerate(code_entries):
        next_code = code_entries[idx + 1][0] if idx + 1 < len(code_entries) else None
        row_start_y = code_entry.box[1] - 1
        next_y = (
            next_code.box[1] - 1
            if next_code is not None
            else min(total_y, code_entry.y + typical_spacing * 1.8)
        )
        name_start_y = row_start_y
        name_end_y = next_y
        numeric_start_y = row_start_y
        numeric_end_y = next_y
        row_entries = [entry for entry in entries if name_start_y <= entry.y < numeric_end_y]

        name_parts: list[str] = []
        if trailing_name:
            name_parts.append(trailing_name)
        numeric_by_col: dict[str, list[float]] = {key: [] for key in positions}
        quantity_candidates: list[tuple[float, int]] = []
        unit_cost_candidates: list[tuple[float, float]] = []
        amount_candidates: list[float] = []
        name_left = positions["code"] + 20
        name_right_candidates = [positions["quantity"] - 20]
        if positions["code"] < positions["unit"] < positions["quantity"]:
            name_right_candidates.append(positions["unit"] - 12)
        name_right = min(name_right_candidates)
        quantity_left = positions["quantity"] - 120
        discount_x = positions.get("discount")
        if discount_x is not None and positions["quantity"] < discount_x < positions["unit_cost"]:
            unit_cost_left = (discount_x + positions["unit_cost"]) / 2
        else:
            unit_cost_left = min(positions["unit_cost"], positions["quantity"]) - 70
        amount_left = positions["amount"] - max(20, (positions["amount"] - positions["unit_cost"]) * 0.25)
        for entry in sorted(row_entries, key=lambda item: (item.y, item.x)):
            if entry is code_entry:
                continue
            col = nearest_column(entry, positions)
            if is_unit_token(entry.text):
                continue
            if numeric_start_y <= entry.y < numeric_end_y:
                number = clean_number(entry.text)
                if number is not None:
                    numeric_by_col[col].append(number)
                if entry.x >= quantity_left:
                    quantity = extract_quantity(entry.text)
                    if quantity is not None:
                        quantity_candidates.append((abs(entry.x - positions["quantity"]), quantity))
                if unit_cost_left <= entry.x < positions["amount"]:
                    unit_cost_value = extract_decimal_money(entry.text)
                    if unit_cost_value is not None:
                        unit_cost_candidates.append((abs(entry.x - positions["unit_cost"]), unit_cost_value))
                if entry.x >= amount_left:
                    amount_value = extract_integer_money(entry.text)
                    if amount_value is not None:
                        amount_candidates.append(amount_value)
            if name_start_y <= entry.y < name_end_y and name_left < entry.x < name_right:
                if not PRODUCT_CODE_RE.match(entry.text) and "以下" not in entry.text:
                    name_parts.append(entry.text)

        quantity = None
        if quantity_candidates:
            quantity = min(quantity_candidates, key=lambda item: item[0])[1]
        elif numeric_by_col["quantity"]:
            integer_values = [int(round(n)) for n in numeric_by_col["quantity"] if abs(n - round(n)) < 0.001]
            quantity = integer_values[0] if integer_values else int(round(numeric_by_col["quantity"][0]))

        unit_cost = None
        if unit_cost_candidates:
            unit_cost = min(unit_cost_candidates, key=lambda item: item[0])[1]
        elif numeric_by_col["unit_cost"]:
            unit_cost = numeric_by_col["unit_cost"][0]

        amount = None
        if amount_candidates:
            amount = amount_candidates[-1]

        name = " ".join(part.strip() for part in name_parts if part.strip())
        name = re.sub(r"\s+", " ", name).strip()
        name = normalize_ocr_name(name)
        if not name or name.upper() == code.upper() or PRODUCT_CODE_RE.fullmatch(name):
            name = code

        issue_parts: list[str] = []
        issue_parts.extend(suspicious_name_issues(name))
        if name.upper() == code.upper() or PRODUCT_CODE_RE.fullmatch(name):
            issue_parts.append("品名未穩定辨識或等於廠商貨號")
        if quantity is None:
            issue_parts.append("數量未穩定辨識")
        if unit_cost is None:
            issue_parts.append("進價未穩定辨識")
        if amount is None:
            issue_parts.append("金額未穩定辨識")

        check = "通過"
        if quantity is not None and unit_cost is not None and amount is not None:
            diff = abs(quantity * unit_cost - amount)
            tolerance = max(1, abs(amount) * 0.01)
            if diff > tolerance:
                swapped = swapped_quantity_unit_cost(quantity, unit_cost, amount, tolerance)
                if swapped is not None:
                    quantity, unit_cost = swapped
                else:
                    inferred = inferred_unit_cost(quantity, amount)
                    if inferred is not None and abs(inferred - unit_cost) > tolerance:
                        check = "不符"
                        issue_parts.append(
                            f"金額÷數量為 {inferred:.3f}，與 OCR 進價 {unit_cost:.3f} 差異過大；未自動覆寫"
                        )
                    else:
                        check = "不符"
                        issue_parts.append(f"數量×進價={quantity * unit_cost:.3f}，與金額 {amount:.3f} 不符")
        else:
            check = "需確認"

        rows.append(
            ProductRow(
                vendor_code=code,
                name=name,
                quantity=quantity,
                unit_cost=compact_amount(unit_cost),
                amount=compact_amount(amount),
                check=check,
                issue="；".join(issue_parts) if issue_parts else "無",
            )
        )

    if not rows:
        global_issues.append("未能從 OCR 座標穩定切出商品列，請查看 OCR測試紀錄工作表。")
    else:
        rows = auto_align_numeric_columns(rows, entries, code_entries, positions, header_y, total_y, typical_spacing)
    return rows, global_issues


def infer_total(entries: list[OcrEntry], rows: list[ProductRow]) -> float | None:
    if not entries:
        return None
    page_bottom = max((entry.box[3] for entry in entries), default=0)
    preferred_terms = (
        "應收總額",
        "应收总额",
        "含稅總額",
        "含税总额",
        "總金額",
        "總金额",
        "总金额",
        "總計",
        "总计",
        "原幣合計",
        "合計",
    )
    labels: list[tuple[int, OcrEntry]] = []
    for entry in entries:
        if "數量" in entry.text or "数量" in entry.text:
            continue
        matched_priority = next(
            (priority for priority, term in enumerate(preferred_terms) if term in entry.text),
            None,
        )
        if matched_priority is None:
            continue
        if entry.y < page_bottom * 0.52 and "原幣合計" not in entry.text:
            continue
        labels.append((matched_priority, entry))

    for _priority, label in sorted(labels, key=lambda item: (item[0], -item[1].y)):
        embedded = clean_number(label.text)
        if embedded is not None and embedded > 0:
            return embedded
        label_height = max(12, label.box[3] - label.box[1])
        max_dy = max(24, label_height * 1.35)
        candidates: list[tuple[float, float, float]] = []
        for candidate in entries:
            if candidate is label or candidate.x <= label.x:
                continue
            if abs(candidate.y - label.y) > max_dy:
                continue
            number = clean_number(candidate.text)
            if number is not None and number > 0:
                candidates.append((abs(candidate.y - label.y), -candidate.x, number))
        if candidates:
            return min(candidates, key=lambda item: (item[0], item[1]))[2]
    return None


def infer_result_total(result: dict[str, Any], rows: list[ProductRow]) -> float | None:
    page_entries = result.get("page_entries") or [result["entries"]]
    return infer_total(page_entries[-1], rows) if page_entries else None


def infer_tax_summary(entries: list[OcrEntry]) -> dict[str, float]:
    if not entries:
        return {}

    def find_label_value(terms: tuple[str, ...]) -> float | None:
        labels = [entry for entry in entries if any(term in entry.text for term in terms)]
        for label in sorted(labels, key=lambda entry: entry.y, reverse=True):
            embedded = clean_number(label.text)
            if embedded is not None and embedded > 0:
                return embedded
            label_height = max(12, label.box[3] - label.box[1])
            candidates = [
                candidate
                for candidate in entries
                if candidate is not label
                and candidate.x > label.x
                and abs(candidate.y - label.y) <= max(24, label_height * 1.35)
            ]
            values = [
                (abs(candidate.y - label.y), -candidate.x, clean_number(candidate.text))
                for candidate in candidates
                if clean_number(candidate.text) is not None and clean_number(candidate.text) > 0
            ]
            if values:
                return min(values, key=lambda item: (item[0], item[1]))[2]
        return None

    subtotal = find_label_value(("未稅", "合計金額", "合計金额"))
    tax = find_label_value(("稅額", "税额", "營業稅", "营业税"))
    total = find_label_value(("含稅總額", "含税总额", "總金額", "總金额", "总金额", "應收總額"))
    summary: dict[str, float] = {}
    if subtotal is not None:
        summary["未稅合計"] = subtotal
    if tax is not None:
        summary["稅額"] = tax
    if total is not None:
        summary["含稅總額"] = total
    return summary


def collect_review_issues(global_issues: list[str], rows: list[ProductRow]) -> list[str]:
    row_issues = [
        f"{row.vendor_code}: {row.issue}"
        for row in rows
        if row.issue != "無" or row.check != "通過"
    ]
    return list(dict.fromkeys([*global_issues, *row_issues]))


def write_workbook(
    image_path: Path,
    output_path: Path,
    vendor: str,
    best: dict[str, Any],
    all_results: list[dict[str, Any]],
    rows: list[ProductRow],
    global_issues: list[str],
    timings: dict[str, float],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "進貨明細"
    is_liying_output = vendor_short(vendor) == "麗嬰" or any(row.retail_price is not None for row in rows)
    max_col = 6 if is_liying_output else 5
    ws.append([f"廠商：{vendor}", *([None] * (max_col - 1))])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    if is_liying_output:
        ws.append(["產品代號", "品名", "零售價", "數量", "進價", "金額"])
    else:
        ws.append(["產品代號", "品名", "數量", "進價", "金額"])
    for row in rows:
        if is_liying_output:
            ws.append(["", row.name, row.retail_price, row.quantity, row.unit_cost, row.amount])
        else:
            ws.append(["", row.name, row.quantity, row.unit_cost, row.amount])
    total = infer_result_total(best, rows)
    if total is None:
        amounts = [row.amount for row in rows if row.amount is not None]
        total = sum(amounts) if amounts else None
    ws.append(["總價格", *([None] * (max_col - 2)), compact_amount(total)])
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=max_col - 1)

    header_fill = PatternFill("solid", fgColor="D9EAF0")
    title_fill = PatternFill("solid", fgColor="1F4E5F")
    thin = Side(style="thin", color="D9E2E6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws["A1"].fill = title_fill
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=(cell.column == 2))
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=3, max_col=max_col):
        for cell in row:
            cell.number_format = "#,##0.###" if cell.column in ({5} if is_liying_output else {4}) else "#,##0"
            cell.alignment = Alignment(horizontal="right")
    for cell in ws["A"]:
        cell.number_format = "@"
    widths = {"A": 12, "B": 58, "C": 12, "D": 12, "E": 12}
    if is_liying_output:
        widths = {"A": 12, "B": 58, "C": 10, "D": 10, "E": 12, "F": 12}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A3"

    note = wb.create_sheet("OCR測試紀錄")
    note.append(["來源圖片", str(image_path)])
    note.append(["最佳版本", best["variant"]])
    note.append(["平均信心", best["avg_score"]])
    note.append(["關鍵字命中", best["key_hits"]])
    note.append(["商品列數", len(rows)])
    note.append(["疑似商品列錨點", estimate_product_anchor_count(best)])
    note.append(["偵測頁數", len(best.get("page_entries") or [best["entries"]])])
    note.append(["模型載入秒數", timings.get("model_load_seconds", "")])
    note.append(["OCR總秒數", timings.get("total_ocr_seconds", "")])
    note.append(["整體疑點", "；".join(global_issues) if global_issues else "無"])
    page_entries = best.get("page_entries") or [best["entries"]]
    tax_summary = infer_tax_summary(page_entries[-1]) if page_entries else {}
    for label in ("未稅合計", "稅額", "含稅總額"):
        if label in tax_summary:
            note.append([label, compact_amount(tax_summary[label])])
    audit_notes = [row.audit_note for row in rows if row.audit_note]
    if audit_notes:
        note.append(["自動修正紀錄", "；".join(dict.fromkeys(audit_notes))])
    note.append([])
    note.append(["版本", "行數", "平均信心", "關鍵字命中", "貨號命中", "辨識秒數", "暫存圖"])
    for result in sorted(
        all_results,
        key=lambda r: (r["product_code_hits"], r["key_hits"], r["line_count"], r["avg_score"]),
        reverse=True,
    ):
        note.append(
            [
                result["variant"],
                result["line_count"],
                result["avg_score"],
                result["key_hits"],
                result["product_code_hits"],
                result.get("predict_seconds", ""),
                result["path"],
            ]
        )
    note.append([])
    if is_liying_output:
        note.append(["列號", "廠商貨號", "品名", "零售價", "數量", "進價", "金額", "金額核對", "OCR疑點"])
    else:
        note.append(["列號", "廠商貨號", "品名", "數量", "進價", "金額", "金額核對", "OCR疑點"])
    for idx, row in enumerate(rows, start=1):
        if is_liying_output:
            note.append(
                [
                    idx,
                    row.vendor_code,
                    row.name,
                    row.retail_price,
                    row.quantity,
                    row.unit_cost,
                    row.amount,
                    row.check,
                    row.issue,
                ]
            )
        else:
            note.append([idx, row.vendor_code, row.name, row.quantity, row.unit_cost, row.amount, row.check, row.issue])
    note.append([])
    note.append(["原始OCR文字", "信心", "x1", "y1", "x2", "y2"])
    for entry in best["entries"]:
        note.append([entry.text, round(entry.score, 4), *entry.box])

    for sheet in [note]:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col in range(1, sheet.max_column + 1):
            letter = get_column_letter(col)
            sheet.column_dimensions[letter].width = 18
        sheet.column_dimensions["A"].width = 36
        sheet.column_dimensions["C"].width = 50
        sheet.freeze_panes = "A2"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Use local PaddleOCR to convert a vendor invoice image into review xlsx.")
    parser.add_argument("image", type=Path, help="進貨單圖片路徑")
    parser.add_argument("--output", type=Path, default=None, help="指定輸出 xlsx 路徑")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR, help="輸出資料夾")
    parser.add_argument("--tmp-dir", type=Path, default=DEFAULT_TMP_DIR, help="OCR 暫存資料夾")
    parser.add_argument("--settings", type=Path, default=DEFAULT_SETTINGS_PATH, help="OCR 設定 JSON 路徑")
    parser.add_argument(
        "--rotation",
        choices=["auto", "none", "cw", "ccw", "180"],
        default=None,
        help="照片前處理旋轉方式；未指定時讀取 OCR設定.json，預設 auto。",
    )
    parser.add_argument("--multi-variant", action="store_true", help="手動開啟舊式多版本 OCR，命中足夠結果後早停。")
    parser.add_argument("--contrast", type=float, default=None, help="灰階對比增強倍率；未指定時讀取 OCR設定.json。")
    parser.add_argument("--sharpness", type=float, default=None, help="銳化倍率；未指定時讀取 OCR設定.json。")
    parser.add_argument(
        "--page-split",
        choices=["auto", "off"],
        default=None,
        help="上下多頁照片自動分割；未指定時讀取 OCR設定.json，預設 auto。",
    )
    args = parser.parse_args()

    image_path = args.image.expanduser().resolve()
    if not image_path.exists():
        raise SystemExit(f"找不到圖片：{image_path}")

    lock_path = acquire_image_lock(image_path)
    try:
        settings = load_settings(args.settings.expanduser().resolve()) if args.settings else {}
        rotation = args.rotation or str(settings.get("rotation", "auto"))
        multi_variant = bool(args.multi_variant or settings.get("multi_variant", False))
        contrast = float(args.contrast if args.contrast is not None else settings.get("contrast", 1.8))
        sharpness = float(args.sharpness if args.sharpness is not None else settings.get("sharpness", 1.15))
        page_split = args.page_split or str(settings.get("page_split", "auto"))

        best, all_results, timings = run_ocr(
            image_path,
            args.tmp_dir / image_path.stem,
            rotation,
            multi_variant,
            contrast,
            sharpness,
            page_split,
        )
        vendor = infer_vendor(best["entries"])
        rows, global_issues = parse_result_rows(best)
        global_issues.extend(row_count_audit_issues(best, rows))
        all_issues = collect_review_issues(global_issues, rows)
        output_path = args.output.expanduser().resolve() if args.output else unique_output_path(args.output_dir, vendor)
        write_workbook(image_path, output_path, vendor, best, all_results, rows, all_issues, timings)
        metrics = result_quality_metrics(best)

        summary = {
            "ok": True,
            "output": str(output_path),
            "vendor": vendor,
            "best_variant": best["variant"],
            "row_count": len(rows),
            "estimated_row_count": estimate_product_anchor_count(best),
            "page_count": len(best.get("page_entries") or [best["entries"]]),
            "model_load_seconds": timings.get("model_load_seconds"),
            "total_ocr_seconds": timings.get("total_ocr_seconds"),
            "valid_row_count": metrics["valid_rows"],
            "complete_row_count": metrics["complete_rows"],
            "total_matches": bool(metrics["total_matches"]),
            "page_error_count": sum(
                1
                for issue in all_issues
                if re.search(r"第\d+頁：未能從 OCR 座標穩定切出商品列", issue)
            ),
            "page_issue_count": metrics["page_penalty"],
            "needs_review": bool(all_issues),
            "issues": all_issues,
        }
        print(json.dumps(summary, ensure_ascii=False, indent=2))
        return 0
    finally:
        release_image_lock(lock_path)


if __name__ == "__main__":
    sys.exit(main())
