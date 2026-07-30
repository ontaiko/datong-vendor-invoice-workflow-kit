from __future__ import annotations

import importlib.util
import csv
import hashlib
import json
import os
import re
import shutil
import subprocess
import sys
from copy import copy
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook


def app_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


APP_DIR = app_dir()
REFERENCE_DIR = APP_DIR / "reference_data"
SCRIPTS_DIR = APP_DIR / "scripts"
CONFIG_PATH = APP_DIR / "app_settings.json"
OCR_SCRIPT = SCRIPTS_DIR / "local_paddleocr_invoice_to_xlsx.py"


def load_app_settings() -> dict[str, Any]:
    if not CONFIG_PATH.exists():
        return {}
    try:
        data = json.loads(CONFIG_PATH.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError) as exc:
        raise RuntimeError(f"APP 設定檔無法讀取：{CONFIG_PATH}\n{exc}") from exc
    return data if isinstance(data, dict) else {}


APP_SETTINGS = load_app_settings()
PROJECT_ROOT = Path(
    os.path.expandvars(
        os.environ.get("DATONG_WORKSPACE")
        or str(APP_SETTINGS.get("workspace_root") or "")
        or str(Path.home() / "Documents" / "大統工作助手")
    )
).expanduser()
_PRODUCT_CSV_OVERRIDE: Path | None = None
_configured_product_csv = os.path.expandvars(str(APP_SETTINGS.get("product_csv_path") or "")).strip()
if _configured_product_csv:
    _configured_product_csv_path = Path(_configured_product_csv).expanduser()
    if _configured_product_csv_path.exists():
        _PRODUCT_CSV_OVERRIDE = _configured_product_csv_path.resolve()


def save_app_settings() -> None:
    CONFIG_PATH.write_text(
        json.dumps(APP_SETTINGS, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


CODE_RE = re.compile(r"^\d{6}$")
SUMMARY_ROW_NAMES = {"總價格", "總價", "總計", "合計", "小計", "稅金", "稅額", "折扣", "總數量", "合計數量"}
DEFAULT_CATEGORY_ROWS = (
    ("1", "PS主機", "PS主機", "PS主機", ""),
    ("2", "PS配件", "PS配件", "PS配件", ""),
    ("3", "PS遊戲", "PS遊戲", "PS遊戲", ""),
    ("4", "PS遊戲(二手)", "二手", "PS遊戲二手;PS4二手;PS5二手", ""),
    ("5", "NS主機", "NS主機", "NS主機;Switch主機", ""),
    ("6", "NS遊戲", "NS遊戲", "NS遊戲;Switch遊戲", ""),
    ("7", "NS遊戲(二手)", "二手", "NS遊戲二手;Switch遊戲二手", ""),
    ("8", "NS Amiibo", "Amiibo", "Amiibo", ""),
    ("9", "NS控制器", "NS控制器", "NS控制器;Joy-Con;Pro控制器", ""),
    ("10", "NS包包類", "NS包包", "NS包包;Switch包包", ""),
    ("11", "NS配件", "NS配件", "NS配件;Switch配件", ""),
    ("13", "PS5控制器", "PS5控制器", "PS5控制器", ""),
    ("14", "XBOX", "XBOX", "XBOX", ""),
    ("15", "NS充電與轉接", "NS充電", "NS充電;轉接", ""),
    ("17", "NS保護貼", "NS保護貼", "NS保護貼;Switch保護貼", ""),
    ("18", "記憶/點數卡", "點數卡", "記憶卡;點數卡", ""),
    ("19", "方向盤類", "方向盤", "方向盤", ""),
    ("21", "攻略本", "攻略本", "攻略本", ""),
    ("25", "TOMICA車車", "TOMICA", "TOMICA;多美小汽車", ""),
    ("26", "一番賞", "一番賞", "一番賞", ""),
    ("27", "卡牌&卡牌周邊", "卡牌", "卡牌;卡套;牌盒;收藏卡", ""),
    ("28", "鑰匙圈", "鑰匙圈", "鑰匙圈", ""),
    ("29", "沐浴球系列", "沐浴球", "沐浴球", ""),
    ("31", "日版授權娃娃", "娃娃", "日版授權娃娃", ""),
    ("33", "動漫盒玩", "盒玩", "動漫盒玩", ""),
    ("34", "大陸盒玩類", "盲盒", "大陸盒玩;盲盒;盒玩;中盒", "萬榮盲盒目前沿用此大類"),
    ("35", "泡泡瑪特盒玩", "泡泡瑪特", "泡泡瑪特;POP MART", ""),
    ("36", "台灣盲盒", "台灣盲盒", "台灣盲盒", ""),
    ("37", "日本盒玩", "日本盒玩", "日本盒玩;日版盒玩", ""),
    ("38", "公仔/吊卡/PVC", "公仔", "公仔;吊卡;PVC", ""),
    ("39", "景品", "景品", "景品;Coreful;Desktop Cute;FIGURIZM;VIBRATION STARS;KING OF ARTIST", ""),
    ("40", "夜燈", "夜燈", "夜燈", ""),
    ("45", "綜合動漫周邊", "動漫周邊", "動漫周邊;綜合動漫周邊", ""),
    ("46", "動漫周邊徽章", "徽章", "徽章", ""),
    ("47", "吉伊卡哇周邊", "吉伊卡哇", "吉伊卡哇;Chiikawa", ""),
    ("48", "寶可夢周邊", "寶可夢", "寶可夢;Pokemon;Pokémon", ""),
    ("51", "卡比周邊", "卡比", "卡比;Kirby", ""),
    ("96", "維修類", "維修", "維修", ""),
    ("97", "禮包/備品/特典", "特典", "禮包;備品;特典", ""),
    ("98", "其他(文具&雜物)", "文具", "文具;雜物;其他", ""),
)


@dataclass
class OcrConfirmRow:
    excel_row: int
    is_existing: bool
    raw_name: str
    quantity: str
    unit_cost: str
    amount: str
    matched_code: str
    matched_name: str
    candidates: str
    status: str


@dataclass
class AdjustmentRow:
    row_id: str
    excel_row: int
    source_row: int
    product_code: str
    name: str
    category: str
    category_display: str
    quantity: str
    unit_cost: str
    amount: str
    status: str


@dataclass
class WorkflowState:
    image_path: Path | None = None
    image_paths: list[Path] = field(default_factory=list)
    output_dir: Path | None = None
    formal_output_dir: Path | None = None
    raw_xlsx: Path | None = None
    raw_xlsx_files: list[Path] = field(default_factory=list)
    match_xlsx: Path | None = None
    suggested_names_txt: Path | None = None
    adjusted_xlsx: Path | None = None
    new_product_file: Path | None = None
    purchase_file: Path | None = None
    vendor: str = ""
    invoice_total: str = ""
    row_count: int = 0
    needs_ocr_review: bool = False
    ocr_issues: list[str] = field(default_factory=list)
    excluded_items: list[str] = field(default_factory=list)
    build_summary: dict[str, Any] = field(default_factory=dict)
    tmp_dir: Path | None = None
    processed_image_path: Path | None = None
    processed_image_paths: list[Path] = field(default_factory=list)


def python_exe(project_root: Path = PROJECT_ROOT) -> Path:
    configured = os.path.expandvars(
        os.environ.get("DATONG_PYTHON_EXE") or str(APP_SETTINGS.get("python_exe") or "")
    )
    candidates = [
        Path(configured).expanduser() if configured else None,
        APP_DIR / "engine" / ".venv" / "Scripts" / "python.exe",
        Path(os.environ.get("LOCALAPPDATA", "")) / "Programs" / "Python" / "Python312" / "python.exe",
        Path(sys.executable) if not getattr(sys, "frozen", False) else None,
        Path(shutil.which("python") or ""),
    ]
    for candidate in candidates:
        if candidate and str(candidate) not in {"", "."} and candidate.exists():
            return candidate.resolve()
    return Path("python.exe")


REFERENCE_FILE_NAMES = (
    "產品資料輸出.CSV",
    "產品比對身份關鍵詞.csv",
    "品牌括號命名規則.csv",
    "廠商代號.xlsx",
    "建檔用.xls",
    "採購單匯入範例.xls",
    "OCR設定.json",
)


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def sync_reference_data(project_root: Path = PROJECT_ROOT) -> list[str]:
    source_dir = project_root / "參考資料"
    if not source_dir.exists():
        return []
    REFERENCE_DIR.mkdir(parents=True, exist_ok=True)
    synced: list[str] = []
    for name in REFERENCE_FILE_NAMES:
        source = source_dir / name
        target = REFERENCE_DIR / name
        if not source.exists():
            continue
        should_copy = not target.exists()
        if target.exists() and file_sha256(source) != file_sha256(target):
            if name == "產品資料輸出.CSV":
                try:
                    assert_product_csv_current(source)
                    source_current = True
                except RuntimeError:
                    source_current = False
                try:
                    assert_product_csv_current(target)
                    target_current = True
                except RuntimeError:
                    target_current = False
                if source_current and not target_current:
                    should_copy = True
                elif source_current and target_current:
                    should_copy = source.stat().st_mtime > target.stat().st_mtime
                else:
                    # Never replace a validated current catalog with an older
                    # workspace file.  When both are stale, keep the selected
                    # copy and let the normal date check report the problem.
                    should_copy = False
            else:
                # Preserve newer APP-side edits during upgrades or portable use.
                # Workspace reference files only replace the installed copy when
                # they are actually newer, rather than merely different.
                should_copy = source.stat().st_mtime > target.stat().st_mtime
        if should_copy:
            shutil.copy2(source, target)
            synced.append(name)
    return synced


def hidden_process_options() -> dict[str, Any]:
    if os.name != "nt":
        return {}
    startupinfo = subprocess.STARTUPINFO()
    startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
    startupinfo.wShowWindow = getattr(subprocess, "SW_HIDE", 0)
    return {
        "creationflags": subprocess.CREATE_NO_WINDOW,
        "startupinfo": startupinfo,
    }


def product_csv() -> Path | None:
    if _PRODUCT_CSV_OVERRIDE is not None and _PRODUCT_CSV_OVERRIDE.exists():
        return _PRODUCT_CSV_OVERRIDE
    candidates = sorted(
        REFERENCE_DIR.glob("產品資料輸出*.CSV"),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    return candidates[0] if candidates else None


def set_product_csv_path(path: Path, persist: bool = False) -> Path:
    global _PRODUCT_CSV_OVERRIDE
    resolved = path.expanduser().resolve()
    if not resolved.exists() or not resolved.is_file():
        raise RuntimeError(f"找不到產品資料輸出 CSV：{resolved}")
    if resolved.suffix.lower() != ".csv":
        raise RuntimeError(f"產品資料輸出必須是 CSV：{resolved}")
    _PRODUCT_CSV_OVERRIDE = resolved
    if persist:
        APP_SETTINGS["product_csv_path"] = str(resolved)
        save_app_settings()
    return resolved


def assert_product_csv_current(path: Path) -> None:
    packaged_marker = path.with_name(f"{path.name}.packaged.sha256")
    if packaged_marker.exists():
        packaged_hash = packaged_marker.read_text(encoding="utf-8-sig").strip().split()[0].lower()
        if packaged_hash and file_sha256(path).lower() == packaged_hash:
            raise RuntimeError(
                "產品資料檔仍是安裝包隨附的舊快照。\n"
                f"目前檔案：{path}\n"
                "請從系統重新匯出今天的產品資料輸出.CSV，再放入 APP 的 reference_data。"
            )
    timezone = ZoneInfo("Asia/Taipei")
    stat = path.stat()
    created_at = datetime.fromtimestamp(stat.st_ctime, timezone)
    modified_at = datetime.fromtimestamp(stat.st_mtime, timezone)
    today = datetime.now(timezone).date()
    if created_at.date() != today and modified_at.date() != today:
        raise RuntimeError(
            "產品資料輸出.CSV 不是今天建立或修改的版本。\n"
            f"目前檔案：{path}\n"
            f"建立時間：{created_at:%Y-%m-%d %H:%M:%S}\n"
            f"修改時間：{modified_at:%Y-%m-%d %H:%M:%S}\n"
            "請先把今天的產品資料輸出.CSV 放進 reference_data。"
        )


def prepare_product_csv_for_use(
    path: Path,
    project_root: Path = PROJECT_ROOT,
    persist: bool = False,
) -> Path:
    resolved = path.expanduser().resolve()
    if not resolved.exists() or not resolved.is_file():
        raise RuntimeError(f"找不到產品資料輸出 CSV：{resolved}")
    if resolved.suffix.lower() != ".csv":
        raise RuntimeError(f"產品資料輸出必須是 CSV：{resolved}")

    try:
        assert_product_csv_current(resolved)
    except RuntimeError as original_error:
        # The packaged APP normally points at its own reference_data copy.
        # If that copy was stale, the user may replace the authoritative file
        # in the workspace while the APP stays open.  Refresh only after the
        # selected file failed validation, so a directly updated APP copy is
        # never overwritten by an older workspace file.
        try:
            resolved.relative_to(REFERENCE_DIR.resolve())
            is_internal_reference = True
        except ValueError:
            is_internal_reference = False
        source = (project_root / "參考資料" / resolved.name).resolve()
        if not is_internal_reference or source == resolved or not source.exists():
            raise original_error
        try:
            assert_product_csv_current(source)
            shutil.copy2(source, resolved)
            assert_product_csv_current(resolved)
        except (OSError, RuntimeError) as refresh_error:
            raise RuntimeError(
                f"{original_error}\n\n"
                "APP 已重新檢查工作區的產品資料，但仍無法更新：\n"
                f"{source}\n{refresh_error}"
            ) from refresh_error

    return set_product_csv_path(resolved, persist=persist)


def read_csv_dicts(path: Path) -> list[dict[str, str]]:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "cp950", "big5", "utf-8"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                return [{k: (v or "").strip() for k, v in row.items()} for row in csv.DictReader(handle)]
        except UnicodeDecodeError as error:
            last_error = error
    raise RuntimeError(f"無法讀取 CSV：{path} ({last_error})")


def category_name_map() -> dict[str, str]:
    return {
        str(row["code"]).strip(): str(row["name"]).strip()
        for row in category_rule_rows()
        if str(row.get("code") or "").strip() and str(row.get("name") or "").strip()
    }


def default_category_rule_rows() -> list[dict[str, str]]:
    return [
        {
            "code": code,
            "name": name,
            "label": label,
            "keywords": keywords,
            "note": note,
        }
        for code, name, label, keywords, note in DEFAULT_CATEGORY_ROWS
    ]


def category_rule_rows() -> list[dict[str, str]]:
    configured = APP_SETTINGS.get("categories")
    source = configured if isinstance(configured, list) else default_category_rule_rows()
    rows: list[dict[str, str]] = []
    seen_codes: set[str] = set()
    for item in source:
        if not isinstance(item, dict):
            continue
        code = str(item.get("code") or item.get("大類代號") or "").strip()
        name = str(item.get("name") or item.get("大類名稱") or "").strip()
        if not code or not name or code in seen_codes:
            continue
        seen_codes.add(code)
        rows.append(
            {
                "code": code,
                "name": name,
                "label": str(item.get("label") or item.get("括號分類") or name).strip(),
                "keywords": str(item.get("keywords") or item.get("關鍵字") or name).strip(),
                "note": str(item.get("note") or item.get("備註") or "").strip(),
            }
        )
    return sorted(
        rows,
        key=lambda row: (
            0,
            int(row["code"]),
        )
        if row["code"].isdigit()
        else (1, row["code"]),
    )


def save_category_rule_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    normalized: list[dict[str, str]] = []
    seen_codes: set[str] = set()
    for item in rows:
        code = str(item.get("code") or "").strip()
        name = str(item.get("name") or "").strip()
        if not code or not name:
            raise RuntimeError("大類代號與大類名稱不可空白。")
        if code in seen_codes:
            raise RuntimeError(f"大類代號重複：{code}")
        seen_codes.add(code)
        normalized.append(
            {
                "code": code,
                "name": name,
                "label": str(item.get("label") or name).strip(),
                "keywords": str(item.get("keywords") or name).strip(),
                "note": str(item.get("note") or "").strip(),
            }
        )
    APP_SETTINGS["categories"] = normalized
    save_app_settings()
    return category_rule_rows()


def reset_category_rule_rows() -> list[dict[str, str]]:
    APP_SETTINGS.pop("categories", None)
    save_app_settings()
    return category_rule_rows()


def runtime_category_rules_path() -> Path:
    runtime_dir = PROJECT_ROOT / ".codex-tmp" / "invoice-app-runtime"
    runtime_dir.mkdir(parents=True, exist_ok=True)
    path = runtime_dir / "大類清單-APP內建.csv"
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=("大類代號", "大類名稱", "括號分類", "關鍵字", "啟用", "備註"),
        )
        writer.writeheader()
        for row in category_rule_rows():
            writer.writerow(
                {
                    "大類代號": row["code"],
                    "大類名稱": row["name"],
                    "括號分類": row["label"],
                    "關鍵字": row["keywords"],
                    "啟用": "1",
                    "備註": row["note"],
                }
            )
    return path


def category_code(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    first = text.split(maxsplit=1)[0].strip()
    return first if first else text


def category_display(value: str, names: dict[str, str] | None = None) -> str:
    code = category_code(value)
    if not code:
        return ""
    lookup = names if names is not None else category_name_map()
    name = lookup.get(code)
    return f"{code} {name}" if name else f"{code} 未知大類"


def normalize_product_code(value: Any) -> str:
    text = str(value or "").strip()
    formula_match = re.fullmatch(r'="\s*(\d+)\s*"', text)
    if formula_match:
        text = formula_match.group(1)
    digits = re.sub(r"\D", "", text)
    if not digits or len(digits) > 6:
        return ""
    return digits.zfill(6)


def parse_pasted_product_codes(value: Any) -> tuple[list[str], list[str]]:
    text = str(value or "").replace("\r\n", "\n").replace("\r", "\n")
    nonempty_lines = [line for line in text.split("\n") if line.strip()]
    if len(nonempty_lines) > 1:
        raw_values = [line.split("\t", 1)[0].strip() for line in nonempty_lines]
    elif "\t" in text:
        raw_values = [cell.strip() for cell in text.split("\t") if cell.strip()]
    else:
        raw_values = [text.strip()] if text.strip() else []

    codes: list[str] = []
    invalid: list[str] = []
    for raw_value in raw_values:
        normalized = normalize_product_code(raw_value)
        if CODE_RE.fullmatch(normalized):
            codes.append(normalized)
        else:
            invalid.append(raw_value)
    return codes, invalid


def parse_pasted_column_values(value: Any) -> list[str]:
    text = str(value or "").replace("\r\n", "\n").replace("\r", "\n")
    lines = text.split("\n")
    while lines and lines[-1] == "":
        lines.pop()
    if len(lines) > 1:
        # Excel copies a selected column as one tab-delimited row per line.
        # Keep internal blank cells so values below a blank row do not shift up.
        return [line.split("\t", 1)[0].strip() for line in lines]
    if "\t" in text:
        return [cell.strip() for cell in text.split("\t")]
    return [text.strip()] if text else []


def catalog_product_names(codes: list[str]) -> dict[str, str]:
    wanted = {
        normalized
        for code in codes
        if CODE_RE.fullmatch(normalized := normalize_product_code(code))
    }
    if not wanted:
        return {}
    csv_path = product_csv()
    if csv_path is None:
        return {}
    result: dict[str, str] = {}
    for row in read_csv_dicts(csv_path):
        row_code = normalize_product_code(
            row.get("1.產品代號")
            or row.get("產品代號")
            or row.get("ProductCode")
            or ""
        )
        if row_code in wanted:
            result[row_code] = str(
                row.get("2.產品名稱")
                or row.get("產品名稱")
                or row.get("品名")
                or row.get("ProductName")
                or ""
            ).strip()
            if len(result) == len(wanted):
                break
    return result


def catalog_product_name(code: str) -> str:
    normalized = normalize_product_code(code)
    return catalog_product_names([normalized]).get(normalized, "")


def validate_runtime(project_root: Path = PROJECT_ROOT) -> list[str]:
    try:
        sync_reference_data(project_root)
    except OSError as exc:
        return [f"同步工作區參考資料失敗：{exc}"]
    issues: list[str] = []
    resolved_python = python_exe(project_root)
    required = [
        resolved_python,
        OCR_SCRIPT,
        SCRIPTS_DIR / "match-existing-products.py",
        SCRIPTS_DIR / "review-invoice-product-check.py",
        SCRIPTS_DIR / "fill-import-templates.ps1",
        REFERENCE_DIR / "產品比對身份關鍵詞.csv",
        REFERENCE_DIR / "品牌括號命名規則.csv",
        REFERENCE_DIR / "廠商代號.xlsx",
        REFERENCE_DIR / "建檔用.xls",
        REFERENCE_DIR / "採購單匯入範例.xls",
    ]
    for path in required:
        if not path.exists():
            issues.append(f"找不到必要檔案：{path}")
    if resolved_python.exists():
        completed = subprocess.run(
            [
                str(resolved_python),
                "-X",
                "utf8",
                "-c",
                "import cv2, openpyxl, paddleocr, paddle; print('LOCAL_ENGINE_OK')",
            ],
            text=True,
            encoding="utf-8",
            errors="replace",
            capture_output=True,
            check=False,
            **hidden_process_options(),
        )
        if completed.returncode != 0:
            detail = "\n".join(part for part in [completed.stdout, completed.stderr] if part).strip()
            issues.append(
                "系統 OCR 尚未安裝完成，請確認全系統 Python 已安裝 PaddleOCR。\n"
                f"Python：{resolved_python}\n{detail}"
            )
    csv_path = product_csv()
    if csv_path is None:
        issues.append(f"找不到產品資料輸出.CSV：{REFERENCE_DIR}")
    return issues


def parse_json_summary(output: str) -> dict[str, Any]:
    decoder = json.JSONDecoder()
    for match in reversed(list(re.finditer(r"\{", output))):
        try:
            data, _ = decoder.raw_decode(output[match.start() :])
        except (json.JSONDecodeError, TypeError):
            continue
        if isinstance(data, dict):
            return data
    raise RuntimeError("程式已結束，但無法讀取結果摘要。")


def read_invoice_total(path: Path) -> str:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        worksheet = workbook["進貨明細"] if "進貨明細" in workbook.sheetnames else workbook.active
        for row in worksheet.iter_rows(values_only=True):
            values = list(row)
            for index, value in enumerate(values):
                text = str(value or "").strip()
                if text in {"總價格", "總價", "總計", "合計"}:
                    for candidate in values[index + 1 :]:
                        if candidate not in (None, ""):
                            return str(candidate).replace(",", "").strip()
        return ""
    finally:
        workbook.close()


def decimal_from_value(value: Any) -> Decimal:
    text = str(value or "").replace(",", "").strip()
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise RuntimeError(f"無法解析單據總額：{value}") from exc


def decimal_text(value: Decimal) -> str:
    text = format(value, "f")
    return text.rstrip("0").rstrip(".") if "." in text else text


def state_image_paths(state: WorkflowState) -> list[Path]:
    if state.image_paths:
        return [Path(path) for path in state.image_paths]
    return [state.image_path] if state.image_path is not None else []


def run_command(command: list[str], cwd: Path, env: dict[str, str] | None = None) -> subprocess.CompletedProcess[str]:
    merged_env = os.environ.copy()
    merged_env["PYTHONUTF8"] = "1"
    if env:
        merged_env.update(env)
    return subprocess.run(
        command,
        cwd=str(cwd),
        env=merged_env,
        text=True,
        encoding="utf-8",
        errors="replace",
        capture_output=True,
        check=False,
        **hidden_process_options(),
    )


def ocr_summary_quality(summary: dict[str, Any]) -> tuple[int, int, int, int, int, int, int, int]:
    row_count = int(summary.get("row_count", 0) or 0)
    estimated = int(summary.get("estimated_row_count", 0) or 0)
    complete = int(summary.get("complete_row_count", 0) or 0)
    valid = int(summary.get("valid_row_count", 0) or 0)
    total_matches = int(bool(summary.get("total_matches")))
    page_errors = int(summary.get("page_error_count", 0) or 0)
    page_issues = int(summary.get("page_issue_count", 0) or 0)
    return (
        min(row_count, estimated or row_count),
        complete,
        valid,
        total_matches,
        -abs(estimated - row_count),
        -page_errors,
        -page_issues,
        int(summary.get("page_count", 1) or 1),
    )


def finalize_vendor_output_name(path: Path, vendor: str) -> Path:
    if not path.exists() or "未辨識廠進貨單-" not in path.name or not vendor or "未辨識" in vendor:
        return path
    prefix = next(
        (
            short
            for marker, short in (
                ("萬榮", "萬榮"),
                ("麗嬰", "麗嬰"),
                ("元大", "元大玩具"),
                ("金亞特", "金亞特"),
                ("南波", "南波"),
                ("鉅霖", "鉅霖"),
            )
            if marker in vendor
        ),
        re.sub(r"[^\w\u4e00-\u9fff]+", "", vendor)[:4],
    )
    desired_name = path.name.replace("未辨識廠進貨單-", f"{prefix}進貨單-", 1)
    target = path.with_name(desired_name)
    if target.exists():
        stem_match = re.match(r"^(.*)-(\d{2})$", target.stem)
        base = stem_match.group(1) if stem_match else target.stem
        for index in range(1, 100):
            candidate = target.with_name(f"{base}-{index:02d}{target.suffix}")
            if not candidate.exists():
                target = candidate
                break
    path.replace(target)
    return target


def run_ocr(state: WorkflowState, project_root: Path = PROJECT_ROOT) -> dict[str, Any]:
    if state.image_path is None or state.output_dir is None:
        raise RuntimeError("尚未選擇圖片或輸出資料夾。")
    state.tmp_dir = project_root / ".codex-tmp" / "invoice-ocr-gui" / state.image_path.stem
    command = [
        str(python_exe(project_root)),
        "-X",
        "utf8",
        str(OCR_SCRIPT),
        str(state.image_path),
        "--output-dir",
        str(state.output_dir),
        "--tmp-dir",
        str(state.tmp_dir),
        "--settings",
        str(REFERENCE_DIR / "OCR設定.json"),
    ]
    completed = run_command(
        command,
        cwd=project_root,
        env={"PADDLE_PDX_ENABLE_MKLDNN_BYDEFAULT": "0"},
    )
    output = "\n".join(part for part in [completed.stdout, completed.stderr] if part)
    if completed.returncode != 0:
        raise RuntimeError(output or f"OCR 程式結束代碼：{completed.returncode}")
    summary = parse_json_summary(output)
    state.raw_xlsx = Path(str(summary.get("output", "")))
    false_page_split = int(summary.get("page_count", 1) or 1) > 1 and any(
        re.search(r"第\d+頁：未能從 OCR 座標穩定切出商品列", str(issue))
        for issue in (summary.get("issues", []) or [])
    )
    if false_page_split:
        original_summary = summary
        original_output = state.raw_xlsx
        retry_output_path = state.tmp_dir / f"{state.image_path.stem}-single-page-retry.xlsx"
        retry_command = command + [
            "--page-split",
            "off",
            "--output",
            str(retry_output_path),
        ]
        retry_completed = run_command(
            retry_command,
            cwd=project_root,
            env={"PADDLE_PDX_ENABLE_MKLDNN_BYDEFAULT": "0"},
        )
        retry_output = "\n".join(
            part for part in [retry_completed.stdout, retry_completed.stderr] if part
        )
        if retry_completed.returncode != 0:
            raise RuntimeError(retry_output or f"OCR 單頁重試結束代碼：{retry_completed.returncode}")
        retry_summary = parse_json_summary(retry_output)
        if ocr_summary_quality(retry_summary) > ocr_summary_quality(original_summary):
            shutil.copy2(Path(str(retry_summary.get("output", retry_output_path))), original_output)
            summary = retry_summary
            summary["output"] = str(original_output)
            summary["page_split_retry"] = "off-selected"
        else:
            summary = original_summary
            summary["page_split_retry"] = "off-rejected"
        state.raw_xlsx = original_output
    state.vendor = str(summary.get("vendor", ""))
    state.raw_xlsx = finalize_vendor_output_name(state.raw_xlsx, state.vendor)
    summary["output"] = str(state.raw_xlsx)
    state.row_count = int(summary.get("row_count", 0) or 0)
    state.invoice_total = read_invoice_total(state.raw_xlsx)
    state.raw_xlsx_files = [state.raw_xlsx]
    state.needs_ocr_review = bool(summary.get("needs_review"))
    state.ocr_issues = list(summary.get("issues", []) or [])
    return summary


def run_ocr_batch(
    state: WorkflowState,
    project_root: Path = PROJECT_ROOT,
    progress_callback=None,
) -> dict[str, Any]:
    image_paths = state_image_paths(state)
    if not image_paths or state.output_dir is None:
        raise RuntimeError("尚未選擇圖片或輸出資料夾。")
    state.image_paths = image_paths
    state.image_path = image_paths[0]
    if len(image_paths) == 1:
        if progress_callback:
            progress_callback(1, 1, image_paths[0])
        summary = run_ocr(state, project_root)
        summary["image_count"] = 1
        summary["source_images"] = [str(image_paths[0])]
        return summary

    child_states: list[WorkflowState] = []
    child_summaries: list[dict[str, Any]] = []
    for index, image_path in enumerate(image_paths, start=1):
        if progress_callback:
            progress_callback(index, len(image_paths), image_path)
        child = WorkflowState(image_path=image_path, image_paths=[image_path], output_dir=state.output_dir)
        child_summaries.append(run_ocr(child, project_root))
        child_states.append(child)

    state.raw_xlsx_files = [
        child.raw_xlsx for child in child_states if child.raw_xlsx is not None
    ]
    missing_vendor = [
        child.image_path.name
        for child in child_states
        if child.image_path is not None and not child.vendor.strip()
    ]
    if missing_vendor:
        raise RuntimeError(
            "以下圖片無法辨識廠商，為避免誤併單已停止：\n"
            + "\n".join(f"- {name}" for name in missing_vendor)
        )
    vendor_groups: dict[str, list[str]] = {}
    for child in child_states:
        vendor_groups.setdefault(child.vendor.strip(), []).append(
            child.image_path.name if child.image_path is not None else "未知圖片"
        )
    if len(vendor_groups) != 1:
        details = []
        for vendor, names in vendor_groups.items():
            details.append(f"- {vendor}：{', '.join(names)}")
        raise RuntimeError(
            "偵測到不同廠商，依規則不可合併。請分開選取同一廠商的圖片後重跑：\n"
            + "\n".join(details)
        )

    merged_path = merge_ocr_workbooks(child_states, state.output_dir)
    state.raw_xlsx = merged_path
    state.vendor = child_states[0].vendor
    state.row_count = sum(child.row_count for child in child_states)
    combined_total = sum(
        (decimal_from_value(child.invoice_total) for child in child_states),
        Decimal("0"),
    )
    state.invoice_total = decimal_text(combined_total)
    state.needs_ocr_review = any(child.needs_ocr_review for child in child_states)
    state.ocr_issues = [
        f"{child.image_path.name}：{issue}"
        for child in child_states
        if child.image_path is not None
        for issue in child.ocr_issues
    ]
    state.tmp_dir = child_states[0].tmp_dir
    return {
        "output": str(merged_path),
        "vendor": state.vendor,
        "row_count": state.row_count,
        "invoice_total": state.invoice_total,
        "needs_review": state.needs_ocr_review,
        "issues": state.ocr_issues,
        "image_count": len(image_paths),
        "source_images": [str(path) for path in image_paths],
        "source_outputs": [str(path) for path in state.raw_xlsx_files],
        "individual_summaries": child_summaries,
    }


def run_match(state: WorkflowState, project_root: Path = PROJECT_ROOT) -> dict[str, Any]:
    if state.raw_xlsx is None or not state.raw_xlsx.exists():
        raise RuntimeError("找不到 OCR 原始 Excel，無法產品比對。")
    csv_path = product_csv()
    if csv_path is None:
        raise RuntimeError(f"找不到產品資料輸出.CSV：{REFERENCE_DIR}")
    assert_product_csv_current(csv_path)
    command = [
        str(python_exe(project_root)),
        "-X",
        "utf8",
        str(SCRIPTS_DIR / "match-existing-products.py"),
        "--input-xlsx",
        str(state.raw_xlsx),
        "--csv",
        str(csv_path),
        "--similar-threshold",
        "0.60",
        "--max-candidates",
        "5",
        "--low-similar-threshold",
        "0.55",
        "--max-low-candidates",
        "3",
        "--no-suggestion-txt",
    ]
    completed = run_command(command, cwd=project_root)
    output = "\n".join(part for part in [completed.stdout, completed.stderr] if part)
    if completed.returncode != 0:
        raise RuntimeError(output or f"產品比對結束代碼：{completed.returncode}")
    state.match_xlsx = state.raw_xlsx.with_name(f"{state.raw_xlsx.stem}_產品比對檢查.xlsx")
    state.suggested_names_txt = state.match_xlsx.with_name(f"{state.match_xlsx.stem}_建議名稱.txt")
    if not state.match_xlsx.exists():
        raise RuntimeError(f"產品比對完成但找不到檢查檔：{state.match_xlsx}")
    return {"message": output.strip(), "output": str(state.match_xlsx)}


def load_match_module() -> Any:
    path = SCRIPTS_DIR / "match-existing-products.py"
    spec = importlib.util.spec_from_file_location("match_existing_products_local", path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"無法載入產品比對腳本：{path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def find_lower_similarity_candidates(
    source_name: str,
    minimum_score: float = 0.35,
    maximum_score: float = 0.55,
    max_candidates: int = 12,
) -> list[tuple[str, str, str, str]]:
    name = str(source_name or "").strip()
    if not name:
        raise RuntimeError("這筆商品沒有可供檢查的 OCR 品名。")
    if not 0 <= minimum_score < maximum_score <= 1:
        raise RuntimeError("更低候選門檻設定錯誤。")
    csv_path = product_csv()
    if csv_path is None:
        raise RuntimeError(f"找不到產品資料輸出.CSV：{REFERENCE_DIR}")
    assert_product_csv_current(csv_path)

    matcher = load_match_module()
    identity_tokens = matcher.load_identity_tokens(csv_path.parent)
    scored: list[tuple[float, str, str]] = []
    seen_codes: set[str] = set()
    for row in matcher.read_csv_rows(csv_path):
        code = matcher.normalize_product_code(
            row.get("1.產品代號")
            or row.get("產品代號")
            or row.get("ProductCode")
            or ""
        )
        product_name = str(
            row.get("2.產品名稱")
            or row.get("產品名稱")
            or row.get("ProductName")
            or ""
        ).strip()
        if not CODE_RE.fullmatch(code) or not product_name or code in seen_codes:
            continue
        score = float(matcher.product_similarity(name, product_name, identity_tokens))
        if minimum_score <= score < maximum_score:
            scored.append((score, code, product_name))
            seen_codes.add(code)

    scored.sort(key=lambda item: (-item[0], item[1]))
    return [
        (code, product_name, f"{score:.4f}", "deep")
        for score, code, product_name in scored[:max_candidates]
    ]


def load_review_module() -> Any:
    path = SCRIPTS_DIR / "review-invoice-product-check.py"
    spec = importlib.util.spec_from_file_location("review_invoice_product_check_local", path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"無法載入覆核腳本：{path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def find_header(ws: Any, required: list[str]) -> tuple[int, dict[str, int]]:
    for row in range(1, min(ws.max_row, 20) + 1):
        headers: dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row, col).value
            if value is not None:
                headers[str(value).strip()] = col
        if all(name in headers for name in required):
            return row, headers
    raise RuntimeError("找不到必要欄位：" + "、".join(required))


def unique_batch_output_path(output_dir: Path, vendor: str) -> Path:
    safe_vendor = re.sub(r'[<>:"/\\|?*]', "", vendor).strip() or "未知廠商"
    stamp = datetime.now(ZoneInfo("Asia/Taipei")).strftime("%Y%m%d-%H%M%S")
    base = output_dir / f"{safe_vendor}進貨單-同廠商合併-{stamp}.xlsx"
    candidate = base
    serial = 2
    while candidate.exists():
        candidate = base.with_name(f"{base.stem}_{serial}{base.suffix}")
        serial += 1
    return candidate


def copy_cell_appearance(source_cell: Any, target_cell: Any) -> None:
    if source_cell.has_style:
        target_cell._style = copy(source_cell._style)
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format
    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.protection = copy(source_cell.protection)


def merge_ocr_workbooks(states: list[WorkflowState], output_dir: Path) -> Path:
    if not states:
        raise RuntimeError("沒有可合併的 OCR 結果。")
    vendors = {state.vendor.strip() for state in states if state.vendor.strip()}
    if len(vendors) != 1:
        raise RuntimeError("OCR 合併只允許同一廠商。")
    source_paths = [
        state.raw_xlsx for state in states if state.raw_xlsx is not None and state.raw_xlsx.exists()
    ]
    if len(source_paths) != len(states):
        raise RuntimeError("部分 OCR 原始 Excel 不存在，無法合併。")

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = unique_batch_output_path(output_dir, states[0].vendor)
    shutil.copy2(source_paths[0], output_path)
    destination_wb = load_workbook(output_path)
    try:
        destination_ws = (
            destination_wb["進貨明細"]
            if "進貨明細" in destination_wb.sheetnames
            else destination_wb.active
        )
        header_row, destination_headers = find_header(
            destination_ws,
            ["產品代號", "品名", "數量", "進價", "金額"],
        )
        total_row = next(
            (
                row
                for row in range(header_row + 1, destination_ws.max_row + 1)
                if str(destination_ws.cell(row, destination_headers["產品代號"]).value or "").strip()
                in SUMMARY_ROW_NAMES
            ),
            destination_ws.max_row,
        )
        total_appearance = {
            column: copy(destination_ws.cell(total_row, column)._style)
            for column in range(1, destination_ws.max_column + 1)
        }
        total_height = destination_ws.row_dimensions[total_row].height
        total_merge_spans = [
            (merged.min_col, merged.max_col)
            for merged in list(destination_ws.merged_cells.ranges)
            if merged.min_row <= total_row <= merged.max_row
        ]
        for min_column, max_column in total_merge_spans:
            destination_ws.unmerge_cells(
                start_row=total_row,
                start_column=min_column,
                end_row=total_row,
                end_column=max_column,
            )
        if destination_ws.max_row > header_row:
            destination_ws.delete_rows(header_row + 1, destination_ws.max_row - header_row)

        output_row = header_row + 1
        merged_row_count = 0
        for state, source_path in zip(states, source_paths):
            source_wb = load_workbook(source_path, data_only=False)
            try:
                source_ws = (
                    source_wb["進貨明細"]
                    if "進貨明細" in source_wb.sheetnames
                    else source_wb.active
                )
                source_header_row, source_headers = find_header(
                    source_ws,
                    ["產品代號", "品名", "數量", "進價", "金額"],
                )
                for source_row in range(source_header_row + 1, source_ws.max_row + 1):
                    product_code = str(
                        source_ws.cell(source_row, source_headers["產品代號"]).value or ""
                    ).strip()
                    product_name = str(
                        source_ws.cell(source_row, source_headers["品名"]).value or ""
                    ).strip()
                    if product_code in SUMMARY_ROW_NAMES or product_name in SUMMARY_ROW_NAMES:
                        continue
                    if not product_name:
                        continue
                    for header, destination_column in destination_headers.items():
                        source_column = source_headers.get(header)
                        if source_column is None:
                            continue
                        source_cell = source_ws.cell(source_row, source_column)
                        target_cell = destination_ws.cell(output_row, destination_column)
                        target_cell.value = source_cell.value
                        copy_cell_appearance(source_cell, target_cell)
                    destination_ws.row_dimensions[output_row].height = (
                        source_ws.row_dimensions[source_row].height
                    )
                    output_row += 1
                    merged_row_count += 1
            finally:
                source_wb.close()

        total_value = sum(
            (decimal_from_value(state.invoice_total) for state in states),
            Decimal("0"),
        )
        for column, style in total_appearance.items():
            destination_ws.cell(output_row, column)._style = copy(style)
        destination_ws.cell(output_row, destination_headers["產品代號"]).value = "總價格"
        destination_ws.cell(output_row, destination_headers["金額"]).value = (
            int(total_value) if total_value == total_value.to_integral() else float(total_value)
        )
        for min_column, max_column in total_merge_spans:
            destination_ws.merge_cells(
                start_row=output_row,
                start_column=min_column,
                end_row=output_row,
                end_column=max_column,
            )
        destination_ws.row_dimensions[output_row].height = total_height
        destination_ws.auto_filter.ref = (
            f"A{header_row}:{destination_ws.cell(output_row, destination_ws.max_column).coordinate}"
        )

        note_ws = (
            destination_wb["OCR測試紀錄"]
            if "OCR測試紀錄" in destination_wb.sheetnames
            else destination_wb.create_sheet("OCR測試紀錄")
        )
        note_ws.append([])
        note_ws.append(["同廠商批次合併", states[0].vendor])
        note_ws.append(["合併圖片張數", len(states)])
        note_ws.append(["合併商品筆數", merged_row_count])
        note_ws.append(["合併總額", decimal_text(total_value)])
        for index, state in enumerate(states, start=1):
            note_ws.append(
                [
                    f"來源圖片 {index}",
                    str(state.image_path or ""),
                    f"OCR檔案：{state.raw_xlsx}",
                    f"商品筆數：{state.row_count}",
                    f"單據總額：{state.invoice_total}",
                ]
            )
        destination_wb.save(output_path)
    finally:
        destination_wb.close()
    return output_path


def ensure_column(ws: Any, header_row: int, headers: dict[str, int], name: str) -> int:
    if name in headers:
        return headers[name]
    column = ws.max_column + 1
    ws.cell(header_row, column).value = name
    headers[name] = column
    return column


def is_summary_row_name(value: str) -> bool:
    text = str(value or "").strip()
    return not text or text in SUMMARY_ROW_NAMES or any(text.startswith(name) for name in SUMMARY_ROW_NAMES)


def worksheet_item_row_numbers(
    ws: Any,
    header_row: int,
    headers: dict[str, int],
) -> list[int]:
    rows: list[int] = []
    for row in range(header_row + 1, ws.max_row + 1):
        name = str(ws.cell(row, headers["品名"]).value or "").strip()
        product_code = str(ws.cell(row, headers["產品代號"]).value or "").strip()
        if is_summary_row_name(name) or product_code in SUMMARY_ROW_NAMES:
            continue
        rows.append(row)
    return rows


def update_worksheet_total(
    ws: Any,
    header_row: int,
    headers: dict[str, int],
) -> Decimal:
    total = sum(
        (
            decimal_from_value(ws.cell(row, headers["金額"]).value)
            for row in worksheet_item_row_numbers(ws, header_row, headers)
        ),
        Decimal("0"),
    )
    for row in range(header_row + 1, ws.max_row + 1):
        product_code = str(ws.cell(row, headers["產品代號"]).value or "").strip()
        name = str(ws.cell(row, headers["品名"]).value or "").strip()
        if product_code in SUMMARY_ROW_NAMES or name in SUMMARY_ROW_NAMES:
            ws.cell(row, headers["金額"]).value = (
                int(total) if total == total.to_integral() else float(total)
            )
            break
    return total


def parse_candidates(candidates: str) -> list[tuple[str, str, str, str]]:
    result: list[tuple[str, str, str, str]] = []
    for line in str(candidates or "").splitlines():
        text = line.strip()
        if not text:
            continue
        match = re.match(
            r"^(?:\[(一般候選|低相似候選|更低候選)\]\s*)?"
            r"(\d{6})\s+(.+?)(?:\s+\(([0-9.]+)\))?$",
            text,
        )
        if match:
            tier = {
                "低相似候選": "low",
                "更低候選": "deep",
            }.get(match.group(1), "normal")
            result.append(
                (
                    match.group(2),
                    match.group(3).strip(),
                    match.group(4) or "",
                    tier,
                )
            )
    return result


def first_candidate(candidates: str) -> tuple[str, str]:
    parsed = parse_candidates(candidates)
    return (parsed[0][0], parsed[0][1]) if parsed else ("", "")


def load_ocr_confirm_rows(state: WorkflowState) -> list[OcrConfirmRow]:
    source_path = (
        state.match_xlsx
        if state.match_xlsx is not None and state.match_xlsx.exists()
        else state.raw_xlsx
    )
    if source_path is None or not source_path.exists():
        raise RuntimeError("找不到 OCR 或產品比對檢查檔。")
    wb = load_workbook(source_path)
    ws = wb.active
    header_row, headers = find_header(ws, ["產品代號", "品名", "數量", "進價", "金額"])
    rows: list[OcrConfirmRow] = []
    for row in range(header_row + 1, ws.max_row + 1):
        raw_name = str(ws.cell(row, headers["品名"]).value or "").strip()
        product_code = str(ws.cell(row, headers["產品代號"]).value or "").strip()
        if is_summary_row_name(raw_name) or product_code in SUMMARY_ROW_NAMES:
            continue
        status = str(ws.cell(row, headers.get("比對狀態", 0)).value or "").strip() if "比對狀態" in headers else ""
        matched_code = str(ws.cell(row, headers.get("已建檔代號", 0)).value or "").strip() if "已建檔代號" in headers else ""
        matched_name = str(ws.cell(row, headers.get("已建檔品名", 0)).value or "").strip() if "已建檔品名" in headers else ""
        candidates = str(ws.cell(row, headers.get("相似候選", 0)).value or "").strip() if "相似候選" in headers else ""
        normalized_matched_code = normalize_product_code(matched_code or product_code)
        rows.append(
            OcrConfirmRow(
                excel_row=row,
                is_existing=status == "已建檔" and bool(CODE_RE.fullmatch(normalized_matched_code)),
                raw_name=raw_name,
                quantity=str(ws.cell(row, headers["數量"]).value or "").strip(),
                unit_cost=str(ws.cell(row, headers["進價"]).value or "").strip(),
                amount=str(ws.cell(row, headers["金額"]).value or "").strip(),
                matched_code=normalized_matched_code,
                matched_name=matched_name,
                candidates=candidates,
                status=status,
            )
        )
    return rows


def save_raw_ocr_rows(state: WorkflowState, rows: list[OcrConfirmRow]) -> None:
    if state.raw_xlsx is None or not state.raw_xlsx.exists():
        raise RuntimeError("找不到 OCR 原始 Excel。")
    wb = load_workbook(state.raw_xlsx)
    try:
        ws = wb["進貨明細"] if "進貨明細" in wb.sheetnames else wb.active
        header_row, headers = find_header(ws, ["產品代號", "品名", "數量", "進價", "金額"])
        original_rows = set(worksheet_item_row_numbers(ws, header_row, headers))
        kept_rows = {item.excel_row for item in rows}
        for item in rows:
            row = item.excel_row
            if row < header_row + 1 or row > ws.max_row:
                continue
            ws.cell(row, headers["品名"]).value = str(item.raw_name).strip()
            ws.cell(row, headers["數量"]).value = item.quantity
            ws.cell(row, headers["進價"]).value = item.unit_cost
            ws.cell(row, headers["金額"]).value = item.amount
        for row in sorted(original_rows - kept_rows, reverse=True):
            ws.delete_rows(row, 1)
        total = update_worksheet_total(ws, header_row, headers)
        wb.save(state.raw_xlsx)
        state.row_count = len(rows)
        state.invoice_total = decimal_text(total)
    finally:
        wb.close()


def save_ocr_confirm_rows(state: WorkflowState, rows: list[OcrConfirmRow]) -> None:
    if state.match_xlsx is None or not state.match_xlsx.exists():
        raise RuntimeError("找不到產品比對檢查檔。")
    wb = load_workbook(state.match_xlsx)
    try:
        ws = wb.active
        header_row, headers = find_header(ws, ["產品代號", "品名", "數量", "進價", "金額"])
        status_col = ensure_column(ws, header_row, headers, "比對狀態")
        matched_code_col = ensure_column(ws, header_row, headers, "已建檔代號")
        matched_name_col = ensure_column(ws, header_row, headers, "已建檔品名")
        candidate_col = ensure_column(ws, header_row, headers, "相似候選")
        original_rows = set(worksheet_item_row_numbers(ws, header_row, headers))
        kept_rows = {item.excel_row for item in rows}

        for item in rows:
            row = item.excel_row
            if row < header_row + 1 or row > ws.max_row:
                continue
            matched_code = normalize_product_code(item.matched_code)
            matched_name = str(item.matched_name).strip()
            if item.is_existing and not CODE_RE.fullmatch(matched_code):
                raise RuntimeError(f"第 {row} 列已勾選已建檔，但缺少六位產品代號。")
            if item.is_existing:
                official_name = catalog_product_name(matched_code)
                if not official_name:
                    raise RuntimeError(
                        f"第 {row} 列已勾選已建檔，但產品代號 {matched_code} 不存在於今天的產品資料。"
                    )
                matched_name = official_name

            ws.cell(row, headers["品名"]).value = str(item.raw_name).strip()
            ws.cell(row, headers["數量"]).value = item.quantity
            ws.cell(row, headers["進價"]).value = item.unit_cost
            ws.cell(row, headers["金額"]).value = item.amount
            ws.cell(row, candidate_col).value = str(item.candidates or "").strip()
            if item.is_existing:
                ws.cell(row, headers["產品代號"]).value = matched_code
                ws.cell(row, headers["產品代號"]).number_format = "@"
                ws.cell(row, status_col).value = "已建檔"
                ws.cell(row, matched_code_col).value = matched_code
                ws.cell(row, matched_name_col).value = matched_name
            else:
                ws.cell(row, headers["產品代號"]).value = ""
                ws.cell(row, headers["產品代號"]).number_format = "@"
                if str(ws.cell(row, status_col).value or "").strip() == "已建檔":
                    ws.cell(row, status_col).value = "有類似產品" if str(item.candidates).strip() else "確認為新品"
                ws.cell(row, matched_code_col).value = ""
                ws.cell(row, matched_name_col).value = ""
        for row in sorted(original_rows - kept_rows, reverse=True):
            ws.delete_rows(row, 1)
        total = update_worksheet_total(ws, header_row, headers)
        wb.save(state.match_xlsx)
        state.row_count = len(rows)
        state.invoice_total = decimal_text(total)
    finally:
        wb.close()


def prepare_review_table(state: WorkflowState) -> list[AdjustmentRow]:
    if state.match_xlsx is None or not state.match_xlsx.exists():
        raise RuntimeError("找不到產品比對檢查檔。")
    csv_path = product_csv()
    if csv_path is None:
        raise RuntimeError(f"找不到產品資料輸出.CSV：{REFERENCE_DIR}")
    review = load_review_module()
    wb = load_workbook(state.match_xlsx)
    ws = wb.active
    header_row, headers = review.find_header_row(ws)
    rows = review.item_rows(ws, header_row, headers)
    excluded_rows = [row for row in rows if review.is_excluded_item_row(ws, row, headers)]
    rows = [row for row in rows if row not in excluded_rows]
    review.fill_existing_product_codes(ws, rows, headers)

    catalog_rows = review.read_catalog(csv_path)
    catalog_names = review.catalog_product_names(catalog_rows)
    brand_rules_path = REFERENCE_DIR / "品牌括號命名規則.csv"
    brand_rule_rows = review.load_brand_rule_rows(brand_rules_path)
    brand_rules = review.load_brand_rules(brand_rules_path)
    category_rules = category_rule_rows()
    name_col = headers["品名"]
    category_col = headers["大類"]
    code_col = headers["產品代號"]

    for row in rows:
        code_cell = ws.cell(row, code_col)
        code_text = "" if code_cell.value is None else str(code_cell.value).strip()
        code_cell.value = code_text
        code_cell.number_format = "@"

        raw_name = str(ws.cell(row, name_col).value or "").strip()
        matched_name_col = headers.get("已建檔品名")
        if review.is_existing_row(ws, row, headers) and matched_name_col:
            matched_name = str(ws.cell(row, matched_name_col).value or "").strip()
            adjusted_name = matched_name or raw_name
        elif review.is_existing_row(ws, row, headers):
            adjusted_name = raw_name
        elif review.CODE_RE.match(code_text) and review.PREFIX_RE.match(raw_name):
            adjusted_name = raw_name
        else:
            adjusted_name = review.adjust_name(raw_name, catalog_names, brand_rules, brand_rule_rows)
        ws.cell(row, name_col).value = adjusted_name

        if not review.is_existing_row(ws, row, headers):
            category_cell = ws.cell(row, category_col)
            if category_cell.value not in (None, ""):
                category_value = review.infer_category_value(category_cell.value, adjusted_name, catalog_rows, category_rules)
            else:
                category_value = review.infer_category_from_brand_rules(adjusted_name, brand_rule_rows)
                if category_value in (None, ""):
                    category_value = review.infer_category_value(category_cell.value, adjusted_name, catalog_rows, category_rules)
            if category_value not in (None, ""):
                category_cell.value = category_value

    wb.save(state.match_xlsx)
    state.excluded_items = [str(ws.cell(row, name_col).value or "").strip() for row in excluded_rows]
    return load_adjustment_rows(state)


def load_adjustment_rows(state: WorkflowState) -> list[AdjustmentRow]:
    if state.match_xlsx is None:
        return []
    wb = load_workbook(state.match_xlsx)
    ws = wb.active
    header_row, headers = find_header(ws, ["產品代號", "品名", "大類", "數量", "進價", "金額"])
    names = category_name_map()
    rows: list[AdjustmentRow] = []
    for row in range(header_row + 1, ws.max_row + 1):
        name = str(ws.cell(row, headers["品名"]).value or "").strip()
        code = str(ws.cell(row, headers["產品代號"]).value or "").strip()
        if not name or name in {"總價格", "總價", "總計", "合計", "小計"} or code in {"總價格", "總價", "總計", "合計", "小計"}:
            continue
        text_for_exclusion = "\n".join(
            str(ws.cell(row, headers[h]).value or "") for h in ("品名", "已建檔品名", "相似候選") if h in headers
        )
        if any(keyword in text_for_exclusion for keyword in ("一番賞", "抽賞", "Ichiban Kuji", "ICHIBAN KUJI", "遮蔽", "已遮蔽", "人工確認重複", "重複品項")):
            continue
        status = str(ws.cell(row, headers.get("比對狀態", 0)).value or "").strip() if "比對狀態" in headers else ""
        if status == "已建檔":
            continue
        category = str(ws.cell(row, headers["大類"]).value or "").strip()
        rows.append(
            AdjustmentRow(
                row_id=str(row),
                excel_row=row,
                source_row=row,
                product_code=code,
                name=name,
                category=category_code(category),
                category_display=category_display(category, names),
                quantity=str(ws.cell(row, headers["數量"]).value or "").strip(),
                unit_cost=str(ws.cell(row, headers["進價"]).value or "").strip(),
                amount=str(ws.cell(row, headers["金額"]).value or "").strip(),
                status=status,
            )
        )
    return rows


def adjustment_candidate_row_numbers(
    ws: Any,
    header_row: int,
    headers: dict[str, int],
) -> list[int]:
    rows: list[int] = []
    for row in range(header_row + 1, ws.max_row + 1):
        name = str(ws.cell(row, headers["品名"]).value or "").strip()
        code = str(ws.cell(row, headers["產品代號"]).value or "").strip()
        if is_summary_row_name(name) or code in SUMMARY_ROW_NAMES:
            continue
        text_for_exclusion = "\n".join(
            str(ws.cell(row, headers[h]).value or "")
            for h in ("品名", "已建檔品名", "相似候選")
            if h in headers
        )
        if any(
            keyword in text_for_exclusion
            for keyword in (
                "一番賞",
                "抽賞",
                "Ichiban Kuji",
                "ICHIBAN KUJI",
                "遮蔽",
                "已遮蔽",
                "人工確認重複",
                "重複品項",
            )
        ):
            continue
        status = (
            str(ws.cell(row, headers["比對狀態"]).value or "").strip()
            if "比對狀態" in headers
            else ""
        )
        if status != "已建檔":
            rows.append(row)
    return rows


def validate_split_totals(ws: Any, headers: dict[str, int], rows: list[AdjustmentRow]) -> None:
    grouped: dict[int, list[AdjustmentRow]] = {}
    for item in rows:
        if item.source_row and item.source_row != item.excel_row:
            grouped.setdefault(item.source_row, []).append(item)
    for source_row, split_rows in grouped.items():
        original_quantity = str(ws.cell(source_row, headers["數量"]).value or "").strip()
        original_amount = str(ws.cell(source_row, headers["金額"]).value or "").strip()
        source_current = next((item for item in rows if item.excel_row == source_row), None)
        group = ([source_current] if source_current else []) + split_rows
        try:
            original_quantity_value = float(original_quantity)
            split_quantity_value = sum(float(str(item.quantity or "0")) for item in group)
        except ValueError:
            original_quantity_value = split_quantity_value = 0
        try:
            original_amount_value = float(original_amount)
            split_amount_value = sum(float(str(item.amount or "0")) for item in group)
        except ValueError:
            original_amount_value = split_amount_value = 0
        if abs(original_quantity_value - split_quantity_value) > 0.0001 or abs(original_amount_value - split_amount_value) > 0.0001:
            name = str(ws.cell(source_row, headers["品名"]).value or "").strip()
            raise RuntimeError(
                f"第 {source_row} 列「{name}」已拆分，但拆分後數量或金額合計不一致。\n"
                f"原數量/金額：{original_quantity} / {original_amount}\n"
                f"拆分後數量/金額：{split_quantity_value:g} / {split_amount_value:g}"
            )


def save_adjustment_rows(state: WorkflowState, rows: list[AdjustmentRow]) -> None:
    if state.match_xlsx is None or not state.match_xlsx.exists():
        raise RuntimeError("找不到產品比對檢查檔。")
    valid_categories = category_name_map()
    invalid_categories = [
        (item.name, category_code(item.category))
        for item in rows
        if category_code(item.category)
        and category_code(item.category) not in valid_categories
    ]
    if invalid_categories:
        preview = "\n".join(
            f"- {name}｜大類代號 {code}"
            for name, code in invalid_categories[:12]
        )
        raise RuntimeError(
            "以下商品使用不存在或已刪除的大類，請重新選擇：\n" + preview
        )
    wb = load_workbook(state.match_xlsx)
    try:
        ws = wb.active
        header_row, headers = find_header(ws, ["產品代號", "品名", "大類", "數量", "進價", "金額"])
        validate_split_totals(ws, headers, rows)
        original_adjustment_rows = set(
            adjustment_candidate_row_numbers(ws, header_row, headers)
        )
        kept_existing_rows = {
            item.excel_row
            for item in rows
            if item.excel_row > 0 and item.excel_row <= ws.max_row
        }
        new_rows: list[AdjustmentRow] = []
        for item in rows:
            row = item.excel_row
            if row <= 0 or row > ws.max_row:
                new_rows.append(item)
                continue
            ws.cell(row, headers["產品代號"]).value = str(item.product_code).strip()
            ws.cell(row, headers["產品代號"]).number_format = "@"
            ws.cell(row, headers["品名"]).value = str(item.name).strip()
            ws.cell(row, headers["大類"]).value = category_code(item.category)
            ws.cell(row, headers["數量"]).value = item.quantity
            ws.cell(row, headers["進價"]).value = item.unit_cost
            ws.cell(row, headers["金額"]).value = item.amount
            if "比對狀態" in headers:
                ws.cell(row, headers["比對狀態"]).value = item.status or "確認為新品"

        for row in sorted(original_adjustment_rows - kept_existing_rows, reverse=True):
            ws.delete_rows(row, 1)

        for item in new_rows:
            row = ws.max_row + 1
            item.excel_row = row
            ws.cell(row, headers["產品代號"]).value = str(item.product_code).strip()
            ws.cell(row, headers["產品代號"]).number_format = "@"
            ws.cell(row, headers["品名"]).value = str(item.name).strip()
            ws.cell(row, headers["大類"]).value = category_code(item.category)
            ws.cell(row, headers["數量"]).value = item.quantity
            ws.cell(row, headers["進價"]).value = item.unit_cost
            ws.cell(row, headers["金額"]).value = item.amount
            if "比對狀態" in headers:
                ws.cell(row, headers["比對狀態"]).value = item.status or "確認為新品"

        total = update_worksheet_total(ws, header_row, headers)
        wb.save(state.match_xlsx)
        state.invoice_total = decimal_text(total)
    finally:
        wb.close()


def generate_adjusted_xlsx(state: WorkflowState, project_root: Path = PROJECT_ROOT) -> dict[str, Any]:
    if state.match_xlsx is None or not state.match_xlsx.exists():
        raise RuntimeError("找不到產品比對檢查檔。")
    csv_path = product_csv()
    if csv_path is None:
        raise RuntimeError(f"找不到產品資料輸出.CSV：{REFERENCE_DIR}")
    command = [
        str(python_exe(project_root)),
        "-X",
        "utf8",
        str(SCRIPTS_DIR / "review-invoice-product-check.py"),
        "--input-xlsx",
        str(state.match_xlsx),
        "--csv",
        str(csv_path),
        "--brand-rules",
        str(REFERENCE_DIR / "品牌括號命名規則.csv"),
        "--category-rules",
        str(runtime_category_rules_path()),
        "--confirmed",
    ]
    completed = run_command(command, cwd=project_root)
    output = "\n".join(part for part in [completed.stdout, completed.stderr] if part)
    if completed.returncode != 0:
        raise RuntimeError(output or f"覆核調整結束代碼：{completed.returncode}")
    state.adjusted_xlsx = state.match_xlsx.with_name(f"{state.match_xlsx.stem}[調整]{state.match_xlsx.suffix}")
    if not state.adjusted_xlsx.exists():
        raise RuntimeError(f"覆核調整完成但找不到調整檔：{state.adjusted_xlsx}")
    return {"message": output.strip(), "output": str(state.adjusted_xlsx)}


def build_import_files(state: WorkflowState, project_root: Path = PROJECT_ROOT) -> dict[str, Any]:
    if state.adjusted_xlsx is None or not state.adjusted_xlsx.exists():
        raise RuntimeError("找不到 [調整].xlsx。")
    output_dir = state.formal_output_dir or (project_root / "建檔進貨用")
    command = [
        "powershell",
        "-NoProfile",
        "-ExecutionPolicy",
        "Bypass",
        "-File",
        str(SCRIPTS_DIR / "fill-import-templates.ps1"),
        "-WorkspaceRoot",
        str(project_root),
        "-ProductsXlsx",
        str(state.adjusted_xlsx),
        "-OutputDir",
        str(output_dir),
        "-NewProductTemplate",
        str(REFERENCE_DIR / "建檔用.xls"),
        "-PurchaseTemplate",
        str(REFERENCE_DIR / "採購單匯入範例.xls"),
        "-InvoiceTotal",
        str(state.invoice_total or read_invoice_total(state.adjusted_xlsx)),
        "-ConfirmedReviewed",
    ]
    completed = run_command(command, cwd=project_root)
    output = "\n".join(part for part in [completed.stdout, completed.stderr] if part)
    if completed.returncode != 0:
        raise RuntimeError(output or f"正式輸出結束代碼：{completed.returncode}")
    summary = parse_json_summary(output)
    new_file = summary.get("newProductFile")
    purchase_file = summary.get("purchaseImportFile")
    state.new_product_file = Path(new_file) if new_file else None
    state.purchase_file = Path(str(purchase_file)) if purchase_file else None
    state.build_summary = summary
    if state.new_product_file and not state.new_product_file.exists():
        raise RuntimeError(f"找不到建檔用成品：{state.new_product_file}")
    if state.purchase_file is None or not state.purchase_file.exists():
        raise RuntimeError(f"找不到採購單成品：{state.purchase_file}")
    processed_paths: list[Path] = []
    for source_path in state_image_paths(state):
        if source_path.exists():
            processed_paths.append(mark_source_processed(source_path))
    if processed_paths:
        state.processed_image_paths = processed_paths
        state.processed_image_path = processed_paths[0]
        state.image_paths = processed_paths
        state.image_path = processed_paths[0]
        summary["sourceProcessedPath"] = str(processed_paths[0])
        summary["sourceProcessedPaths"] = [str(path) for path in processed_paths]
    internal_adjusted = state.adjusted_xlsx
    try:
        internal_adjusted.unlink()
    except OSError as exc:
        summary["internalAdjustedFileRemoved"] = False
        summary["internalAdjustedCleanupWarning"] = str(exc)
    else:
        summary["internalAdjustedFileRemoved"] = True
        state.adjusted_xlsx = None
    return summary


def mark_source_processed(source: Path) -> Path:
    if "_已處理" in source.stem:
        return source
    candidate = source.with_name(f"{source.stem}_已處理{source.suffix}")
    serial = 2
    while candidate.exists():
        candidate = source.with_name(f"{source.stem}_已處理_{serial}{source.suffix}")
        serial += 1
    source.rename(candidate)
    return candidate


def cleanup_intermediate_files(state: WorkflowState) -> list[Path]:
    if not ((state.new_product_file and state.new_product_file.exists()) or (state.purchase_file and state.purchase_file.exists())):
        raise RuntimeError("正式成品不存在，不清理中間檔。")
    candidates: list[Path] = []
    for path in [
        *state.raw_xlsx_files,
        state.raw_xlsx,
        state.match_xlsx,
        state.suggested_names_txt,
        state.adjusted_xlsx,
    ]:
        if path is not None:
            candidates.append(path)
    if state.match_xlsx is not None:
        candidates.extend(state.match_xlsx.parent.glob(f"{state.match_xlsx.stem}*.txt"))
    deleted: list[Path] = []
    workspace_root = PROJECT_ROOT.resolve()
    protected = {
        path.resolve()
        for path in [
            *state.image_paths,
            *state.processed_image_paths,
            state.image_path,
            state.processed_image_path,
            state.new_product_file,
            state.purchase_file,
        ]
        if path is not None and path.exists()
    }
    for path in candidates:
        try:
            resolved = path.resolve()
            if (
                path.exists()
                and path.is_file()
                and resolved not in protected
                and resolved.is_relative_to(workspace_root)
                and not resolved.is_relative_to((PROJECT_ROOT / "參考資料").resolve())
            ):
                path.unlink()
                deleted.append(path)
        except OSError:
            pass
    if state.tmp_dir and state.tmp_dir.exists():
        try:
            resolved_tmp = state.tmp_dir.resolve()
            allowed_tmp = (PROJECT_ROOT / ".codex-tmp").resolve()
            if resolved_tmp.is_relative_to(allowed_tmp):
                shutil.rmtree(state.tmp_dir)
                deleted.append(state.tmp_dir)
        except OSError:
            pass
    return deleted
