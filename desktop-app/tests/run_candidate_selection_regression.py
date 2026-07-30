from __future__ import annotations

import json
import sys
from pathlib import Path
from tkinter import Listbox, Toplevel
from types import SimpleNamespace

from openpyxl import Workbook, load_workbook


APP_DIR = Path(__file__).resolve().parents[1]
if str(APP_DIR) not in sys.path:
    sys.path.insert(0, str(APP_DIR))

import invoice_workflow as workflow
import invoice_ocr_excel_gui as gui_module
from invoice_ocr_excel_gui import InvoiceOcrApp
from invoice_workflow import AdjustmentRow, OcrConfirmRow, WorkflowState


def child_widgets(widget):
    for child in widget.winfo_children():
        yield child
        yield from child_widgets(child)


def catalog_samples(count: int = 3) -> list[tuple[str, str]]:
    csv_path = workflow.product_csv()
    if csv_path is None:
        raise RuntimeError("找不到產品資料 CSV。")
    result: list[tuple[str, str]] = []
    for row in workflow.read_csv_dicts(csv_path):
        code = workflow.normalize_product_code(row.get("1.產品代號", ""))
        name = str(row.get("2.產品名稱", "")).strip()
        if code and name:
            result.append((code, name))
        if len(result) >= count:
            return result
    raise RuntimeError("產品資料不足，無法建立候選測試。")


def create_match_workbook(path: Path, candidates: str) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "進貨明細"
    worksheet.append(["廠商：測試廠商"])
    worksheet.append(
        [
            "產品代號",
            "品名",
            "數量",
            "進價",
            "金額",
            "比對狀態",
            "已建檔代號",
            "已建檔品名",
            "相似候選",
        ]
    )
    worksheet.append(["", "OCR 測試品名", 2, 100, 200, "有類似產品", "", "", candidates])
    worksheet.append(["總價格", "", "", "", 200])
    workbook.save(path)


def create_delete_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "進貨明細"
    worksheet.append(["廠商：測試廠商"])
    worksheet.append(["產品代號", "品名", "數量", "進價", "金額"])
    worksheet.append(["", "刪除測試 1", 1, 100, 100])
    worksheet.append(["", "刪除測試 2", 1, 200, 200])
    worksheet.append(["", "刪除測試 3", 1, 300, 300])
    worksheet.append(["總價格", "", "", "", 600])
    workbook.save(path)


def create_adjustment_delete_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "進貨明細"
    worksheet.append(["廠商：測試廠商"])
    worksheet.append(
        ["產品代號", "品名", "大類", "數量", "進價", "金額", "比對狀態"]
    )
    worksheet.append(["", "調整刪除測試 1", "1", 1, 100, 100, "確認為新品"])
    worksheet.append(["", "調整刪除測試 2", "31", 1, 200, 200, "確認為新品"])
    worksheet.append(["", "調整刪除測試 3", "38", 1, 300, 300, "確認為新品"])
    worksheet.append(["總價格", "", "", "", "", 600, ""])
    workbook.save(path)


def main() -> int:
    output_dir = workflow.PROJECT_ROOT / ".codex-tmp" / "invoice-app-tests" / "candidate-selection"
    output_dir.mkdir(parents=True, exist_ok=True)
    if (workflow.REFERENCE_DIR / "大類清單.csv").exists():
        raise RuntimeError("APP 仍依賴 reference_data/大類清單.csv。")
    default_categories = workflow.category_rule_rows()
    if len(default_categories) < 30 or not workflow.category_name_map():
        raise RuntimeError("APP 內建大類沒有正確載入。")
    original_config_path = workflow.CONFIG_PATH
    original_settings = dict(workflow.APP_SETTINGS)
    category_settings_path = output_dir / "app_settings-category-test.json"
    workflow.CONFIG_PATH = category_settings_path
    workflow.APP_SETTINGS.clear()
    workflow.APP_SETTINGS.update(original_settings)
    try:
        test_categories = workflow.category_rule_rows()
        test_categories.append(
            {
                "code": "99",
                "name": "測試新增大類",
                "label": "測試新增大類",
                "keywords": "測試新增大類",
                "note": "回歸測試",
            }
        )
        workflow.save_category_rule_rows(test_categories)
        if workflow.category_name_map().get("99") != "測試新增大類":
            raise RuntimeError("新增大類沒有保存到 APP 設定。")
        workflow.save_category_rule_rows(
            [row for row in workflow.category_rule_rows() if row["code"] != "99"]
        )
        if "99" in workflow.category_name_map():
            raise RuntimeError("刪除大類沒有保存到 APP 設定。")
    finally:
        workflow.CONFIG_PATH = original_config_path
        workflow.APP_SETTINGS.clear()
        workflow.APP_SETTINGS.update(original_settings)

    workbook_path = output_dir / "候選選擇測試_產品比對檢查.xlsx"
    samples = catalog_samples()
    candidate_text = "\n".join(
        [
            f"[一般候選] {samples[0][0]} {samples[0][1]} (0.90)",
            f"[一般候選] {samples[1][0]} {samples[1][1]} (0.80)",
            f"[低相似候選] {samples[2][0]} {samples[2][1]} (0.57)",
        ]
    )
    create_match_workbook(workbook_path, candidate_text)

    parsed = workflow.parse_candidates(candidate_text)
    if len(parsed) != 3 or [candidate[3] for candidate in parsed] != [
        "normal",
        "normal",
        "low",
    ]:
        raise RuntimeError(f"候選解析筆數錯誤：{len(parsed)}")
    old_candidate_text = f"{samples[0][0]} {samples[0][1]} (0.90)"
    old_parsed = workflow.parse_candidates(old_candidate_text)
    if len(old_parsed) != 1 or old_parsed[0][3] != "normal":
        raise RuntimeError("舊版未標示層級的候選格式無法相容解析。")
    deep_parsed = workflow.parse_candidates(
        f"[更低候選] {samples[2][0]} {samples[2][1]} (0.42)"
    )
    if len(deep_parsed) != 1 or deep_parsed[0][3] != "deep":
        raise RuntimeError("更低候選格式無法解析。")

    state = WorkflowState(match_xlsx=workbook_path)
    chosen_code, chosen_name = samples[1]
    workflow.save_ocr_confirm_rows(
        state,
        [
            OcrConfirmRow(
                excel_row=3,
                is_existing=True,
                raw_name="OCR 測試品名",
                quantity="2",
                unit_cost="100",
                amount="200",
                matched_code=chosen_code,
                matched_name="",
                candidates=candidate_text,
                status="有類似產品",
            )
        ],
    )

    workbook = load_workbook(workbook_path, data_only=True)
    worksheet = workbook["進貨明細"]
    saved_code = str(worksheet.cell(3, 1).value or "").strip()
    saved_status = str(worksheet.cell(3, 6).value or "").strip()
    saved_name = str(worksheet.cell(3, 8).value or "").strip()
    workbook.close()
    if (saved_code, saved_status, saved_name) != (chosen_code, "已建檔", chosen_name):
        raise RuntimeError(
            f"手動已建檔帶入失敗：{saved_code}, {saved_status}, {saved_name}"
        )

    auto_check_path = output_dir / "自動已建檔勾選測試.xlsx"
    create_match_workbook(auto_check_path, candidate_text)
    auto_check_workbook = load_workbook(auto_check_path)
    auto_check_ws = auto_check_workbook["進貨明細"]
    auto_check_ws.cell(3, 1).value = int(chosen_code)
    auto_check_ws.cell(3, 6).value = "已建檔"
    auto_check_ws.cell(3, 7).value = int(chosen_code)
    auto_check_ws.cell(3, 8).value = chosen_name
    auto_check_workbook.save(auto_check_path)
    auto_check_workbook.close()
    auto_check_rows = workflow.load_ocr_confirm_rows(
        WorkflowState(match_xlsx=auto_check_path)
    )
    if (
        len(auto_check_rows) != 1
        or not auto_check_rows[0].is_existing
        or auto_check_rows[0].matched_code != chosen_code
    ):
        raise RuntimeError(f"自動已建檔代號沒有正確勾選：{auto_check_rows}")

    app = InvoiceOcrApp()
    app.withdraw()
    try:
        app.open_category_manager()
        app.update_idletasks()
        category_dialogs = [
            widget
            for widget in app.winfo_children()
            if isinstance(widget, Toplevel) and widget.title() == "大類管理"
        ]
        if len(category_dialogs) != 1:
            raise RuntimeError("頂端大類選單沒有開啟大類管理視窗。")
        category_trees = [
            widget
            for widget in child_widgets(category_dialogs[0])
            if isinstance(widget, gui_module.ttk.Treeview)
        ]
        if (
            len(category_trees) != 1
            or len(category_trees[0].get_children()) != len(default_categories)
        ):
            raise RuntimeError("大類管理視窗沒有完整顯示 APP 內建大類。")
        category_dialogs[0].destroy()

        if Path(app.product_csv_path.get()).resolve() != workflow.product_csv().resolve():
            raise RuntimeError("產品資料輸出路徑欄沒有載入目前 CSV。")
        app.vendor_name.set("麗嬰國際股份有限公司")
        if app.vendor_name.get() != "麗嬰國際股份有限公司":
            raise RuntimeError("偵測廠商欄無法更新。")
        reminder_calls: list[tuple[str, str]] = []
        original_showinfo = gui_module.messagebox.showinfo
        gui_module.messagebox.showinfo = (
            lambda title, message, **_kwargs: reminder_calls.append((title, message))
        )
        try:
            app.state = WorkflowState(
                raw_xlsx=workbook_path,
                vendor="測試廠商",
                row_count=1,
            )
            app.ocr_finished()
        finally:
            gui_module.messagebox.showinfo = original_showinfo
        if len(reminder_calls) != 1 or reminder_calls[0][0] != "OCR 完成":
            raise RuntimeError(f"OCR 完成提醒沒有正確觸發：{reminder_calls}")
        if "請確認產品輸出資料為最新後再繼續" not in reminder_calls[0][1]:
            raise RuntimeError(f"OCR 完成提醒缺少產品資料提示：{reminder_calls}")
        if str(app.match_button.cget("state")) != "normal":
            raise RuntimeError("OCR 完成後沒有啟用產品重複比對按鈕。")
        app.product_match_ready = True
        app.populate_ocr_tree(
            [
                OcrConfirmRow(
                    excel_row=3,
                    is_existing=False,
                    raw_name="OCR 測試品名",
                    quantity="2",
                    unit_cost="100",
                    amount="200",
                    matched_code="",
                    matched_name="",
                    candidates=candidate_text,
                    status="有類似產品",
                )
            ]
        )
        app.tree.selection_set("3")
        collected = app.collect_ocr_rows()
        if len(collected) != 1 or collected[0].candidates != candidate_text:
            raise RuntimeError("表格摘要顯示後遺失完整候選內容。")
        candidate_cell = str(app.tree.item("3", "values")[8])
        if "一般 2" not in candidate_cell or "低相似 1" not in candidate_cell:
            raise RuntimeError(f"候選摘要未顯示雙層筆數：{candidate_cell}")
        if str(app.tree.item("3", "values")[0]) != "☐":
            raise RuntimeError("相似候選被自動勾選為已建檔。")
        app.open_candidate_picker()
        app.update_idletasks()
        dialogs = [widget for widget in app.winfo_children() if isinstance(widget, Toplevel)]
        if len(dialogs) != 1:
            raise RuntimeError(f"候選視窗數量錯誤：{len(dialogs)}")
        listboxes = [widget for widget in child_widgets(dialogs[0]) if isinstance(widget, Listbox)]
        if len(listboxes) != 1 or listboxes[0].size() != 3:
            raise RuntimeError("候選視窗沒有完整顯示三筆候選。")
        listbox_values = listboxes[0].get(0, "end")
        if (
            sum("[一般候選]" in value for value in listbox_values) != 2
            or sum("[低相似候選]" in value for value in listbox_values) != 1
        ):
            raise RuntimeError(f"候選視窗沒有正確標示雙層候選：{listbox_values}")
        if not listboxes[0].bind("<MouseWheel>"):
            raise RuntimeError("候選清單沒有滑鼠滾輪綁定。")
        dialogs[0].destroy()

        batch_rows = [
            OcrConfirmRow(
                excel_row=10 + index,
                is_existing=False,
                raw_name=f"批次代號測試 {index + 1}",
                quantity="1",
                unit_cost="100",
                amount="100",
                matched_code="",
                matched_name="",
                candidates="",
                status="確認為新品",
            )
            for index in range(3)
        ]
        app.populate_ocr_tree(batch_rows)
        batch_text = "\r\n".join(f'="{code}"' for code, _name in samples)
        batch_result = app.apply_pasted_product_codes("10", batch_text)
        if batch_result["applied"] != 3 or batch_result["overflow"] != 0:
            raise RuntimeError(f"批次代號填入筆數錯誤：{batch_result}")
        for index, (code, name) in enumerate(samples):
            values = list(app.tree.item(str(10 + index), "values"))
            if (values[0], values[5], values[6], values[7]) != ("☑", code, name, "已建檔"):
                raise RuntimeError(f"批次代號第 {index + 1} 筆帶入失敗：{values}")
        ocr_column_samples = {
            1: ["OCR整欄名稱1", "OCR整欄名稱2", "OCR整欄名稱3"],
            2: ["7", "", "9"],
            3: ["110", "220", "330"],
            4: ["770", "0", "2970"],
            6: ["已建檔名稱1", "已建檔名稱2", "已建檔名稱3"],
        }
        for column_index, expected in ocr_column_samples.items():
            result = app.apply_pasted_column_values(
                "10",
                column_index,
                "\n".join(expected),
            )
            if result["applied"] != 3:
                raise RuntimeError(f"OCR 欄位 {column_index} 整欄貼上失敗：{result}")
            actual = [
                str(list(app.tree.item(str(10 + index), "values"))[column_index])
                for index in range(3)
            ]
            if actual != expected:
                raise RuntimeError(
                    f"OCR 欄位 {column_index} 整欄貼上內容錯誤：{actual}"
                )

        app.set_adjustment_table()
        if not app.tree.bind("<ButtonRelease-1>"):
            raise RuntimeError("名稱調整表沒有單擊選取格子的綁定。")
        app.populate_adjustment_tree(
            [
                AdjustmentRow(
                    row_id=str(20 + index),
                    excel_row=20 + index,
                    source_row=20 + index,
                    product_code="",
                    name=f"名稱調整測試 {index + 1}",
                    category="20",
                    category_display="20 遊戲軟體",
                    quantity="1",
                    unit_cost="100",
                    amount="100",
                    status="確認為新品",
                )
                for index in range(3)
            ]
        )
        adjustment_result = app.apply_pasted_product_codes(
            "20",
            "\n".join(code for code, _name in samples),
        )
        if adjustment_result["applied"] != 3:
            raise RuntimeError(f"名稱調整階段批次代號填入失敗：{adjustment_result}")
        for index, (code, _name) in enumerate(samples):
            values = list(app.tree.item(str(20 + index), "values"))
            if values[0] != code:
                raise RuntimeError(f"名稱調整階段第 {index + 1} 筆代號錯誤：{values}")

        pasted_product_names = [f"整欄產品名稱 {index + 1}" for index in range(3)]
        name_paste_result = app.apply_pasted_column_values(
            "20",
            1,
            "\n".join(pasted_product_names),
        )
        if name_paste_result["applied"] != 3:
            raise RuntimeError(f"產品名稱整欄貼上失敗：{name_paste_result}")
        actual_product_names = [
            str(list(app.tree.item(str(20 + index), "values"))[1])
            for index in range(3)
        ]
        if actual_product_names != pasted_product_names:
            raise RuntimeError(f"產品名稱整欄貼上內容錯誤：{actual_product_names}")
        adjustment_numeric_samples = {
            4: ["4", "", "6"],
            5: ["125", "250", "375"],
            6: ["500", "0", "2250"],
        }
        for column_index, expected in adjustment_numeric_samples.items():
            result = app.apply_pasted_column_values(
                "20",
                column_index,
                "\n".join(expected),
            )
            if result["applied"] != 3:
                raise RuntimeError(f"調整欄位 {column_index} 整欄貼上失敗：{result}")
            actual = [
                str(list(app.tree.item(str(20 + index), "values"))[column_index])
                for index in range(3)
            ]
            if actual != expected:
                raise RuntimeError(
                    f"調整欄位 {column_index} 整欄貼上內容錯誤：{actual}"
                )

        category_samples = list(workflow.category_name_map().items())[:3]
        if len(category_samples) < 3:
            raise RuntimeError("大類清單不足，無法執行大類整欄貼上測試。")
        category_code_result = app.apply_pasted_column_values(
            "20",
            2,
            "\n".join(code for code, _name in category_samples),
        )
        if category_code_result["applied"] != 3:
            raise RuntimeError(f"大類代號整欄貼上失敗：{category_code_result}")
        for index, (code, name) in enumerate(category_samples):
            values = list(app.tree.item(str(20 + index), "values"))
            if (values[2], values[3]) != (code, name):
                raise RuntimeError(f"大類代號／名稱拆欄帶入錯誤：{values}")

        app.select_column_for_copy("產品名稱")
        app.copy_selected_column()
        copied_raw = app.clipboard_get()
        copied_names = copied_raw.splitlines()
        expected_copied_names = pasted_product_names
        if copied_names != expected_copied_names:
            raise RuntimeError(f"整欄複製內容錯誤：{copied_names}")
        if "\n\n" in copied_raw or "\r\n\r\n" in copied_raw:
            raise RuntimeError("整欄複製在商品名稱之間產生空白列。")
        if not app.tree.bind("<Control-Shift-C>") or not app.tree.bind("<Button-3>"):
            raise RuntimeError("整欄複製快捷鍵或右鍵選單沒有綁定。")

        app.state = WorkflowState(
            raw_xlsx=workbook_path,
            match_xlsx=workbook_path,
        )
        app.workflow_stage = "ready"
        app.refresh_navigation_buttons()
        if str(app.back_button.cget("state")) != "normal":
            raise RuntimeError("正式輸出前沒有啟用回到上一步。")
        app.go_back()
        if app.workflow_stage != "adjust" or str(app.build_button.cget("state")) != "disabled":
            raise RuntimeError("正式輸出前返回名稱調整失敗。")
        app.go_back()
        if app.workflow_stage != "match" or len(app.adjustment_draft_rows or []) != 3:
            raise RuntimeError("名稱調整返回已建檔勾選時沒有保留草稿。")
        app.prepare_review()
        if app.workflow_stage != "adjust" or len(app.tree.get_children()) != 3:
            raise RuntimeError("返回後沒有恢復名稱調整草稿。")
        restored_names = [
            str(list(app.tree.item(item, "values"))[1])
            for item in app.tree.get_children()
        ]
        if restored_names != expected_copied_names:
            raise RuntimeError(f"恢復的名稱調整草稿錯誤：{restored_names}")

        app.deiconify()
        app.tree.see("20")
        app.update()
        cell_box = app.tree.bbox("20", "#2")
        if not cell_box:
            raise RuntimeError("無法取得名稱調整格子位置。")
        x, y, width, height = cell_box
        app.select_adjustment_cell(
            SimpleNamespace(x=x + width // 2, y=y + height // 2)
        )
        if app.edit_entry is None or not app.edit_entry.selection_present():
            raise RuntimeError("單擊格子後沒有選取完整內容。")
        app.clipboard_clear()
        app.clipboard_append("貼上測試名稱")
        app.edit_entry.event_generate("<<Paste>>")
        app.update()
        if list(app.tree.item("20", "values"))[1] != "貼上測試名稱":
            raise RuntimeError("格子貼上內容沒有保存。")
        app.go_back()
        if app.workflow_stage != "match":
            raise RuntimeError("名稱調整沒有返回已建檔勾選。")
        app.go_back()
        if app.workflow_stage != "ocr" or str(app.back_button.cget("state")) != "disabled":
            raise RuntimeError("已建檔勾選沒有返回 OCR 原文確認。")

        delete_workbook_path = output_dir / "刪除項目測試.xlsx"
        create_delete_workbook(delete_workbook_path)
        app.state = WorkflowState(
            raw_xlsx=delete_workbook_path,
            row_count=3,
            invoice_total="600",
        )
        app.workflow_stage = "ocr"
        app.product_match_ready = False
        app.populate_ocr_tree(workflow.load_ocr_confirm_rows(app.state))
        app.tree.selection_set("4")
        original_askyesno = gui_module.messagebox.askyesno
        gui_module.messagebox.askyesno = lambda *_args, **_kwargs: True
        try:
            app.delete_selected_items()
        finally:
            gui_module.messagebox.askyesno = original_askyesno
        remaining_delete_rows = workflow.load_ocr_confirm_rows(
            WorkflowState(raw_xlsx=delete_workbook_path)
        )
        if [row.raw_name for row in remaining_delete_rows] != [
            "刪除測試 1",
            "刪除測試 3",
        ]:
            raise RuntimeError(f"刪除項目沒有同步到 Excel：{remaining_delete_rows}")
        if workflow.read_invoice_total(delete_workbook_path) != "400":
            raise RuntimeError("刪除項目後沒有重算總額為 400。")

        match_delete_path = output_dir / "已建檔勾選刪除項目測試.xlsx"
        create_adjustment_delete_workbook(match_delete_path)
        app.state = WorkflowState(
            raw_xlsx=match_delete_path,
            match_xlsx=match_delete_path,
            row_count=3,
            invoice_total="600",
        )
        app.workflow_stage = "match"
        app.product_match_ready = True
        app.populate_ocr_tree(workflow.load_ocr_confirm_rows(app.state))
        app.tree.selection_set("4")
        original_askyesno = gui_module.messagebox.askyesno
        gui_module.messagebox.askyesno = lambda *_args, **_kwargs: True
        try:
            app.delete_selected_items()
        finally:
            gui_module.messagebox.askyesno = original_askyesno
        remaining_match_rows = workflow.load_ocr_confirm_rows(
            WorkflowState(match_xlsx=match_delete_path)
        )
        if [row.raw_name for row in remaining_match_rows] != [
            "調整刪除測試 1",
            "調整刪除測試 3",
        ]:
            raise RuntimeError(
                f"已建檔勾選刪除沒有同步到 Excel：{remaining_match_rows}"
            )
        if workflow.read_invoice_total(match_delete_path) != "400":
            raise RuntimeError("已建檔勾選刪除後沒有重算總額為 400。")

        adjustment_delete_path = output_dir / "名稱調整刪除項目測試.xlsx"
        create_adjustment_delete_workbook(adjustment_delete_path)
        app.state = WorkflowState(
            match_xlsx=adjustment_delete_path,
            row_count=3,
            invoice_total="600",
        )
        app.workflow_stage = "adjust"
        app.product_match_ready = True
        app.set_adjustment_table()
        app.populate_adjustment_tree(workflow.load_adjustment_rows(app.state))
        app.tree.selection_set("4")
        original_askyesno = gui_module.messagebox.askyesno
        gui_module.messagebox.askyesno = lambda *_args, **_kwargs: True
        try:
            app.delete_selected_items()
        finally:
            gui_module.messagebox.askyesno = original_askyesno
        remaining_adjustment_rows = workflow.load_adjustment_rows(
            WorkflowState(match_xlsx=adjustment_delete_path)
        )
        if [row.name for row in remaining_adjustment_rows] != [
            "調整刪除測試 1",
            "調整刪除測試 3",
        ]:
            raise RuntimeError(
                f"名稱調整刪除沒有同步到 Excel：{remaining_adjustment_rows}"
            )
        if workflow.read_invoice_total(adjustment_delete_path) != "400":
            raise RuntimeError("名稱調整刪除後沒有重算總額為 400。")
        app.withdraw()
    finally:
        app.destroy()

    print(
        json.dumps(
            {
                "ok": True,
                "candidate_count": len(parsed),
                "normal_candidate_count": sum(
                    candidate[3] == "normal" for candidate in parsed
                ),
                "low_candidate_count": sum(
                    candidate[3] == "low" for candidate in parsed
                ),
                "deep_candidate_format": True,
                "legacy_candidate_format_compatible": True,
                "similar_candidates_not_auto_checked": True,
                "candidate_summary_keeps_full_data": True,
                "mouse_wheel_bound": True,
                "manual_existing_code": chosen_code,
                "automatic_existing_code_checked": True,
                "catalog_name_auto_filled": chosen_name,
                "product_csv_path_field": True,
                "detected_vendor_field": True,
                "ocr_completion_reminder": True,
                "ocr_and_product_match_are_separate": True,
                "batch_product_code_paste": True,
                "adjustment_batch_product_code_paste": True,
                "batch_product_name_paste": True,
                "all_ocr_data_columns_batch_paste": True,
                "all_adjustment_data_columns_batch_paste": True,
                "batch_paste_preserves_blank_rows": True,
                "category_code_and_name_columns": True,
                "batch_category_paste": True,
                "embedded_categories": True,
                "category_manager_dialog": True,
                "category_add_delete_persistence": True,
                "adjustment_single_click_cell_selection": True,
                "copy_entire_column": True,
                "copied_names_have_no_blank_rows": True,
                "copy_column_shortcut_and_menu": True,
                "go_back_preserves_adjustment_draft": True,
                "go_back_match_to_ocr": True,
                "delete_item_button": True,
                "delete_item_persisted_to_workbook": True,
                "delete_match_item_persisted": True,
                "delete_adjustment_item_persisted": True,
                "delete_item_total_recalculated": True,
                "workbook": str(workbook_path),
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
