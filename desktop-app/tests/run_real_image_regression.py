from __future__ import annotations

import argparse
import json
import shutil
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


APP_DIR = Path(__file__).resolve().parents[1]
if str(APP_DIR) not in sys.path:
    sys.path.insert(0, str(APP_DIR))

import invoice_workflow as workflow


SUMMARY_NAMES = {"總價格", "總價", "總計", "合計", "小計"}


def normalized_rows(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        worksheet = workbook["進貨明細"] if "進貨明細" in workbook.sheetnames else workbook.active
        header_row = 0
        headers: dict[str, int] = {}
        for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            current = {str(value).strip(): index for index, value in enumerate(row) if value not in (None, "")}
            if "品名" in current and "數量" in current and "進價" in current and "金額" in current:
                header_row = row_number
                headers = current
                break
        if not header_row:
            raise RuntimeError(f"找不到進貨明細表頭：{path}")

        result: list[dict[str, str]] = []
        for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
            name = str(row[headers["品名"]] or "").strip()
            if not name or name in SUMMARY_NAMES:
                continue
            item = {
                "品名": name,
                "數量": str(row[headers["數量"]] or "").strip(),
                "進價": str(row[headers["進價"]] or "").strip(),
                "金額": str(row[headers["金額"]] or "").strip(),
            }
            if "零售價" in headers:
                item["零售價"] = str(row[headers["零售價"]] or "").strip()
            result.append(item)
        return result
    finally:
        workbook.close()


def match_columns(path: Path) -> list[str]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        worksheet = workbook["進貨明細"] if "進貨明細" in workbook.sheetnames else workbook.active
        for row in worksheet.iter_rows(values_only=True):
            values = [str(value).strip() if value is not None else "" for value in row]
            if "品名" in values and "比對狀態" in values:
                return [value for value in values if value]
        return []
    finally:
        workbook.close()


def run(args: argparse.Namespace) -> dict[str, Any]:
    source_image = args.image.resolve()
    expected_raw = args.expected_raw.resolve()
    expected_adjusted = args.expected_adjusted.resolve() if args.expected_adjusted else None
    job_dir = args.output_dir.resolve()
    input_dir = job_dir / "input"
    raw_dir = job_dir / "raw"
    formal_dir = job_dir / "formal"
    for directory in (input_dir, raw_dir, formal_dir):
        directory.mkdir(parents=True, exist_ok=True)

    clean_stem = source_image.stem.replace("_已處理", "")
    test_image = input_dir / f"{clean_stem}{source_image.suffix.lower()}"
    shutil.copy2(source_image, test_image)

    runtime_issues = workflow.validate_runtime()
    if runtime_issues:
        raise RuntimeError("\n".join(runtime_issues))

    state = workflow.WorkflowState(
        image_path=test_image,
        output_dir=raw_dir,
        formal_output_dir=formal_dir,
    )
    ocr_summary = workflow.run_ocr(state)
    expected_rows = normalized_rows(expected_raw)
    actual_rows = normalized_rows(state.raw_xlsx)
    if actual_rows != expected_rows:
        raise RuntimeError(
            "實圖 OCR 與歷史確認結果不同。\n"
            + json.dumps({"expected": expected_rows, "actual": actual_rows}, ensure_ascii=False, indent=2)
        )

    match_summary = workflow.run_match(state)
    required_match_columns = {"比對狀態", "已建檔代號", "已建檔品名", "相似候選"}
    actual_match_columns = set(match_columns(state.match_xlsx))
    if not required_match_columns.issubset(actual_match_columns):
        raise RuntimeError(f"產品比對欄位不完整：{sorted(actual_match_columns)}")

    build_summary: dict[str, Any] | None = None
    if expected_adjusted:
        adjusted_copy = raw_dir / expected_adjusted.name
        shutil.copy2(expected_adjusted, adjusted_copy)
        state.adjusted_xlsx = adjusted_copy
        state.invoice_total = workflow.read_invoice_total(adjusted_copy)
        build_summary = workflow.build_import_files(state)

    return {
        "ok": True,
        "source_image": str(source_image),
        "test_image_after_processing": str(state.processed_image_path or state.image_path),
        "ocr_output": str(state.raw_xlsx),
        "ocr_rows": len(actual_rows),
        "invoice_total": state.invoice_total,
        "ocr_needs_review": state.needs_ocr_review,
        "ocr_issues": state.ocr_issues,
        "page_split_retry": ocr_summary.get("page_split_retry", ""),
        "match_output": str(state.match_xlsx),
        "match_columns_ok": True,
        "formal_build": build_summary,
        "historical_raw_match": True,
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--image", type=Path, required=True)
    parser.add_argument("--expected-raw", type=Path, required=True)
    parser.add_argument("--expected-adjusted", type=Path)
    parser.add_argument("--output-dir", type=Path, required=True)
    args = parser.parse_args()
    try:
        result = run(args)
    except Exception as exc:
        print(json.dumps({"ok": False, "error": str(exc)}, ensure_ascii=False, indent=2))
        return 1
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
